from __future__ import annotations

from datetime import timedelta
from io import BytesIO

import pytest
from django.test import override_settings
from openpyxl import load_workbook

from core.models import AdminAuditLog, Game, ScheduleGridDraft, ScheduleImportBatch, Venue
from core.services.schedule_drafts import (
    export_schedule_draft_xlsx,
    get_or_create_schedule_draft,
    import_schedule_draft_xlsx,
    replace_schedule_draft,
    serialize_schedule_draft,
    validate_schedule_draft,
)
from core.services.schedule_imports_v3 import (
    GRID_START_COLUMN,
    ScheduleImportError,
    confirm_schedule_import,
    generate_schedule_template,
)
from core.tests.test_schedule_imports_v3 import _filled_workbook, _setup

pytestmark = pytest.mark.django_db


def test_new_draft_uses_defaults_without_reading_legacy_grid_configuration():
    setup = _setup()
    draft = get_or_create_schedule_draft(
        actor=setup["actor"], season=setup["season"]
    )
    result = serialize_schedule_draft(draft)

    assert result["version"] == 1
    assert [row["venue_name"] for row in result["columns"]] == [
        "五四东一",
        "五四东二",
        "五四东三",
        "邱德拔",
    ]
    assert result["columns"][-1]["final_only"] is True
    assert result["summary"]["draft_game_count"] == 0


def test_archived_schedule_draft_reads_without_creating_or_mutating():
    setup = _setup()
    draft = get_or_create_schedule_draft(actor=setup["actor"], season=setup["season"])
    setup["season"].status = setup["season"].Status.ARCHIVED
    setup["season"].save(update_fields=["status", "updated_at"])
    before_audits = AdminAuditLog.objects.count()

    loaded = get_or_create_schedule_draft(actor=setup["actor"], season=setup["season"])

    assert loaded.id == draft.id
    assert AdminAuditLog.objects.count() == before_audits
    with pytest.raises(ScheduleImportError, match="已归档赛季只读"):
        replace_schedule_draft(
            actor=setup["actor"],
            season=setup["season"],
            expected_version=draft.version,
            columns=[],
            cells=[],
        )


def test_archived_season_without_draft_does_not_create_one():
    setup = _setup()
    setup["season"].status = setup["season"].Status.ARCHIVED
    setup["season"].save(update_fields=["status", "updated_at"])

    with pytest.raises(ScheduleImportError, match="不能新建草稿"):
        get_or_create_schedule_draft(actor=setup["actor"], season=setup["season"])

    assert not ScheduleGridDraft.objects.filter(season=setup["season"]).exists()


def test_xlsx_headers_replace_draft_and_allow_free_venue_text():
    setup = _setup()
    draft = get_or_create_schedule_draft(
        actor=setup["actor"], season=setup["season"]
    )
    workbook = load_workbook(BytesIO(_filled_workbook(setup)))
    workbook["赛程网格"].cell(6, GRID_START_COLUMN, "临时馆")
    output = BytesIO()
    workbook.save(output)

    imported = import_schedule_draft_xlsx(
        actor=setup["actor"],
        season=setup["season"],
        expected_version=draft.version,
        content=output.getvalue(),
        source_name="自定义赛程.xlsx",
    )
    result = serialize_schedule_draft(imported)

    assert result["source_name"] == "自定义赛程.xlsx"
    assert result["summary"]["draft_game_count"] == 7
    assert result["columns"][0]["venue_name"] == "临时馆"
    assert not Venue.objects.filter(season=setup["season"], name="临时馆").exists()


def test_structural_xlsx_error_does_not_replace_existing_draft():
    setup = _setup()
    draft = get_or_create_schedule_draft(
        actor=setup["actor"], season=setup["season"]
    )
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    workbook["赛程网格"].cell(5, GRID_START_COLUMN, "13:37")
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ScheduleImportError) as error:
        import_schedule_draft_xlsx(
            actor=setup["actor"],
            season=setup["season"],
            expected_version=draft.version,
            content=output.getvalue(),
            source_name="坏表头.xlsx",
        )
    assert error.value.code == "UNKNOWN_GRID_PERIOD"
    draft.refresh_from_db()
    assert draft.version == 1
    assert draft.cells.count() == 0


def test_online_draft_validation_preserves_game_lock_and_clears_cells_on_confirm(
    tmp_path,
):
    setup = _setup()
    draft = get_or_create_schedule_draft(
        actor=setup["actor"], season=setup["season"]
    )
    draft = import_schedule_draft_xlsx(
        actor=setup["actor"],
        season=setup["season"],
        expected_version=draft.version,
        content=_filled_workbook(setup),
        source_name="完整赛程.xlsx",
    )
    serialized = serialize_schedule_draft(draft)
    cells = [dict(item) for item in serialized["cells"]]
    cells[0]["leader_adjustable"] = False
    draft = replace_schedule_draft(
        actor=setup["actor"],
        season=setup["season"],
        expected_version=draft.version,
        columns=[dict(item) for item in serialized["columns"]],
        cells=cells,
    )

    with override_settings(MEDIA_ROOT=tmp_path):
        batch = validate_schedule_draft(
            actor=setup["actor"],
            season=setup["season"],
            expected_version=draft.version,
        )
        assert batch.source_kind == ScheduleImportBatch.SourceKind.ONLINE_DRAFT
        assert batch.summary["error_count"] == 0
        confirmed = confirm_schedule_import(
            actor=setup["actor"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )

    assert confirmed.status == ScheduleImportBatch.Status.CONFIRMED
    assert Game.objects.filter(season=setup["season"], leader_adjustable=False).count() == 1
    draft.refresh_from_db()
    assert draft.cells.count() == 0
    assert draft.version == batch.source_draft_version + 1


def test_draft_update_uses_optimistic_version_and_xlsx_omits_lock_policy():
    setup = _setup()
    draft = get_or_create_schedule_draft(
        actor=setup["actor"], season=setup["season"]
    )
    draft = import_schedule_draft_xlsx(
        actor=setup["actor"],
        season=setup["season"],
        expected_version=draft.version,
        content=_filled_workbook(setup),
        source_name="完整赛程.xlsx",
    )
    result = serialize_schedule_draft(draft)
    cells = [dict(item) for item in result["cells"]]
    cells[0]["leader_adjustable"] = False
    updated = replace_schedule_draft(
        actor=setup["actor"],
        season=setup["season"],
        expected_version=draft.version,
        columns=[dict(item) for item in result["columns"]],
        cells=cells,
    )
    with pytest.raises(ScheduleImportError) as error:
        replace_schedule_draft(
            actor=setup["actor"],
            season=setup["season"],
            expected_version=draft.version,
            columns=[dict(item) for item in result["columns"]],
            cells=cells,
        )
    assert error.value.code == "DRAFT_VERSION_CONFLICT"

    workbook = load_workbook(
        BytesIO(export_schedule_draft_xlsx(actor=setup["actor"], season=setup["season"]))
    )
    values = [
        str(cell.value)
        for row in workbook["填写说明"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert updated.cells.filter(leader_adjustable=False).count() == 1
    assert not any("leader_adjustable" in value or "不可调" in value for value in values)


def test_online_draft_keeps_special_dates_outside_planning_range():
    setup = _setup()
    draft = get_or_create_schedule_draft(actor=setup["actor"], season=setup["season"])
    result = serialize_schedule_draft(draft)
    target_date = setup["season"].ends_on + timedelta(days=2)

    updated = replace_schedule_draft(
        actor=setup["actor"],
        season=setup["season"],
        expected_version=draft.version,
        columns=[dict(item) for item in result["columns"]],
        cells=[
            {
                "column_id": result["columns"][0]["id"],
                "date": target_date.isoformat(),
                "matchup": "A1vsA2",
                "leader_adjustable": True,
            }
        ],
    )
    serialized = serialize_schedule_draft(updated)
    workbook = load_workbook(
        BytesIO(export_schedule_draft_xlsx(actor=setup["actor"], season=setup["season"]))
    )

    assert serialized["dates"][-1]["date"] == target_date.isoformat()
    assert any(
        cell.value == target_date
        or (hasattr(cell.value, "date") and cell.value.date() == target_date)
        for cell in workbook["赛程网格"]["A"]
    )
