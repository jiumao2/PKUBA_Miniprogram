from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from django.test import override_settings
from openpyxl import load_workbook

from core.models import (
    Account,
    AdminAuditLog,
    ImportIssue,
    PeriodCapacity,
    ScheduleImportBatch,
)
from core.services.schedule_imports import (
    GRID_HEADER_ROW,
    GRID_START_ROW,
    VENUE_START_COLUMN,
    ScheduleImportError,
    confirm_schedule_import,
    generate_schedule_template,
    validate_schedule_upload,
)
from core.tests.factories import reschedule_setup

pytestmark = [
    pytest.mark.django_db,
    pytest.mark.skip(reason="V1 签名赛程协议已退役；当前服务器仅接受 V3。"),
]


def import_setup():
    setup = reschedule_setup(capacity=3)
    for weekday in range(7):
        PeriodCapacity.objects.update_or_create(
            season=setup["season"],
            weekday=weekday,
            period=setup["period"],
            defaults={"capacity": 3},
        )
    superadmin = Account.objects.create_user(
        username="schedule-superadmin",
        password="test-password",
        role=Account.Role.SUPERADMIN,
    )
    return setup, superadmin


def workbook_bytes(workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def game_cells(workbook) -> dict[str, str]:
    sheet = workbook["赛程网格"]
    cells: dict[str, str] = {}
    for row in range(GRID_START_ROW, sheet.max_row + 1):
        for column in range(VENUE_START_COLUMN, sheet.max_column + 1):
            value = sheet.cell(row, column).value
            if isinstance(value, str) and value.startswith("RS-G"):
                cells[value] = sheet.cell(row, column).coordinate
    return cells


def grid_cell(workbook, target_date: date, period_code: str, venue_code: str):
    sheet = workbook["赛程网格"]
    venue_column = next(
        column
        for column in range(VENUE_START_COLUMN, sheet.max_column + 1)
        if sheet.cell(GRID_HEADER_ROW - 1, column).value == venue_code
    )
    target_row = next(
        row
        for row in range(GRID_START_ROW, sheet.max_row + 1)
        if sheet.cell(row, 1).value.date() == target_date
        and sheet.cell(row, 3).value == period_code
    )
    return sheet.cell(target_row, venue_column)


def validate_bytes(*, setup, actor, content, media_root):
    with override_settings(MEDIA_ROOT=media_root):
        return validate_schedule_upload(
            actor=actor,
            season=setup["season"],
            content=content,
            source_name="schedule.xlsx",
        )


def test_generated_template_round_trips_to_validated_batch(tmp_path):
    setup, superadmin = import_setup()
    content = generate_schedule_template(setup["season"])

    batch = validate_bytes(
        setup=setup,
        actor=superadmin,
        content=content,
        media_root=tmp_path,
    )

    assert batch.status == ScheduleImportBatch.Status.VALIDATED
    assert batch.summary["error_count"] == 0
    assert batch.summary["expected"] == 2
    assert batch.summary["actual_unique"] == 2
    assert batch.summary["modified"] == []
    assert batch.issues.filter(severity=ImportIssue.Severity.WARNING).count() == 2


def test_non_superadmin_cannot_generate_import_batch(tmp_path):
    setup, _ = import_setup()
    content = generate_schedule_template(setup["season"])

    with pytest.raises(ScheduleImportError, match="超级管理员") as denied:
        validate_bytes(
            setup=setup,
            actor=setup["admin"],
            content=content,
            media_root=tmp_path,
        )

    assert denied.value.code == "PERMISSION_DENIED"
    assert ScheduleImportBatch.objects.count() == 0


def test_formula_in_user_area_is_server_side_error(tmp_path):
    setup, superadmin = import_setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    cells = game_cells(workbook)
    first_code = setup["games"][0].code
    first_cell = workbook["赛程网格"][cells[first_code]]
    duplicate_target = grid_cell(
        workbook,
        setup["target_date"],
        setup["period"].code,
        setup["venues"][2].code,
    )
    duplicate_target.value = first_code
    first_cell.value = "=\"RS-G001\""

    batch = validate_bytes(
        setup=setup,
        actor=superadmin,
        content=workbook_bytes(workbook),
        media_root=tmp_path,
    )

    codes = set(batch.issues.values_list("code", flat=True))
    assert "FORMULA_FORBIDDEN" in codes
    assert "MISSING_GAME" not in codes
    assert batch.summary["error_count"] >= 1


def test_duplicate_missing_and_unknown_codes_are_reported_together(tmp_path):
    setup, superadmin = import_setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    cells = game_cells(workbook)
    first_code = setup["games"][0].code
    second_code = setup["games"][1].code
    duplicate_target = grid_cell(
        workbook,
        setup["target_date"],
        setup["period"].code,
        setup["venues"][2].code,
    )
    duplicate_target.value = first_code
    workbook["赛程网格"][cells[second_code]].value = "NOT-A-GAME"

    batch = validate_bytes(
        setup=setup,
        actor=superadmin,
        content=workbook_bytes(workbook),
        media_root=tmp_path,
    )

    codes = set(batch.issues.values_list("code", flat=True))
    assert {"DUPLICATE_GAME", "MISSING_GAME", "UNKNOWN_GAME_CODE"} <= codes


def test_template_identity_tampering_requires_fresh_download(tmp_path):
    setup, superadmin = import_setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    workbook["赛程网格"]["D4"] = "another-season"

    batch = validate_bytes(
        setup=setup,
        actor=superadmin,
        content=workbook_bytes(workbook),
        media_root=tmp_path,
    )

    assert batch.issues.filter(code="TEMPLATE_SIGNATURE_INVALID").exists()


def test_external_hyperlink_is_rejected_before_workbook_parsing(tmp_path):
    setup, superadmin = import_setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    workbook["填写说明"]["B4"].hyperlink = "https://example.invalid/collect"

    batch = validate_bytes(
        setup=setup,
        actor=superadmin,
        content=workbook_bytes(workbook),
        media_root=tmp_path,
    )

    assert batch.issues.filter(code="EXTERNAL_LINK_FORBIDDEN").exists()


def test_capacity_is_recomputed_with_proposed_games(tmp_path):
    setup, superadmin = import_setup()
    PeriodCapacity.objects.filter(
        season=setup["season"],
        weekday=setup["target_date"].weekday(),
        period=setup["period"],
    ).update(capacity=1)
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    cells = game_cells(workbook)
    for coordinate in cells.values():
        workbook["赛程网格"][coordinate].value = None
    for index, game in enumerate(setup["games"]):
        grid_cell(
            workbook,
            setup["target_date"],
            setup["period"].code,
            setup["venues"][index].code,
        ).value = game.code

    batch = validate_bytes(
        setup=setup,
        actor=superadmin,
        content=workbook_bytes(workbook),
        media_root=tmp_path,
    )

    assert batch.issues.filter(code="CAPACITY_EXCEEDED").exists()
    assert batch.summary["error_count"] >= 1


def test_confirm_swap_preserves_game_ids_and_records_policy(tmp_path):
    setup, superadmin = import_setup()
    first, second = setup["games"]
    original = {
        first.code: (first.id, first.date, first.period_id, first.venue_id),
        second.code: (second.id, second.date, second.period_id, second.venue_id),
    }
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    cells = game_cells(workbook)
    first_cell = workbook["赛程网格"][cells[first.code]]
    second_cell = workbook["赛程网格"][cells[second.code]]
    first_cell.value, second_cell.value = second.code, first.code

    with override_settings(MEDIA_ROOT=tmp_path):
        batch = validate_schedule_upload(
            actor=superadmin,
            season=setup["season"],
            content=workbook_bytes(workbook),
            source_name="schedule.xlsx",
        )
        assert batch.summary["error_count"] == 0
        confirmed = confirm_schedule_import(
            actor=superadmin,
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
            leader_adjustable_by_game={first.code: False, second.code: True},
        )

    first.refresh_from_db()
    second.refresh_from_db()
    assert confirmed.status == ScheduleImportBatch.Status.CONFIRMED
    assert first.id == original[first.code][0]
    assert second.id == original[second.code][0]
    assert (first.date, first.period_id, first.venue_id) == original[second.code][1:]
    assert (second.date, second.period_id, second.venue_id) == original[first.code][1:]
    assert first.leader_adjustable is False
    assert second.leader_adjustable is True
    assert AdminAuditLog.objects.filter(
        action="SCHEDULE_IMPORT_CONFIRMED",
        object_id=batch.id,
    ).exists()


def test_confirming_unchanged_schedule_does_not_bump_game_or_season_versions(tmp_path):
    setup, superadmin = import_setup()
    content = generate_schedule_template(setup["season"])
    original_season_version = setup["season"].version
    original_game_versions = {game.code: game.version for game in setup["games"]}

    with override_settings(MEDIA_ROOT=tmp_path):
        batch = validate_schedule_upload(
            actor=superadmin,
            season=setup["season"],
            content=content,
            source_name="schedule.xlsx",
        )
        confirm_schedule_import(
            actor=superadmin,
            batch_id=batch.id,
            expected_season_version=original_season_version,
            leader_adjustable_by_game={game.code: True for game in setup["games"]},
        )

    setup["season"].refresh_from_db()
    assert setup["season"].version == original_season_version
    for game in setup["games"]:
        game.refresh_from_db()
        assert game.version == original_game_versions[game.code]
