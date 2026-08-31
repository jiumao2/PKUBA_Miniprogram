from __future__ import annotations

from datetime import time, timedelta
from io import BytesIO

import pytest
from django.test import override_settings
from django.utils import timezone
from openpyxl import load_workbook

from core.models import (
    Account,
    CompetitionGroup,
    Division,
    Game,
    ParticipantSlot,
    Period,
    PeriodCapacity,
    RescheduleRequest,
    ScheduleImportBatch,
    ScheduleSlotFamily,
    Season,
    SlotReservation,
    Team,
    Venue,
)
from core.services.schedule_imports_v3 import (
    GRID_START_COLUMN,
    GRID_START_ROW,
    SLOT_START_ROW,
    TEMPLATE_VERSION,
    confirm_schedule_import,
    generate_schedule_template,
    reset_schedule_imports,
    schedule_import_readiness,
    schedule_import_reset_preview,
    validate_schedule_upload,
)

pytestmark = pytest.mark.django_db


def _setup():
    starts_on = timezone.localdate() + timedelta(days=30)
    season = Season.objects.create(
        name="V3 测试赛季",
        competition_type=Season.CompetitionType.PKU_CUP,
        year=starts_on.year,
        status=Season.Status.SETUP,
        starts_on=starts_on,
        ends_on=starts_on + timedelta(days=4),
    )
    men = Division.objects.create(
        season=season,
        code="men-a",
        name="男甲",
        gender=Division.Gender.MEN,
        sort_order=1,
    )
    women = Division.objects.create(
        season=season,
        code="women-a",
        name="女甲",
        gender=Division.Gender.WOMEN,
        sort_order=2,
    )
    for index in range(1, 5):
        Team.objects.create(season=season, division=men, name=f"男队 {index}")
    for index in range(1, 3):
        Team.objects.create(season=season, division=women, name=f"女队 {index}")
    regular = Period.objects.create(
        season=season,
        code="p1",
        name="第一时段",
        start_time=time(12, 50),
        sort_order=1,
    )
    final = Period.objects.create(
        season=season,
        code="p4",
        name="决赛早场",
        start_time=time(18, 30),
        sort_order=2,
    )
    for day_type in PeriodCapacity.DayType.values:
        PeriodCapacity.objects.create(
            season=season,
            period=regular,
            day_type=day_type,
            capacity=2,
        )
        PeriodCapacity.objects.create(
            season=season,
            period=final,
            day_type=day_type,
            capacity=1,
        )
    Venue.objects.create(
        season=season, name="五四东一", sort_order=1, active=True
    )
    Venue.objects.create(
        season=season, name="五四东二", sort_order=2, active=True
    )
    Venue.objects.create(season=season, name="邱德拔", sort_order=3, active=True)
    ScheduleSlotFamily.objects.create(
        season=season,
        division=men,
        stage=Game.Stage.GROUP,
        prefix="A",
        slot_count=4,
        sort_order=1,
    )
    ScheduleSlotFamily.objects.create(
        season=season,
        division=women,
        stage=Game.Stage.FINAL,
        prefix="A",
        slot_count=2,
        sort_order=2,
    )
    actor = Account.objects.create_user(
        username="v3-superadmin",
        password="test-password",
        role=Account.Role.SUPERADMIN,
    )
    return {"season": season, "actor": actor, "men": men, "women": women}


def _filled_workbook(setup) -> bytes:
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    grid = workbook["赛程网格"]
    # V3.3 的列头属于工作簿本身：测试把默认网格改成两列常规场地和一列决赛场地。
    grid.cell(5, GRID_START_COLUMN + 2, "18:30")
    grid.cell(6, GRID_START_COLUMN + 2, "邱德拔（仅决赛）")
    grid.cell(5, GRID_START_COLUMN + 3).value = None
    grid.cell(6, GRID_START_COLUMN + 3).value = None
    values = [
        (0, GRID_START_COLUMN, "A1vsA2"),
        (0, GRID_START_COLUMN + 1, "A3vsA4"),
        (1, GRID_START_COLUMN, "A1vsA3"),
        (1, GRID_START_COLUMN + 1, "A2vsA4"),
        (2, GRID_START_COLUMN, "A1vsA4"),
        (2, GRID_START_COLUMN + 1, "A2vsA3"),
        (3, GRID_START_COLUMN + 2, "A1vsA2（女）"),
    ]
    for day_offset, column, value in values:
        grid.cell(GRID_START_ROW + day_offset, column, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _upload(setup, tmp_path, content: bytes):
    with override_settings(MEDIA_ROOT=tmp_path):
        return validate_schedule_upload(
            actor=setup["actor"],
            season=setup["season"],
            content=content,
            source_name="schedule-v3.xlsx",
        )


def test_active_reservation_venue_is_hidden_from_import_conflict(tmp_path):
    setup = _setup()
    season = setup["season"]
    period = season.periods.get(code="p1")
    reserved_venue = season.venues.get(name="五四东一")
    original_venue = season.venues.get(name="五四东二")
    teams = list(setup["men"].teams.order_by("name")[:2])
    game = Game.objects.create(
        season=season,
        division=setup["men"],
        code="ACTIVE-RESCHEDULE",
        date=season.ends_on,
        period=period,
        start_time=period.start_time,
        venue_name=original_venue.name,
        home_team=teams[0],
        away_team=teams[1],
    )
    reservation = SlotReservation.objects.create(
        season=season,
        date=season.starts_on,
        period=period,
        venue=reserved_venue,
        venue_name=reserved_venue.name,
    )
    request = RescheduleRequest.objects.create(
        game=game,
        requester_team=teams[0],
        requester=setup["actor"],
        request_type=RescheduleRequest.RequestType.SAME_WEEK,
        target_date=reservation.date,
        target_period=period,
        target_start_time=period.start_time,
        target_venue_name=reserved_venue.name,
        reservation=reservation,
        original_game_snapshot={
            "date": game.date.isoformat(),
            "start_time": game.start_time.strftime("%H:%M"),
            "venue_name": game.venue_name,
        },
        game_version_at_submit=game.version,
        submit_deadline=timezone.now() + timedelta(days=1),
        confirmation_deadline=timezone.now() + timedelta(days=2),
    )
    game.active_reschedule_request = request
    game.save(update_fields=["active_reschedule_request", "updated_at"])

    batch = _upload(setup, tmp_path, _filled_workbook(setup))
    issue = batch.issues.get(code="VENUE_OCCUPIED")

    assert reserved_venue.name not in issue.message
    assert issue.cell == ""
    assert issue.context == {
        "date": season.starts_on.isoformat(),
        "period_code": "P1",
        "venue_hidden_until_reschedule_effective": True,
    }
    assert str(reservation.id) not in str(issue.context)


def test_dynamic_template_uses_three_sheets_daily_rows_and_gender_namespace():
    setup = _setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))

    assert workbook.sheetnames == ["填写说明", "签位定义", "赛程网格"]
    assert workbook["填写说明"]["B4"].value == TEMPLATE_VERSION
    for row in range(4, 13):
        for column in range(1, 3):
            cell = workbook["填写说明"].cell(row, column)
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"
            assert cell.border.bottom.style == "thin"
    summary_end = 15 + setup["season"].divisions.count()
    for row in range(15, summary_end + 1):
        for column in range(1, 5):
            cell = workbook["填写说明"].cell(row, column)
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"
            assert cell.border.bottom.style == "thin"
    instruction_values = {
        str(cell.value)
        for row in workbook["填写说明"].iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert not any("可调政策" in value or "调赛政策" in value for value in instruction_values)
    assert workbook["签位定义"].max_row == SLOT_START_ROW + 6 - 1
    assert workbook["赛程网格"].max_row == GRID_START_ROW + 5 - 1
    assert workbook["赛程网格"]["B5"].value == "星期"
    assert workbook["赛程网格"]["B6"].value == "自动生成"
    assert workbook["赛程网格"]["B7"].value.startswith("周")
    assert [
        workbook["赛程网格"].cell(5, column).value
        for column in range(GRID_START_COLUMN, GRID_START_COLUMN + 4)
    ] == [
        "12:50",
        "12:50",
        "12:50",
        "18:30",
    ]
    assert [
        workbook["赛程网格"].cell(6, column).value
        for column in range(GRID_START_COLUMN, GRID_START_COLUMN + 4)
    ] == [
        "五四东一",
        "五四东二",
        "五四东三",
        "邱德拔（仅决赛）",
    ]
    for row in range(5, GRID_START_ROW + 5):
        for column in range(1, GRID_START_COLUMN + 4):
            cell = workbook["赛程网格"].cell(row, column)
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"
            assert cell.border.bottom.style == "thin"


def test_valid_v3_upload_stages_then_atomically_confirms_and_resets(tmp_path):
    setup = _setup()
    batch = _upload(setup, tmp_path, _filled_workbook(setup))

    assert batch.status == ScheduleImportBatch.Status.VALIDATED
    assert batch.summary["error_count"] == 0
    assert batch.summary["warning_count"] == 0
    assert batch.summary["new_group_count"] == 1
    assert batch.summary["new_slot_count"] == 6
    assert batch.summary["new_game_count"] == 7
    assert batch.summary["covered_game_count"] == batch.summary["prerequisites"][
        "expected_game_count"
    ]
    assert not Game.objects.filter(season=setup["season"]).exists()

    with override_settings(MEDIA_ROOT=tmp_path):
        confirmed = confirm_schedule_import(
            actor=setup["actor"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )
    assert confirmed.status == ScheduleImportBatch.Status.CONFIRMED
    assert Game.objects.filter(season=setup["season"]).count() == 7
    assert CompetitionGroup.objects.filter(division=setup["men"], code="A").exists()
    assert ParticipantSlot.objects.filter(division__season=setup["season"]).count() == 6
    assert not Game.objects.filter(season=setup["season"], leader_adjustable=False).exists()
    women_final = Game.objects.get(division=setup["women"])
    assert women_final.stage == Game.Stage.FINAL
    assert women_final.venue_name == "邱德拔"
    assert women_final.start_time == time(18, 30)

    setup["season"].refresh_from_db()
    preview = schedule_import_reset_preview(
        actor=setup["actor"], season=setup["season"], now=timezone.now()
    )
    assert preview["eligible"] is True
    result = reset_schedule_imports(
        actor=setup["actor"],
        season_id=setup["season"].id,
        expected_season_version=setup["season"].version,
        season_name=setup["season"].name,
        now=timezone.now(),
    )
    assert result["game_count"] == 7
    assert not Game.objects.filter(season=setup["season"]).exists()
    assert ScheduleSlotFamily.objects.filter(season=setup["season"]).count() == 2


def test_published_season_can_download_template_but_cannot_import(tmp_path):
    setup = _setup()
    setup["season"].status = Season.Status.PUBLISHED
    setup["season"].save()

    readiness = schedule_import_readiness(setup["season"])
    assert readiness["template_ready"] is True
    assert readiness["template_blockers"] == []
    assert readiness["ready"] is False
    assert {item["code"] for item in readiness["blockers"]} == {"SEASON_NOT_SETUP"}

    content = _filled_workbook(setup)
    batch = _upload(setup, tmp_path, content)
    assert batch.summary["error_count"] > 0
    assert batch.issues.filter(code="SEASON_NOT_SETUP").exists()


def test_reverse_duplicate_missing_pair_and_final_only_misuse_are_blocking(tmp_path):
    setup = _setup()
    workbook = load_workbook(BytesIO(_filled_workbook(setup)))
    grid = workbook["赛程网格"]
    grid.cell(GRID_START_ROW + 1, GRID_START_COLUMN, "A2vsA1")
    grid.cell(GRID_START_ROW + 4, GRID_START_COLUMN + 2, "A1vsA4")
    output = BytesIO()
    workbook.save(output)

    batch = _upload(setup, tmp_path, output.getvalue())
    codes = {issue.code for issue in batch.issues.all()}
    assert "DUPLICATE_MATCHUP" in codes
    assert "MISSING_ROUND_ROBIN_MATCHUPS" in codes
    assert "FINAL_ONLY_COLUMN" in codes


def test_second_sheet_is_informational_and_does_not_change_configured_slots(tmp_path):
    setup = _setup()
    workbook = load_workbook(BytesIO(_filled_workbook(setup)))
    slots = workbook["签位定义"]
    slots.cell(5, 1, "管理员可以修改这里")
    slots.cell(SLOT_START_ROW, 1, "不存在的组别")
    slots.cell(SLOT_START_ROW, 2, "不存在的阶段")
    slots.cell(SLOT_START_ROW, 3, "Z999")
    slots.delete_rows(SLOT_START_ROW + 1, slots.max_row - SLOT_START_ROW)
    output = BytesIO()
    workbook.save(output)

    batch = _upload(setup, tmp_path, output.getvalue())
    assert batch.summary["error_count"] == 0
    assert batch.summary["new_slot_count"] == 6
    assert batch.summary["new_game_count"] == 7


def test_v2_workbook_is_not_accepted(tmp_path):
    setup = _setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    workbook.create_sheet("比赛清单", 2)
    output = BytesIO()
    workbook.save(output)

    batch = _upload(setup, tmp_path, output.getvalue())
    assert batch.summary["error_count"] == 1
    assert batch.issues.get().code == "SHEET_STRUCTURE_CHANGED"
