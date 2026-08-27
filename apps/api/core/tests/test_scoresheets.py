from __future__ import annotations

import copy
import hashlib
import io
import json
from datetime import datetime, timedelta

import pytest
from django.conf import settings as django_settings
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, override_settings
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image
from pydantic import ValidationError as PydanticValidationError
from pypdf import PdfReader

from core.models import (
    Account,
    AdminAuditLog,
    ApiIdempotencyRecord,
    Game,
    GameMediaAsset,
    GamePlayerStat,
    GameScoresheet,
    GameTeamStat,
    RosterPlayer,
    ScoresheetEditLease,
    ScoresheetPublication,
    ScoresheetRecognitionRun,
)
from core.scoresheet_schema_v2 import ScoresheetDocumentError, merge_recognition_result
from core.scoresheet_v2.models import FoulEntry, PeriodScore, ScoreEvent, ScoresheetDocument
from core.scoresheet_v2.recognition import (
    QWEN_DATA_URI_MAX_BYTES,
    RecognitionImageError,
    _data_url_size,
    _encode_jpeg,
    prepare_image,
)
from core.scoresheet_v2.renderer import build_scene
from core.scoresheet_v2.scoring import derive_score_events
from core.scoresheet_v2.validation import validate_document as validate_v2_document
from core.services.game_media import replace_game_media, upload_game_media
from core.services.scoresheet_recognition import (
    RecognitionAttemptError,
    _renew_worker_lease,
    claim_next_run,
    execute_claim,
    run_once,
)
from core.services.scoresheet_renderer import render_scoresheet_pdf
from core.services.scoresheets import (
    ScoresheetError,
    _build_stats,
    acknowledge_warnings,
    acquire_edit_lease,
    apply_recognition_regions,
    force_takeover_edit_lease,
    publish_scoresheet,
    release_edit_lease,
    retry_recognition,
    review_region,
    save_draft_changes,
    sync_scoresheet,
    validate_scoresheet,
)
from core.services.wechat import issue_session
from core.tests.factories import reschedule_setup

pytestmark = pytest.mark.django_db(transaction=True)


@pytest.fixture(autouse=True)
def _disable_live_qwen_credentials(settings):
    settings.QWEN_API_KEY = ""


def image_file(name: str = "scoresheet.jpg") -> SimpleUploadedFile:
    content = io.BytesIO()
    Image.new("RGB", (800, 1130), color=(246, 244, 239)).save(
        content, format="JPEG", quality=90
    )
    return SimpleUploadedFile(name, content.getvalue(), content_type="image/jpeg")


def create_scoresheet():
    setup = reschedule_setup()
    setup["ordinary_admin"] = setup["admin"]
    setup["admin"] = setup["superadmin"]
    game = setup["games"][0]
    player_rows = {}
    for side, team, prefix in (
        ("A", game.home_team, "甲队"),
        ("B", game.away_team, "乙队"),
    ):
        created = [
            RosterPlayer.objects.create(
                team=team,
                name=f"{prefix}{index}号",
                jersey_number=str(index + 3),
            )
            for index in range(1, 6)
        ]
        player_rows[side] = created[0]
    asset = upload_game_media(
        actor=setup["superadmin"],
        game=game,
        kind=GameMediaAsset.Kind.SCORESHEET,
        scoresheet_complete_confirmed=True,
        uploaded_file=image_file(),
    )
    return setup, game, player_rows, asset, GameScoresheet.objects.get(game=game)


def valid_document(scoresheet: GameScoresheet) -> dict[str, object]:
    document = copy.deepcopy(scoresheet.draft)
    player_a = document["teams"][0]["players"][0]
    player_b = document["teams"][1]["players"][0]
    for player in document["teams"][0]["players"][:5]:
        player["participation"] = "starter"
    for player in document["teams"][1]["players"][:5]:
        player["participation"] = "starter"
    document["score_events"] = [
        {
            "sequence": 1,
            "team": "A",
            "period": 1,
            "points": 2,
            "cumulative_score": 2,
            "scorer_jersey": player_a["jersey_number"],
            "mark": "diagonal",
            "scorer_circled": False,
            "boundary": "none",
            "ink_role": "q1_q3",
        },
        {
            "sequence": 2,
            "team": "B",
            "period": 1,
            "points": 1,
            "cumulative_score": 1,
            "scorer_jersey": player_b["jersey_number"],
            "mark": "filled_dot",
            "scorer_circled": False,
            "boundary": "game_end",
            "ink_role": "q1_q3",
        },
    ]
    document["stated_period_scores"] = [
        {"period": 1, "team_a": 2, "team_b": 1},
        {"period": 2, "team_a": 0, "team_b": 0},
        {"period": 3, "team_a": 0, "team_b": 0},
        {"period": 4, "team_a": 0, "team_b": 0},
    ]
    document["final_score"] = {
        "team_a": 2,
        "team_b": 1,
        "winner_name": document["teams"][0]["name"],
        "ended_at": "14:20",
    }
    for official in document["officials"]:
        if official["role"] == "scorer":
            official.update({"name": "记录员", "signature": "present"})
        elif official["role"] == "timer":
            official.update({"name": "计时员", "signature": "present"})
        elif official["role"] == "crew_chief":
            official["signature"] = "present"
    return document


def obtain_lease(
    scoresheet: GameScoresheet,
    actor: Account,
    client_id: str = "web-1",
    *,
    archived_correction_confirmed: bool = False,
) -> str:
    _, token, read_only, _reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=actor,
        client_id=client_id,
        surface=ScoresheetEditLease.Surface.WEB,
        archived_correction_confirmed=archived_correction_confirmed,
    )
    assert read_only is False
    assert token
    return token


def test_combined_overtime_period_contract_rejects_period_six():
    assert FoulEntry(slot=1, code="P", period=5).period == 5
    assert PeriodScore(period=5, team_a=2, team_b=1).period == 5
    assert ScoreEvent(
        sequence=1,
        team="A",
        period=5,
        points=2,
        cumulative_score=2,
        scorer_jersey="7",
    ).period == 5

    with pytest.raises(PydanticValidationError):
        FoulEntry(slot=1, code="P", period=6)
    with pytest.raises(PydanticValidationError):
        PeriodScore(period=6, team_a=2, team_b=1)
    with pytest.raises(PydanticValidationError):
        ScoreEvent(
            sequence=1,
            team="A",
            period=6,
            points=2,
            cumulative_score=2,
            scorer_jersey="7",
        )


def test_fixed_score_cell_validation_reports_duplicates_crossings_and_missing_boundaries(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        _, _, _, _, scoresheet = create_scoresheet()
    base = valid_document(scoresheet)

    duplicate = copy.deepcopy(base)
    duplicate["score_events"].append({**duplicate["score_events"][0], "sequence": 3})
    rule_profiles_path = (
        django_settings.BASE_DIR / "core" / "assets" / "scoresheet" / "rule_profiles.json"
    )
    duplicate_report = validate_v2_document(
        ScoresheetDocument.model_validate(duplicate), rule_profiles_path
    )
    assert "DUPLICATE_SCORE_CELL" in {issue.code for issue in duplicate_report.issues}

    crossed = copy.deepcopy(base)
    crossed["score_events"][0]["cumulative_score"] = 3
    crossed["score_events"][0]["points"] = 3
    crossed_report = validate_v2_document(
        ScoresheetDocument.model_validate(crossed), rule_profiles_path
    )
    crossed_codes = {issue.code for issue in crossed_report.issues}
    assert "SCORE_EVENT_CROSSES_PERIOD_BOUNDARY" in crossed_codes
    assert "PERIOD_BOUNDARY_WITHOUT_EVENT" in crossed_codes


def test_fixed_score_cells_derive_combined_overtime_without_an_intermediate_boundary(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        _, _, _, _, scoresheet = create_scoresheet()
    draft = valid_document(scoresheet)
    draft["stated_period_scores"].append({"period": 5, "team_a": 2, "team_b": 1})
    jersey_a = draft["teams"][0]["players"][0]["jersey_number"]
    jersey_b = draft["teams"][1]["players"][0]["jersey_number"]
    draft["score_events"].extend(
        [
            {
                "sequence": 3,
                "team": "A",
                "period": 1,
                "points": 1,
                "cumulative_score": 3,
                "scorer_jersey": jersey_a,
                "mark": "filled_dot",
                "scorer_circled": False,
                "boundary": "period_end",
                "ink_role": "neutral",
            },
            {
                "sequence": 4,
                "team": "A",
                "period": 1,
                "points": 1,
                "cumulative_score": 4,
                "scorer_jersey": jersey_a,
                "mark": "filled_dot",
                "scorer_circled": False,
                "boundary": "none",
                "ink_role": "neutral",
            },
            {
                "sequence": 5,
                "team": "B",
                "period": 1,
                "points": 1,
                "cumulative_score": 2,
                "scorer_jersey": jersey_b,
                "mark": "filled_dot",
                "scorer_circled": False,
                "boundary": "none",
                "ink_role": "neutral",
            },
        ]
    )
    draft["final_score"].update(team_a=4, team_b=2)

    document = derive_score_events(ScoresheetDocument.model_validate(draft))
    overtime_a = [
        event
        for event in document.score_events
        if event.team.value == "A" and event.cumulative_score >= 3
    ]

    assert [event.period for event in overtime_a] == [5, 5]
    assert [event.boundary.value for event in overtime_a] == ["none", "game_end"]
    final_b = next(
        event
        for event in document.score_events
        if event.team.value == "B" and event.cumulative_score == 2
    )
    assert final_b.boundary.value == "game_end"


def test_renderer_recomputes_forged_score_metadata_before_drawing(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        _, _, _, _, scoresheet = create_scoresheet()
    draft = valid_document(scoresheet)
    first_a = next(event for event in draft["score_events"] if event["team"] == "A")
    first_a.update(
        points=1,
        period=4,
        mark="filled_dot",
        scorer_circled=False,
        boundary="none",
        ink_role="neutral",
    )

    scene = build_scene(ScoresheetDocument.model_validate(draft))
    score_mark = next(item for item in scene if item.get("field_id") == "score.A.002.mark")
    boundary = [item for item in scene if item.get("field_id") == "score.A.002.boundary"]

    assert score_mark["type"] == "line"
    assert [item["type"] for item in boundary].count("circle") == 1
    assert [item["type"] for item in boundary].count("line") == 2


def make_ready(
    scoresheet: GameScoresheet,
    actor: Account,
    token: str,
    *,
    client_id: str = "web-1",
) -> GameScoresheet:
    save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=actor,
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id=client_id,
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": valid_document(scoresheet)}],
        explicit_save=True,
    )
    scoresheet.refresh_from_db()
    validate_scoresheet(
        scoresheet_id=scoresheet.id,
        actor=actor,
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id=client_id,
        surface=ScoresheetEditLease.Surface.WEB,
    )
    scoresheet.refresh_from_db()
    warning_ids = [row["id"] for row in scoresheet.validation_report.get("warnings", [])]
    if warning_ids:
        acknowledge_warnings(
            scoresheet_id=scoresheet.id,
            actor=actor,
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id=client_id,
            surface=ScoresheetEditLease.Surface.WEB,
            warning_ids=warning_ids,
        )
        scoresheet.refresh_from_db()
    assert scoresheet.validation_report["errors"] == []
    return scoresheet


def test_single_editor_lease_expiry_and_superadmin_takeover(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    second_admin = Account.objects.create_user(
        username="scoresheet-admin-2", password="password", role=Account.Role.SUPERADMIN
    )
    superadmin = Account.objects.create_user(
        username="scoresheet-root", password="password", role=Account.Role.SUPERADMIN
    )

    old_token = obtain_lease(scoresheet, setup["admin"], "web-owner")
    holder, other_token, read_only, reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=second_admin,
        client_id="mini-reader",
        surface=ScoresheetEditLease.Surface.MINIAPP,
    )
    assert read_only is True
    assert other_token is None
    assert reason
    assert holder is not None
    assert holder.account_id == setup["admin"].id

    with pytest.raises(ScoresheetError, match="二次确认"):
        force_takeover_edit_lease(
            scoresheet_id=scoresheet.id,
            actor=superadmin,
            client_id="root-mini",
            surface=ScoresheetEditLease.Surface.MINIAPP,
            confirmed=False,
        )
    _, root_token = force_takeover_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=superadmin,
        client_id="root-mini",
        surface=ScoresheetEditLease.Surface.MINIAPP,
        confirmed=True,
    )
    assert root_token
    with pytest.raises(ScoresheetError) as lost:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=old_token,
            client_id="web-owner",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/header/game_number", "value": "失效客户端"}],
        )
    assert lost.value.code == "LEASE_LOST"

    lease = ScoresheetEditLease.objects.get(scoresheet=scoresheet)
    lease.expires_at = timezone.now() - timedelta(seconds=1)
    lease.save(update_fields=["expires_at"])
    new_holder, new_token, read_only, _reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=second_admin,
        client_id="mini-reader",
        surface=ScoresheetEditLease.Surface.MINIAPP,
    )
    assert read_only is False
    assert new_token
    assert new_holder.account_id == second_admin.id
    assert AdminAuditLog.objects.filter(action="SCORESHEET_LEASE_FORCE_TAKEN").exists()


def test_ordinary_admin_is_read_only_during_ai_then_can_edit_after_failure(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path, QWEN_API_KEY="test-key"):
        setup, _, _, _, scoresheet = create_scoresheet()
    ordinary = setup["ordinary_admin"]

    holder, token, read_only, reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=ordinary,
        client_id="ordinary-before-ai",
        surface=ScoresheetEditLease.Surface.WEB,
    )
    assert holder is None
    assert token is None
    assert read_only is True
    assert "识别" in reason

    ScoresheetRecognitionRun.objects.filter(
        scoresheet=scoresheet,
        source_version=scoresheet.source_version,
    ).update(
        status=ScoresheetRecognitionRun.Status.FAILED,
        last_error_code="CREDENTIALS_MISSING",
        finished_at=timezone.now(),
    )
    holder, token, read_only, _reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=ordinary,
        client_id="ordinary-after-ai",
        surface=ScoresheetEditLease.Surface.WEB,
    )

    assert read_only is False
    assert token
    assert holder.account_id == ordinary.id


def test_field_change_invalidates_region_and_is_available_to_sync(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    token = obtain_lease(scoresheet, setup["admin"])
    review_region(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
        region="OFFICIALS",
        reviewed=True,
    )
    scoresheet.refresh_from_db()
    after_event = scoresheet.event_sequence
    old_version = scoresheet.draft_version
    save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=old_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/officials/0/name", "value": "同步测试记录员"}],
    )
    scoresheet.refresh_from_db()
    assert scoresheet.draft_version == old_version + 1
    assert "OFFICIALS" not in scoresheet.reviewed_regions
    events = sync_scoresheet(scoresheet, after_event)
    assert [event.event_type for event in events] == ["FIELD_EDIT"]
    assert events[0].changed_fields[0]["path"] == "/officials/scorer/name"
    with pytest.raises(ScoresheetError) as stale:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=old_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/header/game_number", "value": "旧值覆盖"}],
        )
    assert stale.value.code == "VERSION_CONFLICT"


def test_game_prior_fields_are_locked_and_final_result_is_derived(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    token = obtain_lease(scoresheet, setup["admin"])

    with pytest.raises(ScoresheetDocumentError) as locked:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/header/game_number", "value": "不可修改"}],
        )
    assert locked.value.code == "SCORESHEET_FIELD_LOCKED"

    replacement = copy.deepcopy(scoresheet.draft)
    replacement["header"]["game_number"] = "也不可通过整表覆盖"
    replacement["stated_period_scores"] = [
        {"period": 1, "team_a": 2, "team_b": 3},
        {"period": 2, "team_a": 3, "team_b": 1},
    ]
    replacement["score_events"] = [
        {
            "sequence": 1,
            "team": "A",
            "period": 99,
            "points": 0,
            "cumulative_score": 2,
            "scorer_jersey": replacement["teams"][0]["players"][0]["jersey_number"],
            "mark": "forged",
            "scorer_circled": True,
            "boundary": "forged",
            "ink_role": "forged",
        },
        {
            "sequence": 1,
            "team": "A",
            "period": 8,
            "points": 9,
            "cumulative_score": 5,
            "scorer_jersey": replacement["teams"][0]["players"][0]["jersey_number"],
            "mark": None,
            "scorer_circled": False,
            "boundary": "none",
            "ink_role": "neutral",
        },
        {
            "sequence": 93,
            "team": "B",
            "period": 8,
            "points": 9,
            "cumulative_score": 3,
            "scorer_jersey": replacement["teams"][1]["players"][0]["jersey_number"],
            "mark": None,
            "scorer_circled": False,
            "boundary": "none",
            "ink_role": "neutral",
        },
        {
            "sequence": 94,
            "team": "B",
            "period": 8,
            "points": 9,
            "cumulative_score": 4,
            "scorer_jersey": replacement["teams"][1]["players"][0]["jersey_number"],
            "mark": None,
            "scorer_circled": False,
            "boundary": "none",
            "ink_role": "neutral",
        },
    ]
    replacement["final_score"].update(team_a=99, team_b=98, winner_name="伪造胜队")
    saved = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": replacement}],
    )
    assert saved.draft["header"]["game_number"] == scoresheet.draft["header"]["game_number"]
    assert saved.draft["final_score"] == {
        "team_a": 5,
        "team_b": 4,
        "winner_name": saved.draft["teams"][0]["name"],
        "ended_at": "",
    }
    assert [
        (
            event["team"],
            event["cumulative_score"],
            event["points"],
            event["period"],
            event["mark"],
            event["scorer_circled"],
            event["boundary"],
            event["ink_role"],
        )
        for event in saved.draft["score_events"]
    ] == [
        ("A", 2, 2, 1, "diagonal", False, "period_end", "q1_q3"),
        ("B", 3, 3, 1, "diagonal", True, "period_end", "q1_q3"),
        ("A", 5, 3, 2, "diagonal", True, "game_end", "q2_q4_ot"),
        ("B", 4, 1, 2, "filled_dot", False, "game_end", "q2_q4_ot"),
    ]

    with pytest.raises(ScoresheetDocumentError) as derived_locked:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=saved.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/score_events/0/period", "value": 8}],
        )
    assert derived_locked.value.code == "SCORESHEET_FIELD_LOCKED"

    invalid_gap = copy.deepcopy(saved.draft)
    invalid_gap["score_events"] = [
        event
        for event in invalid_gap["score_events"]
        if not (event["team"] == "A" and event["cumulative_score"] == 2)
    ]
    invalid_saved = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=saved.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": invalid_gap}],
    )
    invalid_event = next(
        event
        for event in invalid_saved.draft["score_events"]
        if event["team"] == "A" and event["cumulative_score"] == 5
    )
    assert invalid_event["points"] == 5
    assert invalid_event["mark"] is None
    validated = validate_scoresheet(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=invalid_saved.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
    )
    error_codes = {row["code"] for row in validated.validation_report["errors"]}
    assert {"INVALID_SCORE_POINTS", "SCORE_SEQUENCE_GAP"} <= error_codes


def test_unassigned_table_personnel_can_be_saved_before_recognition(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    assert scoresheet.draft["recognition"] is None
    token = obtain_lease(scoresheet, setup["admin"], "manual-personnel-tab")
    incoming = copy.deepcopy(scoresheet.draft)
    incoming["recognition"] = {
        "run_id": "manual-table-personnel",
        "notes": "",
        "table_personnel": ["无法归类人员"],
        "problem_paths": [],
        "issues": [],
        "applied_at": timezone.now().isoformat(),
    }

    saved = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="manual-personnel-tab",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": incoming}],
    )

    assert saved.draft["status"] == "draft"
    assert saved.draft["recognition"]["run_id"] == "manual-table-personnel"
    assert saved.draft["recognition"]["notes"] == ""
    assert saved.draft["recognition"]["problem_paths"] == []
    assert saved.draft["recognition"]["table_personnel"] == ["无法归类人员"]
    assert saved.change_logs.order_by("-event_sequence").first().changed_fields == [
        {"path": "/table_personnel/0", "before": None, "after": "无法归类人员"}
    ]

    recognized = copy.deepcopy(saved.draft)
    recognized["recognition"] = {
        "run_id": "real-recognition-run",
        "notes": "",
        "table_personnel": ["模型识别人名", "无法归类人员"],
        "problem_paths": [],
        "issues": [],
        "applied_at": timezone.now().isoformat(),
    }
    merged = merge_recognition_result(
        saved.draft,
        recognized,
        saved.roster_snapshot,
        run_id="real-recognition-run",
    )
    assert merged["recognition"]["table_personnel"] == [
        "无法归类人员",
        "模型识别人名",
    ]
    assert merged["recognition"]["run_id"] == "real-recognition-run"

    cleared = copy.deepcopy(saved.draft)
    cleared["recognition"]["table_personnel"] = []
    saved = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=saved.draft_version,
        lease_token=token,
        client_id="manual-personnel-tab",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": cleared}],
    )
    assert saved.draft["recognition"] is None
    assert saved.draft["status"] == "draft"


def test_recognition_source_fields_are_server_owned_and_problems_only_shrink(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    scoresheet.draft["recognition"] = {
        "run_id": "server-run",
        "notes": "模型原始备注",
        "table_personnel": ["原人员"],
        "problem_paths": ["/teams/0/players/0/name", "/officials/scorer/name"],
        "issues": [
            {
                "code": "MODEL_WARNING",
                "path": "/officials/scorer/name",
                "message": "待人工核对",
                "observed": None,
                "expected": None,
            }
        ],
        "applied_at": timezone.now().isoformat(),
    }
    scoresheet.save(update_fields=["draft", "updated_at"])
    token = obtain_lease(scoresheet, setup["admin"], "recognition-guard")

    forged = copy.deepcopy(scoresheet.draft)
    forged["header"]["crew_chief"] = "正常人工修改"
    forged["recognition"]["run_id"] = "client-forged"
    with pytest.raises(ScoresheetDocumentError) as root_locked:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="recognition-guard",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/", "operation": "SET", "value": forged}],
        )
    assert root_locked.value.code == "SCORESHEET_FIELD_LOCKED"
    scoresheet.refresh_from_db()
    assert scoresheet.draft["header"]["crew_chief"] != "正常人工修改"

    for field, forged_value in (
        ("run_id", "client-forged"),
        ("notes", "伪造备注"),
        ("applied_at", "2099-01-01T00:00:00+00:00"),
    ):
        for operation in ("SET", "DELETE"):
            with pytest.raises(ScoresheetDocumentError) as patch_locked:
                save_draft_changes(
                    scoresheet_id=scoresheet.id,
                    actor=setup["admin"],
                    expected_version=scoresheet.draft_version,
                    lease_token=token,
                    client_id="recognition-guard",
                    surface=ScoresheetEditLease.Surface.WEB,
                    changes=[
                        {
                            "path": f"/recognition/{field}",
                            "operation": operation,
                            "value": forged_value,
                        }
                    ],
                )
            assert patch_locked.value.code == "SCORESHEET_FIELD_LOCKED"

    original_crew_chief = scoresheet.draft["header"]["crew_chief"]
    with pytest.raises(ScoresheetDocumentError) as mixed_locked:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="recognition-guard",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[
                {"path": "/header/crew_chief", "value": "不能部分写入"},
                {"path": "/recognition/run_id", "value": "client-forged"},
            ],
        )
    assert mixed_locked.value.code == "SCORESHEET_FIELD_LOCKED"
    scoresheet.refresh_from_db()
    assert scoresheet.draft["header"]["crew_chief"] == original_crew_chief

    forbidden_problem_edits = (
        {
            "path": "/recognition/problem_paths",
            "value": [*scoresheet.draft["recognition"]["problem_paths"], "/header/venue"],
        },
        {"path": "/recognition/problem_paths/0", "operation": "DELETE"},
        {"path": "/recognition/issues/0/message", "operation": "DELETE"},
        {
            "path": "/recognition/issues",
            "value": [
                {
                    **scoresheet.draft["recognition"]["issues"][0],
                    "message": "客户端改写的问题",
                }
            ],
        },
    )
    for change in forbidden_problem_edits:
        with pytest.raises(ScoresheetDocumentError) as problem_locked:
            save_draft_changes(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=token,
                client_id="recognition-guard",
                surface=ScoresheetEditLease.Surface.WEB,
                changes=[change],
            )
        assert problem_locked.value.code == "SCORESHEET_FIELD_LOCKED"

    forged_problem_document = copy.deepcopy(scoresheet.draft)
    forged_problem_document["recognition"]["issues"].append(
        {
            "code": "CLIENT_WARNING",
            "path": "/header/venue",
            "message": "客户端新增的问题",
            "observed": None,
            "expected": None,
        }
    )
    with pytest.raises(ScoresheetDocumentError) as document_locked:
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="recognition-guard",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/", "operation": "SET", "value": forged_problem_document}],
        )
    assert document_locked.value.code == "SCORESHEET_FIELD_LOCKED"

    reduced = copy.deepcopy(scoresheet.draft)
    reduced["recognition"]["problem_paths"] = ["/officials/scorer/name"]
    reduced["recognition"]["issues"] = []
    saved = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="recognition-guard",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": reduced}],
    )
    assert saved.draft["recognition"]["problem_paths"] == ["/officials/scorer/name"]
    assert saved.draft["recognition"]["issues"] == []


def test_scoresheet_queue_is_server_prioritized_searchable_and_paginated(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    scoresheet.status = GameScoresheet.Status.DRAFT
    scoresheet.save(update_fields=["status", "updated_at"])

    client = Client()
    client.force_login(setup["admin"])
    first_page = client.get(
        "/api/v1/scoresheets/",
        {"scope": "ACTION_REQUIRED", "page": 1, "page_size": 1},
    )
    assert first_page.status_code == 200
    first_payload = first_page.json()
    assert first_payload["total"] == 2
    assert first_payload["division_names"] == [setup["division"].name]
    assert len(first_payload["items"]) == 1
    assert first_payload["items"][0]["game_id"] == str(scoresheet.game_id)
    assert first_payload["items"][0]["status"] == GameScoresheet.Status.DRAFT

    second_page = client.get(
        "/api/v1/scoresheets/",
        {"scope": "ACTION_REQUIRED", "page": 2, "page_size": 1},
    )
    assert second_page.status_code == 200
    assert second_page.json()["items"][0]["status"] == GameScoresheet.Status.NO_SOURCE

    search = client.get(
        "/api/v1/scoresheets/",
        {"scope": "ALL", "query": "测试球队 1", "page": 1, "page_size": 20},
    )
    assert search.status_code == 200
    assert search.json()["total"] == 1
    assert search.json()["items"][0]["game_id"] == str(scoresheet.game_id)

    in_progress = client.get(
        "/api/v1/scoresheets/",
        {"scope": "IN_PROGRESS", "page": 1, "page_size": 20},
    )
    assert in_progress.status_code == 200
    assert in_progress.json()["total"] == 0

    exact = client.get(
        "/api/v1/scoresheets/",
        {"game_id": scoresheet.game_id, "page": 1, "page_size": 1},
    )
    assert exact.status_code == 200
    assert exact.json()["total"] == 1
    assert exact.json()["items"][0]["game_id"] == str(scoresheet.game_id)

    upload_only = client.get(
        "/api/v1/scoresheets/",
        {
            "season_id": setup["season"].id,
            "processing": "UPLOAD",
            "page": 1,
            "page_size": 20,
        },
    )
    assert upload_only.status_code == 200
    assert upload_only.json()["total"] == 1
    assert upload_only.json()["items"][0]["status"] == GameScoresheet.Status.NO_SOURCE


def test_publish_does_not_require_region_reviews_and_keeps_field_change_log(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    token = obtain_lease(scoresheet, setup["admin"])
    scoresheet = make_ready(scoresheet, setup["admin"], token)

    updated = copy.deepcopy(scoresheet.draft)
    updated["header"]["crew_chief"] = "只重核工作人员"
    save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": updated}],
    )
    scoresheet.refresh_from_db()
    assert scoresheet.reviewed_regions == {}
    latest_change = scoresheet.change_logs.order_by("-event_sequence").first()
    assert latest_change is not None
    assert [row["path"] for row in latest_change.changed_fields] == ["/header/crew_chief"]

    validate_scoresheet(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
    )
    scoresheet.refresh_from_db()
    warning_ids = [row["id"] for row in scoresheet.validation_report.get("warnings", [])]
    if warning_ids:
        acknowledge_warnings(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
            warning_ids=warning_ids,
        )
    publication = publish_scoresheet(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="web-1",
        surface=ScoresheetEditLease.Surface.WEB,
    )
    assert publication.publication_number == 1


def test_publish_endpoint_replays_same_idempotency_key(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    lease_token = obtain_lease(scoresheet, setup["admin"])
    scoresheet = make_ready(scoresheet, setup["admin"], lease_token)
    payload = {
        "expected_version": scoresheet.draft_version,
        "lease_token": lease_token,
        "client_id": "web-1",
        "surface": ScoresheetEditLease.Surface.WEB,
    }
    client = Client()
    client.force_login(setup["admin"])
    path = f"/api/v1/scoresheets/{scoresheet.id}/publish"

    first = client.post(
        path,
        data=json.dumps(payload),
        content_type="application/json",
        HTTP_IDEMPOTENCY_KEY="scoresheet-publish-test",
    )
    replayed = client.post(
        path,
        data=json.dumps(payload),
        content_type="application/json",
        HTTP_IDEMPOTENCY_KEY="scoresheet-publish-test",
    )
    conflicting = client.post(
        path,
        data=json.dumps({**payload, "expected_version": payload["expected_version"] + 1}),
        content_type="application/json",
        HTTP_IDEMPOTENCY_KEY="scoresheet-publish-test",
    )

    assert first.status_code == 200, first.content
    assert replayed.status_code == 200
    assert replayed.json()["publication"] == first.json()["publication"]
    assert conflicting.status_code == 409
    assert conflicting.json()["code"] == "IDEMPOTENCY_KEY_REUSED"
    assert ScoresheetPublication.objects.filter(scoresheet=scoresheet).count() == 1
    assert ApiIdempotencyRecord.objects.filter(operation="scoresheet.publish").count() == 1


def test_recognition_retry_endpoint_replays_same_idempotency_key(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    latest = scoresheet.recognition_runs.order_by("-created_at").first()
    assert latest is not None
    latest.status = ScoresheetRecognitionRun.Status.FAILED
    latest.last_error_code = "PROVIDER_FAILED"
    latest.save(update_fields=["status", "last_error_code", "updated_at"])
    scoresheet.status = GameScoresheet.Status.RECOGNITION_FAILED
    scoresheet.save(update_fields=["status", "updated_at"])
    lease_token = obtain_lease(scoresheet, setup["admin"], "web-idempotency")
    payload = {
        "expected_version": scoresheet.draft_version,
        "lease_token": lease_token,
        "client_id": "web-idempotency",
        "surface": ScoresheetEditLease.Surface.WEB,
        "confirmed_overwrite": True,
    }
    client = Client()
    client.force_login(setup["admin"])
    path = f"/api/v1/scoresheets/{scoresheet.id}/recognition/retry"

    with override_settings(QWEN_API_KEY="test-key"):
        first = client.post(
            path,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_IDEMPOTENCY_KEY="scoresheet-recognition-retry-test",
        )
        replayed = client.post(
            path,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_IDEMPOTENCY_KEY="scoresheet-recognition-retry-test",
        )
        conflicting = client.post(
            path,
            data=json.dumps({**payload, "expected_version": payload["expected_version"] + 1}),
            content_type="application/json",
            HTTP_IDEMPOTENCY_KEY="scoresheet-recognition-retry-test",
        )

    assert first.status_code == 200, first.content
    assert replayed.status_code == 200
    assert replayed.json()["id"] == first.json()["id"]
    assert conflicting.status_code == 409
    assert conflicting.json()["code"] == "IDEMPOTENCY_KEY_REUSED"
    assert scoresheet.recognition_runs.filter(
        trigger=ScoresheetRecognitionRun.Trigger.MANUAL_RETRY
    ).count() == 1
    assert ApiIdempotencyRecord.objects.filter(
        operation="scoresheet.recognition.retry"
    ).count() == 1


def test_web_edit_is_returned_by_miniapp_sync_endpoint(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    second_admin = Account.objects.create_user(
        username="mini-sync-admin", password="password", role=Account.Role.ADMIN
    )
    web = Client()
    web.force_login(setup["admin"])
    lease_response = web.post(
        f"/api/v1/scoresheets/{scoresheet.id}/lease",
        data=json.dumps({"client_id": "web-sync", "surface": "WEB"}),
        content_type="application/json",
    )
    assert lease_response.status_code == 200
    lease_token = lease_response.json()["lease_token"]
    start_event = scoresheet.event_sequence
    saved = web.patch(
        f"/api/v1/scoresheets/{scoresheet.id}/draft",
        data=json.dumps(
            {
                "expected_version": scoresheet.draft_version,
                "lease_token": lease_token,
                "client_id": "web-sync",
                "surface": "WEB",
                    "changes": [{"path": "/header/crew_chief", "value": "跨端同步裁判"}],
            }
        ),
        content_type="application/json",
    )
    assert saved.status_code == 200

    mini_token = issue_session(second_admin)
    synced = Client().get(
        f"/api/v1/scoresheets/{scoresheet.id}/sync",
        data={"after_version": scoresheet.draft_version, "after_event": start_event},
        HTTP_AUTHORIZATION=f"Bearer {mini_token}",
    )
    assert synced.status_code == 200
    assert synced.json()["current_version"] == scoresheet.draft_version + 1
    assert synced.json()["events"][-1]["changed_fields"][0]["after"] == "跨端同步裁判"
    assert synced.json()["lease"]["surface"] == "WEB"


def test_miniapp_bearer_can_write_without_csrf_cookie(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    mini_token = issue_session(setup["admin"])
    mini = Client(enforce_csrf_checks=True)

    lease_response = mini.post(
        f"/api/v1/scoresheets/{scoresheet.id}/lease",
        data=json.dumps({"client_id": "mini-csrf", "surface": "MINIAPP"}),
        content_type="application/json",
        HTTP_AUTHORIZATION=f"Bearer {mini_token}",
    )

    assert lease_response.status_code == 200
    assert lease_response.json()["read_only"] is False
    lease_token = lease_response.json()["lease_token"]
    saved = mini.patch(
        f"/api/v1/scoresheets/{scoresheet.id}/draft",
        data=json.dumps(
            {
                "expected_version": scoresheet.draft_version,
                "lease_token": lease_token,
                "client_id": "mini-csrf",
                "surface": "MINIAPP",
                "changes": [{"path": "/header/crew_chief", "value": "MINI-NO-CSRF"}],
            }
        ),
        content_type="application/json",
        HTTP_AUTHORIZATION=f"Bearer {mini_token}",
    )

    assert saved.status_code == 200
    assert saved.json()["draft"]["header"]["crew_chief"] == "MINI-NO-CSRF"


@pytest.mark.parametrize("surface", ["WEB", "MINIAPP"])
def test_scoresheet_get_root_roundtrip_accepts_same_recognition_timestamp_instant(
    tmp_path,
    surface,
):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    authoritative_applied_at = "2026-08-25T04:05:06.123456+00:00"
    draft = copy.deepcopy(scoresheet.draft)
    draft["recognition"] = {
        "run_id": "roundtrip-recognition",
        "notes": "服务器识别来源",
        "table_personnel": [],
        "problem_paths": [],
        "issues": [],
        "applied_at": authoritative_applied_at,
    }
    scoresheet.draft = draft
    scoresheet.save(update_fields=["draft", "updated_at"])

    client = Client(enforce_csrf_checks=surface == "MINIAPP")
    headers = {}
    if surface == "WEB":
        client.force_login(setup["admin"])
    else:
        headers["HTTP_AUTHORIZATION"] = f"Bearer {issue_session(setup['admin'])}"
    lease = client.post(
        f"/api/v1/scoresheets/{scoresheet.id}/lease",
        data=json.dumps({"client_id": f"{surface.lower()}-roundtrip", "surface": surface}),
        content_type="application/json",
        **headers,
    )
    assert lease.status_code == 200
    detail = client.get(f"/api/v1/scoresheets/{scoresheet.id}", **headers)
    assert detail.status_code == 200
    incoming = detail.json()["draft"]
    incoming["recognition"]["applied_at"] = "2026-08-25T04:05:06.123Z"
    incoming["header"]["crew_chief"] = f"{surface} 往返裁判"

    saved = client.patch(
        f"/api/v1/scoresheets/{scoresheet.id}/draft",
        data=json.dumps(
            {
                "expected_version": detail.json()["draft_version"],
                "lease_token": lease.json()["lease_token"],
                "client_id": f"{surface.lower()}-roundtrip",
                "surface": surface,
                "changes": [{"path": "/", "operation": "SET", "value": incoming}],
            }
        ),
        content_type="application/json",
        **headers,
    )

    assert saved.status_code == 200, saved.content
    scoresheet.refresh_from_db()
    assert scoresheet.draft["header"]["crew_chief"] == f"{surface} 往返裁判"
    assert datetime.fromisoformat(
        scoresheet.draft["recognition"]["applied_at"].replace("Z", "+00:00")
    ) == datetime.fromisoformat(authoritative_applied_at)
    assert scoresheet.draft["recognition"]["run_id"] == "roundtrip-recognition"
    assert scoresheet.draft["recognition"]["notes"] == "服务器识别来源"


def test_retryable_recognition_uses_initial_call_plus_three_retries(tmp_path, monkeypatch):
    from core.services import scoresheet_recognition as recognition

    with override_settings(MEDIA_ROOT=tmp_path, QWEN_API_KEY="test-key"):
        _, _, _, _, scoresheet = create_scoresheet()

        def fail(_run):
            raise RecognitionAttemptError(
                "QWEN_NETWORK_ERROR", "temporary", retryable=True
            )

        monkeypatch.setattr(recognition, "call_qwen", fail)
        expected_delays = [30, 30, 30]
        for attempt in range(4):
            before = timezone.now()
            outcome = run_once("test-worker")
            run = ScoresheetRecognitionRun.objects.get(scoresheet=scoresheet)
            assert run.attempt_count == attempt + 1
            if attempt < 3:
                assert outcome == "retry_wait"
                assert run.status == ScoresheetRecognitionRun.Status.RETRY_WAIT
                remaining = (run.next_attempt_at - before).total_seconds()
                assert expected_delays[attempt] - 2 <= remaining <= expected_delays[attempt] + 2
                run.next_attempt_at = timezone.now() - timedelta(seconds=1)
                run.save(update_fields=["next_attempt_at"])
            else:
                assert outcome == "failed"
                assert run.status == ScoresheetRecognitionRun.Status.FAILED
        scoresheet.refresh_from_db()
        assert scoresheet.status == GameScoresheet.Status.RECOGNITION_FAILED


def test_running_recognition_worker_lease_renews_during_long_provider_calls(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path, QWEN_API_KEY="test-key"):
        _, _, _, _, scoresheet = create_scoresheet()
        claim = claim_next_run("long-call-worker")
        assert claim is not None
        run = ScoresheetRecognitionRun.objects.get(scoresheet=scoresheet)
        run.worker_lease_expires_at = timezone.now() + timedelta(seconds=1)
        run.save(update_fields=["worker_lease_expires_at"])

        assert _renew_worker_lease(claim) is True

        run.refresh_from_db()
        assert run.worker_lease_expires_at >= timezone.now() + timedelta(minutes=4)


def test_qwen_prior_contains_only_team_and_player_names(tmp_path):
    from core.services.scoresheet_recognition import _prompt, _provider_prior

    with override_settings(MEDIA_ROOT=tmp_path):
        _, _, _, _, scoresheet = create_scoresheet()
    prior = _provider_prior(scoresheet)
    serialized = json.dumps(prior, ensure_ascii=False)
    prompt = _prompt(prior)
    assert set(prior) == {"teams"}
    assert set(prior["teams"]["A"]) == {"name", "players"}
    assert set(prior["teams"]["A"]["players"][0]) == {"name"}
    assert str(scoresheet.id) not in serialized
    assert scoresheet.game_prior_snapshot["date"] not in serialized
    assert scoresheet.game_prior_snapshot["venue"] not in serialized
    assert "甲队1号" in prompt and "乙队1号" in prompt


def test_late_success_is_retained_but_never_overwrites_human_edits(tmp_path, monkeypatch):
    from core.services import scoresheet_recognition as recognition

    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/header/crew_chief", "value": "人工确认裁判"}],
        )
        scoresheet.refresh_from_db()
        with override_settings(QWEN_API_KEY="test-key"):
            retry_recognition(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=token,
                client_id="web-1",
                surface=ScoresheetEditLease.Surface.WEB,
                confirmed_overwrite=True,
            )
            claim = claim_next_run("race-worker")
        assert claim is not None
        # Only edits made *after* the explicit overwrite confirmation are stale
        # for this run. Pre-retry manual edits are intentionally replaced.
        scoresheet.refresh_from_db()
        scoresheet.draft_version += 1
        scoresheet.draft["revision"] = scoresheet.draft_version
        scoresheet.save(update_fields=["draft_version", "draft"])
        monkeypatch.setattr(
            recognition,
            "call_qwen",
            lambda _run: ({"game": {"venue": "模型错误场地"}}, {"total_tokens": 12}),
        )
        assert execute_claim(claim) == "stored_not_applied"
    scoresheet.refresh_from_db()
    run = (
        ScoresheetRecognitionRun.objects.filter(scoresheet=scoresheet)
        .order_by("-cycle")
        .first()
    )
    assert run is not None
    assert scoresheet.draft["header"]["crew_chief"] == "人工确认裁判"
    assert run.status == ScoresheetRecognitionRun.Status.SUCCEEDED
    assert run.provider_result["game"]["venue"] == "模型错误场地"
    assert scoresheet.change_logs.filter(
        event_type="RECOGNITION_STORED_NOT_APPLIED"
    ).exists()


@pytest.mark.parametrize(
    ("status", "source_offset", "provider_result", "expected_code"),
    [
        (
            ScoresheetRecognitionRun.Status.FAILED,
            0,
            {"header": {"crew_chief": "不应写入的识别结果"}},
            "RECOGNITION_NOT_READY",
        ),
        (
            ScoresheetRecognitionRun.Status.SUPERSEDED,
            0,
            {"header": {"crew_chief": "不应写入的识别结果"}},
            "RECOGNITION_NOT_READY",
        ),
        (
            ScoresheetRecognitionRun.Status.SUCCEEDED,
            1,
            {"header": {"crew_chief": "不应写入的识别结果"}},
            "RECOGNITION_SUPERSEDED",
        ),
        (ScoresheetRecognitionRun.Status.SUCCEEDED, 0, {}, "RECOGNITION_NOT_READY"),
    ],
)
def test_apply_recognition_revalidates_locked_current_run_without_writes(
    tmp_path,
    status,
    source_offset,
    provider_result,
    expected_code,
):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    run = scoresheet.recognition_runs.get()
    run.status = status
    run.source_version = scoresheet.source_version + source_offset
    run.provider_result = provider_result
    run.save(update_fields=["status", "source_version", "provider_result", "updated_at"])
    token = obtain_lease(scoresheet, setup["admin"], "recognition-apply-guard")
    before_draft = copy.deepcopy(scoresheet.draft)
    before_version = scoresheet.draft_version
    before_events = scoresheet.change_logs.count()

    with pytest.raises(ScoresheetError) as blocked:
        apply_recognition_regions(
            scoresheet_id=scoresheet.id,
            run_id=run.id,
            actor=setup["admin"],
            expected_version=before_version,
            lease_token=token,
            client_id="recognition-apply-guard",
            surface=ScoresheetEditLease.Surface.WEB,
            regions=["SOURCE_GAME"],
        )

    assert blocked.value.code == expected_code
    scoresheet.refresh_from_db()
    run.refresh_from_db()
    assert scoresheet.draft == before_draft
    assert scoresheet.draft_version == before_version
    assert scoresheet.change_logs.count() == before_events
    assert run.applied_draft_version is None


def test_apply_recognition_accepts_only_current_successful_provider_result(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    run = scoresheet.recognition_runs.get()
    run.status = ScoresheetRecognitionRun.Status.SUCCEEDED
    run.provider_result = {"header": {"crew_chief": "当前识别裁判"}}
    run.save(update_fields=["status", "provider_result", "updated_at"])
    token = obtain_lease(scoresheet, setup["admin"], "recognition-apply-current")

    saved = apply_recognition_regions(
        scoresheet_id=scoresheet.id,
        run_id=run.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="recognition-apply-current",
        surface=ScoresheetEditLease.Surface.WEB,
        regions=["SOURCE_GAME"],
    )

    run.refresh_from_db()
    assert saved.draft["header"]["crew_chief"] == "当前识别裁判"
    assert run.applied_draft_version == saved.draft_version

    event_count = saved.change_logs.count()
    with pytest.raises(ScoresheetError) as repeated:
        apply_recognition_regions(
            scoresheet_id=saved.id,
            run_id=run.id,
            actor=setup["admin"],
            expected_version=saved.draft_version,
            lease_token=token,
            client_id="recognition-apply-current",
            surface=ScoresheetEditLease.Surface.WEB,
            regions=["SOURCE_GAME"],
        )
    assert repeated.value.code == "RECOGNITION_ALREADY_APPLIED"
    saved.refresh_from_db()
    assert saved.draft_version == run.applied_draft_version
    assert saved.change_logs.count() == event_count


def test_reupload_supersedes_old_claim_and_resets_attempt_budget(tmp_path, monkeypatch):
    from core.services import scoresheet_recognition as recognition

    with override_settings(MEDIA_ROOT=tmp_path, QWEN_API_KEY="test-key"):
        setup, _, _, old_asset, scoresheet = create_scoresheet()
        old_claim = claim_next_run("old-worker")
        assert old_claim is not None
        replacement = replace_game_media(
            actor=setup["admin"],
            asset_id=old_asset.id,
            expected_version=old_asset.version,
            scoresheet_complete_confirmed=True,
            uploaded_file=image_file("replacement.jpg"),
        )
        monkeypatch.setattr(
            recognition,
            "call_qwen",
            lambda _run: ({"game": {"venue": "迟到结果"}}, {}),
        )
        assert execute_claim(old_claim) == "superseded"
    scoresheet.refresh_from_db()
    runs = list(scoresheet.recognition_runs.order_by("source_version"))
    assert replacement.id == scoresheet.source_asset_id
    assert runs[0].status == ScoresheetRecognitionRun.Status.SUPERSEDED
    assert runs[1].status == ScoresheetRecognitionRun.Status.QUEUED
    assert runs[1].attempt_count == 0
    assert scoresheet.draft["header"]["venue"] != "迟到结果"


def test_publish_is_atomic_generates_stats_and_limits_leader_view(tmp_path, monkeypatch):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, asset, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)

        original_bulk_create = GameTeamStat.objects.bulk_create

        def fail_bulk_create(*_args, **_kwargs):
            raise RuntimeError("simulated stats failure")

        monkeypatch.setattr(GameTeamStat.objects, "bulk_create", fail_bulk_create)
        with pytest.raises(RuntimeError, match="simulated stats failure"):
            publish_scoresheet(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=token,
                client_id="web-1",
                surface=ScoresheetEditLease.Surface.WEB,
            )
        assert ScoresheetPublication.objects.count() == 0
        game.refresh_from_db()
        asset.refresh_from_db()
        assert game.status == Game.Status.SCHEDULED
        assert asset.review_status == GameMediaAsset.ReviewStatus.PENDING
        assert ScoresheetEditLease.objects.filter(scoresheet=scoresheet).exists()

        monkeypatch.setattr(GameTeamStat.objects, "bulk_create", original_bulk_create)
        publication = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )

    game.refresh_from_db()
    asset.refresh_from_db()
    scoresheet.refresh_from_db()
    assert publication.publication_number == 1
    assert game.status == Game.Status.COMPLETED
    assert [game.home_score, game.away_score] == [2, 1]
    assert asset.review_status == GameMediaAsset.ReviewStatus.APPROVED
    assert scoresheet.current_publication_id == publication.id
    assert GameTeamStat.objects.filter(publication=publication).count() == 2
    assert GamePlayerStat.objects.filter(publication=publication).count() == 10
    assert sum(
        GamePlayerStat.objects.filter(publication=publication).values_list("points", flat=True)
    ) == 3
    assert not ScoresheetEditLease.objects.filter(scoresheet=scoresheet).exists()

    leader_token = issue_session(setup["accounts"][0])
    leader_view = Client().get(
        f"/api/v1/game-media/games/{game.id}",
        HTTP_AUTHORIZATION=f"Bearer {leader_token}",
    )
    assert leader_view.status_code == 200
    assert [row["id"] for row in leader_view.json()["assets"]] == [str(asset.id)]
    public = Client().get(f"/api/v1/public/scoresheet-stats?game_id={game.id}")
    assert public.status_code == 200
    assert public.json()[0]["home_score"] == 2
    assert public.json()[0]["player_stats"][0].keys() >= {
        "points",
        "one_point_events",
        "two_point_events",
        "three_point_events",
        "personal_fouls",
    }

    with pytest.raises(ScoresheetError) as correction:
        publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token="released-token",
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )
    assert correction.value.code in {"LEASE_REQUIRED", "SUPERADMIN_REQUIRED"}


def test_stats_builder_skips_fully_blank_paper_player_rows(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        _, _, _, _, scoresheet = create_scoresheet()
    snapshot = copy.deepcopy(scoresheet.draft)
    snapshot["teams"][0]["players"].append(
        {
            "row": 12,
            "license_number": "",
            "name": "",
            "jersey_number": "",
            "captain": False,
            "participation": "none",
            "fouls": [],
            "post_foul_markers": [],
        }
    )
    publication = ScoresheetPublication(
        scoresheet=scoresheet,
        snapshot=snapshot,
        validation_report={"computed": {"player_points": {}}},
    )
    _, players = _build_stats(publication, scoresheet)
    assert len(players) == 10
    assert all(player.player_name or player.jersey_number for player in players)


def test_published_source_correction_is_superadmin_only_and_old_publication_stays_live(
    tmp_path,
):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, asset, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        publication = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        asset.refresh_from_db()

        admin_client = Client()
        admin_client.force_login(setup["ordinary_admin"])
        blocked_replace = admin_client.post(
            f"/api/v1/admin/game-media/{asset.id}/replace",
            data={
                "expected_version": asset.version,
                "scoresheet_complete_confirmed": "true",
                "image": image_file("admin-correction.jpg"),
            },
        )
        blocked_delete = admin_client.delete(
            f"/api/v1/admin/game-media/{asset.id}",
            data=json.dumps({"expected_version": asset.version}),
            content_type="application/json",
        )
        assert blocked_replace.status_code == 403
        assert blocked_replace.json()["code"] == "SUPERADMIN_REQUIRED"
        assert blocked_delete.status_code == 400
        assert blocked_delete.json()["code"] == "SCORESHEET_DELETE_FORBIDDEN"

        superadmin = Account.objects.create_user(
            username="scoresheet-correction-root",
            password="password",
            role=Account.Role.SUPERADMIN,
        )
        admin_client.force_login(superadmin)
        replacement = admin_client.post(
            f"/api/v1/admin/game-media/{asset.id}/replace",
            data={
                "expected_version": asset.version,
                "scoresheet_complete_confirmed": "true",
                "image": image_file("superadmin-correction.jpg"),
            },
        )
        assert replacement.status_code == 201

        scoresheet.refresh_from_db()
        asset.refresh_from_db()
        assert str(scoresheet.source_asset_id) == replacement.json()["id"]
        assert scoresheet.current_publication_id == publication.id
        assert scoresheet.recognition_runs.filter(
            source_version=scoresheet.source_version,
            status=ScoresheetRecognitionRun.Status.FAILED,
            last_error_code="CREDENTIALS_MISSING",
        ).exists()
        assert asset.deleted_at is not None

        leader_token = issue_session(setup["accounts"][0])
        leader_view = Client().get(
            f"/api/v1/game-media/games/{game.id}",
            HTTP_AUTHORIZATION=f"Bearer {leader_token}",
        )
        assert leader_view.status_code == 200
        assert [row["id"] for row in leader_view.json()["assets"]] == [str(asset.id)]
        old_source_url = leader_view.json()["assets"][0]["content_url"]
        assert Client().get(old_source_url).status_code == 200

        public = Client().get(f"/api/v1/public/scoresheet-stats?game_id={game.id}")
        assert public.status_code == 200
        assert public.json()[0]["publication_number"] == 1
        assert [public.json()[0]["home_score"], public.json()[0]["away_score"]] == [2, 1]

        admin_client.force_login(setup["ordinary_admin"])
        blocked_lease = admin_client.post(
            f"/api/v1/scoresheets/{scoresheet.id}/lease",
            data=json.dumps({"client_id": "normal-admin", "surface": "WEB"}),
            content_type="application/json",
        )
        assert blocked_lease.status_code == 200
        assert blocked_lease.json()["read_only"] is True
        assert "普通管理员" in blocked_lease.json()["read_only_reason"]


def test_archived_published_scoresheet_requires_explicit_web_correction_and_republishes(
    tmp_path,
):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, _, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        first_publication = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        scoresheet.refresh_from_db()

        prearchive_token = obtain_lease(scoresheet, setup["admin"], "archive-correction")
        game.season.status = game.season.Status.ARCHIVED
        game.season.save(update_fields=["status", "updated_at"])

        with pytest.raises(ScoresheetError) as stale_lease:
            save_draft_changes(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=prearchive_token,
                client_id="archive-correction",
                surface=ScoresheetEditLease.Surface.WEB,
                changes=[{"path": "/officials/0/name", "value": "不应写入"}],
            )
        assert stale_lease.value.code == "SEASON_ARCHIVED"

        for actor, surface, confirmed in (
            (setup["admin"], ScoresheetEditLease.Surface.WEB, False),
            (setup["ordinary_admin"], ScoresheetEditLease.Surface.WEB, True),
            (setup["admin"], ScoresheetEditLease.Surface.MINIAPP, True),
        ):
            _, blocked_token, read_only, reason = acquire_edit_lease(
                scoresheet_id=scoresheet.id,
                actor=actor,
                client_id=f"blocked-{surface}-{actor.id}",
                surface=surface,
                archived_correction_confirmed=confirmed,
            )
            assert read_only is True
            assert blocked_token is None
            assert "已归档赛季" in reason

        with pytest.raises(ScoresheetError) as force_without_record_confirmation:
            force_takeover_edit_lease(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                client_id="archive-force-without-confirmation",
                surface=ScoresheetEditLease.Surface.WEB,
                confirmed=True,
                archived_correction_confirmed=False,
            )
        assert force_without_record_confirmation.value.code == "SEASON_ARCHIVED"

        change_count = scoresheet.change_logs.count()
        audit_count = AdminAuditLog.objects.count()
        read_client = Client()
        read_client.force_login(setup["admin"])
        assert read_client.get(f"/api/v1/scoresheets/{scoresheet.id}").status_code == 200
        assert read_client.get(f"/api/v1/scoresheets/{scoresheet.id}/sync").status_code == 200
        assert scoresheet.change_logs.count() == change_count
        assert AdminAuditLog.objects.count() == audit_count

        lease, correction_token, read_only, _ = acquire_edit_lease(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            client_id="archive-correction",
            surface=ScoresheetEditLease.Surface.WEB,
            resume_token=prearchive_token,
            archived_correction_confirmed=True,
        )
        assert read_only is False
        assert correction_token == prearchive_token
        assert lease is not None and lease.archived_correction is True
        assert AdminAuditLog.objects.filter(
            action="ARCHIVED_SCORESHEET_CORRECTION_OPENED",
            object_id=scoresheet.id,
        ).count() == 1

        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=correction_token,
            client_id="archive-correction",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/officials/0/name", "value": "归档纠错记录员"}],
            explicit_save=True,
        )
        scoresheet.refresh_from_db()
        validate_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=correction_token,
            client_id="archive-correction",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        scoresheet.refresh_from_db()
        warning_ids = [row["id"] for row in scoresheet.validation_report.get("warnings", [])]
        if warning_ids:
            acknowledge_warnings(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=correction_token,
                client_id="archive-correction",
                surface=ScoresheetEditLease.Surface.WEB,
                warning_ids=warning_ids,
            )
            scoresheet.refresh_from_db()
        second_publication = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=correction_token,
            client_id="archive-correction",
            surface=ScoresheetEditLease.Surface.WEB,
        )

        assert second_publication.publication_number == 2
        assert second_publication.supersedes_id == first_publication.id
        assert second_publication.snapshot["officials"][0]["name"] == "归档纠错记录员"
        assert ScoresheetPublication.objects.filter(scoresheet=scoresheet).count() == 2
        game.season.refresh_from_db()
        assert game.season.status == game.season.Status.ARCHIVED


def test_publish_rejects_tampered_validation_and_stale_game_prior(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, _, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)

        report = copy.deepcopy(scoresheet.validation_report)
        report["draft_digest"] = "tampered"
        scoresheet.validation_report = report
        scoresheet.save(update_fields=["validation_report", "updated_at"])
        with pytest.raises(ScoresheetError) as tampered:
            publish_scoresheet(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=token,
                client_id="web-1",
                surface=ScoresheetEditLease.Surface.WEB,
            )
        assert tampered.value.code == "VALIDATION_STALE"

        validate_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        game.version += 1
        game.venue_name = "新场地"
        game.save(update_fields=["version", "venue_name", "updated_at"])
        scoresheet.refresh_from_db()
        with pytest.raises(ScoresheetError) as stale_game:
            publish_scoresheet(
                scoresheet_id=scoresheet.id,
                actor=setup["admin"],
                expected_version=scoresheet.draft_version,
                lease_token=token,
                client_id="web-1",
                surface=ScoresheetEditLease.Surface.WEB,
            )
        assert stale_game.value.code == "GAME_CONTEXT_REVIEW_REQUIRED"


def test_publication_stats_are_immutable_and_pdf_is_single_page(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
        document = valid_document(scoresheet)
        rendered = render_scoresheet_pdf(document)
        assert rendered.startswith(b"%PDF")
        assert len(PdfReader(io.BytesIO(rendered)).pages) == 1

        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        publication = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        team_stat = publication.team_stats.first()
        player_stat = publication.player_stats.first()
        assert team_stat is not None and player_stat is not None
        team_stat.total_score = 99
        player_stat.points = 99
        with pytest.raises(ValidationError):
            team_stat.save()
        with pytest.raises(ValidationError):
            player_stat.save()
        with pytest.raises(ValidationError):
            publication.delete()


def test_published_scoresheet_export_endpoints(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, _, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )

        client = Client()
        client.force_login(setup["admin"])
        pdf_response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/exports/pdf")
        csv_response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/exports/csv")
        xlsx_response = client.get(
            f"/api/v1/scoresheets/exports/seasons/{game.season_id}/xlsx"
        )

    assert pdf_response.status_code == 200
    assert pdf_response["Content-Type"] == "application/pdf"
    assert len(PdfReader(io.BytesIO(pdf_response.content)).pages) == 1

    assert csv_response.status_code == 200
    assert csv_response["Content-Type"].startswith("text/csv")
    csv_text = csv_response.content.decode("utf-8-sig")
    assert "甲队1号" in csv_text and "乙队1号" in csv_text

    assert xlsx_response.status_code == 200
    assert xlsx_response["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    assert workbook.sheetnames == ["球队单场统计", "球员单场统计"]
    assert workbook["球队单场统计"].max_row == 3
    assert workbook["球员单场统计"].max_row == 11


def test_superadmin_can_correct_and_republish_the_same_source(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, asset, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        first = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )

        scoresheet.refresh_from_db()
        game.refresh_from_db()
        assert scoresheet.game_prior_snapshot["game_version"] == game.version

        superadmin = Account.objects.create_user(
            username="same-source-correction-root",
            password="password",
            role=Account.Role.SUPERADMIN,
        )
        root_token = obtain_lease(scoresheet, superadmin, "root-web")
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=superadmin,
            expected_version=scoresheet.draft_version,
            lease_token=root_token,
            client_id="root-web",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/officials/0/name", "value": "纠错记录员"}],
        )
        scoresheet.refresh_from_db()
        assert scoresheet.reviewed_regions == {}
        validate_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=superadmin,
            expected_version=scoresheet.draft_version,
            lease_token=root_token,
            client_id="root-web",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        scoresheet.refresh_from_db()
        warning_ids = [row["id"] for row in scoresheet.validation_report.get("warnings", [])]
        if warning_ids:
            acknowledge_warnings(
                scoresheet_id=scoresheet.id,
                actor=superadmin,
                expected_version=scoresheet.draft_version,
                lease_token=root_token,
                client_id="root-web",
                surface=ScoresheetEditLease.Surface.WEB,
                warning_ids=warning_ids,
            )
        second = publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=superadmin,
            expected_version=scoresheet.draft_version,
            lease_token=root_token,
            client_id="root-web",
            surface=ScoresheetEditLease.Surface.WEB,
        )

        scoresheet.refresh_from_db()
        assert second.publication_number == 2
        assert second.supersedes_id == first.id
        assert second.source_asset_id == asset.id
        assert scoresheet.current_publication_id == second.id

        game.refresh_from_db()
        blocked_payload = {
            "expected_version": game.version,
            "date": game.date.isoformat(),
            "period_id": str(game.period_id),
            "start_time": game.start_time.strftime("%H:%M"),
            "standard_venue_id": str(setup["venues"][0].id),
            "venue_name": game.venue_name,
            "home_team_id": str(game.home_team_id),
            "away_team_id": str(game.away_team_id),
            "home_score": 88,
            "away_score": 77,
            "status": Game.Status.COMPLETED,
            "leader_adjustable": game.leader_adjustable,
            "override_rules": True,
            "confirmed": True,
        }
        blocked = Client().put(
            f"/api/v1/admin/mobile/games/{game.id}",
            data=json.dumps(blocked_payload),
            content_type="application/json",
            HTTP_AUTHORIZATION=f"Bearer {issue_session(superadmin)}",
        )
        assert blocked.status_code == 409
        assert blocked.json()["code"] == "SCORESHEET_REPUBLICATION_REQUIRED"


def test_public_stats_and_exports_use_the_immutable_publication_snapshot(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, game, _, _, scoresheet = create_scoresheet()
        published_home_name = game.home_team.name
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )

        game.home_score = 99
        game.away_score = 98
        game.save(update_fields=["home_score", "away_score", "updated_at"])
        game.home_team.name = "后来修改的队名"
        game.home_team.save(update_fields=["name", "updated_at"])

        public = Client().get(f"/api/v1/public/scoresheet-stats?game_id={game.id}")
        assert public.status_code == 200
        assert [public.json()[0]["home_score"], public.json()[0]["away_score"]] == [2, 1]
        assert public.json()[0]["home_name"] == published_home_name
        assert public.json()[0]["team_stats"][0]["team_name"] in {
            published_home_name,
            game.away_team.name,
        }

        client = Client()
        client.force_login(setup["admin"])
        csv_response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/exports/csv")
        xlsx_response = client.get(
            f"/api/v1/scoresheets/exports/seasons/{game.season_id}/xlsx"
        )
        assert published_home_name in csv_response.content.decode("utf-8-sig")
        assert "后来修改的队名" not in csv_response.content.decode("utf-8-sig")
        workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
        team_names = {
            row[2]
            for row in workbook["球队单场统计"].iter_rows(min_row=2, values_only=True)
        }
        assert published_home_name in team_names
        assert "后来修改的队名" not in team_names


def test_same_tab_resumes_exact_lease_token_without_log_noise(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    actor = setup["admin"]
    lease, token, read_only, reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=actor,
        client_id="tab-a",
        surface=ScoresheetEditLease.Surface.WEB,
    )
    assert lease and token and read_only is False and reason == ""

    same_holder, missing_token, read_only, reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=actor,
        client_id="tab-a",
        surface=ScoresheetEditLease.Surface.WEB,
    )
    assert same_holder and missing_token is None and read_only is True
    assert "凭据" in reason

    resumed, resumed_token, read_only, reason = acquire_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=actor,
        client_id="tab-a",
        surface=ScoresheetEditLease.Surface.WEB,
        resume_token=token,
    )
    assert resumed and resumed.id == lease.id
    assert resumed_token == token
    assert read_only is False and reason == ""
    assert scoresheet.change_logs.filter(event_type="LEASE_ACQUIRED").count() == 1
    assert not scoresheet.change_logs.filter(event_type="LEASE_RESUMED").exists()

    for _ in range(2):
        release_edit_lease(
            scoresheet_id=scoresheet.id,
            actor=actor,
            lease_token=token,
            client_id="tab-a",
            surface=ScoresheetEditLease.Surface.WEB,
        )
    assert not ScoresheetEditLease.objects.filter(scoresheet=scoresheet).exists()


def test_changes_endpoint_only_returns_chinese_human_events_with_true_values(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    before = scoresheet.draft["header"]["crew_chief"]
    token = obtain_lease(scoresheet, setup["admin"], "human-log-tab")
    save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="human-log-tab",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/header/crew_chief", "value": "人工日志-42"}],
    )
    release_edit_lease(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        lease_token=token,
        client_id="human-log-tab",
        surface=ScoresheetEditLease.Surface.WEB,
    )

    client = Client()
    client.force_login(setup["admin"])
    response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/changes")

    assert response.status_code == 200
    assert [row["action"] for row in response.json()["items"]] == ["human_edit"]
    entry = response.json()["items"][0]
    assert entry["summary"] == "人工编辑 · 1 项"
    assert entry["changes"] == [
        {
            "path": "/header/crew_chief",
            "before": before,
            "after": "人工日志-42",
        }
    ]


def test_score_change_log_only_exposes_the_selected_jersey(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    token = obtain_lease(scoresheet, setup["admin"], "score-log-tab")

    period_document = copy.deepcopy(scoresheet.draft)
    period_document["stated_period_scores"] = [
        {"period": period, "team_a": 2 if period == 1 else 0, "team_b": 0}
        for period in range(1, 5)
    ]
    scoresheet = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="score-log-tab",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": period_document}],
    )

    jersey = scoresheet.draft["teams"][0]["players"][0]["jersey_number"]
    score_document = copy.deepcopy(scoresheet.draft)
    score_document["score_events"] = [
        {
            "sequence": 99,
            "team": "A",
            "period": 8,
            "points": 9,
            "cumulative_score": 2,
            "scorer_jersey": jersey,
            "mark": None,
            "scorer_circled": False,
            "boundary": "none",
            "ink_role": "neutral",
        }
    ]
    save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="score-log-tab",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": score_document}],
    )

    latest = scoresheet.change_logs.order_by("-event_sequence").first()
    assert latest is not None
    assert latest.changed_fields[0]["path"] == "/score_events/A/cumulative/2"
    assert latest.changed_fields[0]["after"]["points"] == 2

    client = Client()
    client.force_login(setup["admin"])
    response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/changes")

    assert response.status_code == 200
    entry = response.json()["items"][0]
    assert entry["summary"] == "人工编辑 · 1 项"
    assert entry["changes"] == [
        {
            "path": "/score_events/A/cumulative/2/scorer_jersey",
            "before": None,
            "after": jersey,
        }
    ]


def test_duplicate_score_cells_keep_distinct_stable_audit_paths(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
    token = obtain_lease(scoresheet, setup["admin"], "duplicate-score-log-tab")
    jersey = scoresheet.draft["teams"][0]["players"][0]["jersey_number"]
    duplicate_document = copy.deepcopy(scoresheet.draft)
    duplicate_document["stated_period_scores"] = [
        {"period": period, "team_a": 2 if period == 1 else 0, "team_b": 0}
        for period in range(1, 5)
    ]
    event = {
        "sequence": 1,
        "team": "A",
        "period": 1,
        "points": 2,
        "cumulative_score": 2,
        "scorer_jersey": jersey,
        "mark": "diagonal",
        "scorer_circled": False,
        "boundary": "none",
        "ink_role": "q1_q3",
    }
    duplicate_document["score_events"] = [event, {**event, "sequence": 2}]

    saved = save_draft_changes(
        scoresheet_id=scoresheet.id,
        actor=setup["admin"],
        expected_version=scoresheet.draft_version,
        lease_token=token,
        client_id="duplicate-score-log-tab",
        surface=ScoresheetEditLease.Surface.WEB,
        changes=[{"path": "/", "operation": "SET", "value": duplicate_document}],
    )
    latest = saved.change_logs.order_by("-event_sequence").first()

    assert latest is not None
    assert {
        change["path"]
        for change in latest.changed_fields
        if change["path"].startswith("/score_events/")
    } == {
        "/score_events/A/cumulative/2",
        "/score_events/A/cumulative/2#2",
    }

    client = Client()
    client.force_login(setup["admin"])
    response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/changes")
    assert response.status_code == 200
    assert {
        change["path"]
        for change in response.json()["items"][0]["changes"]
        if change["path"].startswith("/score_events/")
    } == {
        "/score_events/A/cumulative/2/scorer_jersey",
        "/score_events/A/cumulative/2#2/scorer_jersey",
    }


def test_recognition_capability_and_removed_stop_endpoint(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path, QWEN_API_KEY=""):
        setup, _, _, _, scoresheet = create_scoresheet()
        client = Client()
        client.force_login(setup["admin"])
        capability = client.get("/api/v1/scoresheets/recognition/capabilities")
        stopped = client.post(f"/api/v1/scoresheets/{scoresheet.id}/recognition/stop")

    assert capability.status_code == 200
    assert capability.json() == {
        "configured": False,
        "provider": "QWEN",
        "model": "qwen3.8-max",
        "prompt_version": "scoresheet-2026-08-24-v25-cn",
        "max_attempts": 4,
        "retry_delays_seconds": [30, 30, 30],
    }
    # No POST handler exists. Ninja's adjacent read-only dynamic route may
    # surface this as 405 instead of 404, but the mutation endpoint is gone.
    assert stopped.status_code in {404, 405}


def test_provider_retry_after_longer_than_default_is_honored(tmp_path, monkeypatch):
    from core.services import scoresheet_recognition as recognition

    with override_settings(MEDIA_ROOT=tmp_path, QWEN_API_KEY="test-key"):
        _, _, _, _, scoresheet = create_scoresheet()

        def fail(_run):
            raise RecognitionAttemptError(
                "QWEN_HTTP_429",
                "rate limited",
                retryable=True,
                retry_after_seconds=75,
            )

        monkeypatch.setattr(recognition, "call_qwen", fail)
        before = timezone.now()
        assert run_once("retry-after-worker") == "retry_wait"
        run = ScoresheetRecognitionRun.objects.get(scoresheet=scoresheet)
        remaining = (run.next_attempt_at - before).total_seconds()
        assert 73 <= remaining <= 77


def test_qwen_request_uses_strict_private_contract_and_audits_image(tmp_path, monkeypatch):
    import base64
    import hashlib
    import sys
    from types import SimpleNamespace

    from core.scoresheet_v2.recognition import PROMPT_VERSION, SYSTEM_PROMPT
    from core.services.scoresheet_recognition import call_qwen

    captured: dict[str, object] = {}

    class FakeCompletions:
        def create(self, **kwargs):
            captured.update(kwargs)
            return []

    class FakeOpenAI:
        def __init__(self, **kwargs):
            captured["client"] = kwargs
            self.chat = SimpleNamespace(completions=FakeCompletions())

    monkeypatch.setitem(sys.modules, "openai", SimpleNamespace(OpenAI=FakeOpenAI))
    with override_settings(
        MEDIA_ROOT=tmp_path,
        QWEN_API_KEY="private-test-key",
        QWEN_MODEL="qwen3.8-max",
        QWEN_REASONING_EFFORT="xhigh",
        SCORESHEET_RECOGNITION_UPSCALE_TARGET_PIXELS=8_000_000,
        SCORESHEET_RECOGNITION_TIMEOUT_SECONDS=180,
    ):
        _, _, _, _, scoresheet = create_scoresheet()
        run = ScoresheetRecognitionRun.objects.get(scoresheet=scoresheet)
        with pytest.raises(RecognitionAttemptError) as failure:
            call_qwen(run)

    assert failure.value.code == "PROVIDER_SCHEMA_INVALID"
    assert hashlib.sha256(SYSTEM_PROMPT.encode()).hexdigest() == (
        "ad8bd0f20d5b0d18496187555d3ad7d7559b734cac08cca1612a56e9e50b8586"
    )
    assert captured["model"] == "qwen3.8-max"
    assert captured["client"]["timeout"] == 180
    assert captured["seed"] == 1234
    assert captured["stream"] is True
    assert captured["stream_options"] == {"include_usage": True}
    assert captured["extra_body"] == {
        "enable_thinking": True,
        "reasoning_effort": "xhigh",
        "vl_high_resolution_images": True,
        "preserve_thinking": False,
    }
    response_format = captured["response_format"]
    assert response_format["json_schema"]["strict"] is True
    schema_text = json.dumps(response_format["json_schema"]["schema"], ensure_ascii=False)
    messages = captured["messages"]
    user_content = messages[1]["content"]
    user_text = next(item["text"] for item in user_content if item["type"] == "text")
    image_url = next(
        item["image_url"]["url"] for item in user_content if item["type"] == "image_url"
    )
    assert scoresheet.draft["teams"][0]["players"][0]["name"] in schema_text
    assert scoresheet.draft["teams"][1]["players"][0]["name"] in schema_text
    assert scoresheet.draft["teams"][0]["name"] in user_text
    assert scoresheet.draft["teams"][1]["name"] in user_text
    forbidden_values = {
        str(scoresheet.id),
        str(scoresheet.game_id),
        scoresheet.game_prior_snapshot["competition"],
        scoresheet.game_prior_snapshot["date"],
        scoresheet.game_prior_snapshot["venue"],
    }
    assert all(value not in user_text + schema_text for value in forbidden_values if value)
    processed_image = base64.b64decode(image_url.split(",", 1)[1])
    run.refresh_from_db()
    assert run.model_name == "qwen3.8-max"
    assert run.prompt_version == PROMPT_VERSION
    assert run.image_sha256 == hashlib.sha256(processed_image).hexdigest()


def test_small_whole_image_is_upscaled_toward_target_with_two_times_axis_cap(tmp_path):
    source = tmp_path / "small.png"
    Image.new("RGB", (100, 200), "white").save(source)

    payload, _, _ = prepare_image(source, 80_000)

    with Image.open(io.BytesIO(payload)) as prepared:
        assert prepared.size == (200, 400)
        assert prepared.width * prepared.height == 80_000


@pytest.mark.parametrize(
    ("image_format", "mime_type"),
    [("JPEG", "image/jpeg"), ("PNG", "image/png")],
)
def test_native_large_jpeg_and_png_are_not_resampled_or_reencoded(
    tmp_path,
    image_format: str,
    mime_type: str,
):
    source = tmp_path / f"large.{image_format.lower()}"
    Image.new("RGB", (40, 30), "white").save(source, format=image_format)
    original = source.read_bytes()

    payload, data_url, digest = prepare_image(source, 1_000)

    assert payload == original
    assert data_url.startswith(f"data:{mime_type};base64,")
    assert len(data_url.encode("ascii")) == _data_url_size(len(payload), mime_type)
    assert digest == hashlib.sha256(original).hexdigest()


def test_native_webp_is_preserved_below_direct_resolution_threshold(tmp_path):
    source = tmp_path / "native.webp"
    Image.new("RGB", (40, 30), "white").save(source, format="WEBP")
    original = source.read_bytes()

    payload, data_url, _ = prepare_image(source, 1_000)

    assert payload == original
    assert data_url.startswith("data:image/webp;base64,")


def test_large_webp_is_converted_to_same_size_jpeg(tmp_path, monkeypatch):
    monkeypatch.setattr("core.scoresheet_v2.recognition.QWEN_WEBP_MAX_LONG_EDGE", 10)
    monkeypatch.setattr("core.scoresheet_v2.recognition.QWEN_WEBP_MAX_SHORT_EDGE", 5)
    source = tmp_path / "large.webp"
    Image.new("RGB", (20, 10), "white").save(source, format="WEBP")

    payload, data_url, _ = prepare_image(source, 1)

    assert data_url.startswith("data:image/jpeg;base64,")
    with Image.open(io.BytesIO(payload)) as prepared:
        assert prepared.format == "JPEG"
        assert prepared.size == (20, 10)


def test_exif_rotation_is_applied_only_to_qwen_copy(tmp_path):
    source = tmp_path / "rotated.jpg"
    image = Image.new("RGB", (30, 20), "white")
    exif = Image.Exif()
    exif[274] = 6
    image.save(source, format="JPEG", exif=exif)
    original = source.read_bytes()

    payload, data_url, _ = prepare_image(source, 1)

    assert source.read_bytes() == original
    assert payload != original
    assert data_url.startswith("data:image/jpeg;base64,")
    with Image.open(io.BytesIO(payload)) as prepared:
        assert prepared.size == (20, 30)


def test_transparent_background_is_flattened_to_white_for_qwen_jpeg(tmp_path):
    source = tmp_path / "transparent.png"
    Image.new("RGBA", (20, 20), (0, 0, 0, 0)).save(source)

    payload, data_url, _ = prepare_image(source, 1_600)

    assert data_url.startswith("data:image/jpeg;base64,")
    with Image.open(io.BytesIO(payload)) as prepared:
        assert prepared.size == (40, 40)
        assert prepared.convert("RGB").getpixel((20, 20)) == (255, 255, 255)


def test_exact_data_url_boundary_preserves_native_bytes(tmp_path):
    source = tmp_path / "boundary.png"
    Image.new("RGB", (40, 30), "white").save(source, format="PNG")
    original = source.read_bytes()
    limit = _data_url_size(len(original), "image/png")

    payload, data_url, _ = prepare_image(source, 1, max_data_uri_bytes=limit)

    assert payload == original
    assert len(data_url.encode("ascii")) == limit


def test_oversize_data_url_uses_highest_fitting_quality_without_resizing(tmp_path):
    source = tmp_path / "noise.png"
    image = Image.effect_noise((96, 96), 100).convert("RGB")
    image.save(source, format="PNG")
    encoded_by_quality = {quality: _encode_jpeg(image, quality) for quality in range(1, 96)}
    limit = _data_url_size(len(encoded_by_quality[50]), "image/jpeg")
    expected_quality = max(
        quality
        for quality, payload in encoded_by_quality.items()
        if _data_url_size(len(payload), "image/jpeg") <= limit
    )
    assert _data_url_size(source.stat().st_size, "image/png") > limit

    payload, data_url, _ = prepare_image(source, 1, max_data_uri_bytes=limit)

    assert payload == encoded_by_quality[expected_quality]
    assert len(data_url.encode("ascii")) <= limit
    with Image.open(io.BytesIO(payload)) as prepared:
        assert prepared.size == image.size
    image.close()


def test_image_that_cannot_fit_at_quality_one_fails(tmp_path):
    source = tmp_path / "sheet.png"
    Image.new("RGB", (20, 20), "white").save(source)

    with pytest.raises(RecognitionImageError, match="不缩小分辨率"):
        prepare_image(source, 1, max_data_uri_bytes=10)


def test_preparation_failure_skips_qwen_and_all_automatic_retries(
    tmp_path,
    monkeypatch,
):
    import sys
    from types import SimpleNamespace

    from core.services import scoresheet_recognition as recognition

    provider_constructions = 0

    class ForbiddenOpenAI:
        def __init__(self, **kwargs):
            del kwargs
            nonlocal provider_constructions
            provider_constructions += 1

    def fail_preparation(*args, **kwargs):
        del args, kwargs
        raise RecognitionImageError(
            "图片在不缩小分辨率的条件下无法满足 Qwen 20 MB Base64 Data URI 限制。"
        )

    monkeypatch.setitem(sys.modules, "openai", SimpleNamespace(OpenAI=ForbiddenOpenAI))
    monkeypatch.setattr(recognition, "build_context", fail_preparation)
    with override_settings(
        MEDIA_ROOT=tmp_path,
        QWEN_API_KEY="must-not-be-used",
        SCORESHEET_RECOGNITION_UPSCALE_TARGET_PIXELS=8_000_000,
        SCORESHEET_RECOGNITION_TIMEOUT_SECONDS=180,
    ):
        _, _, _, _, scoresheet = create_scoresheet()
        assert run_once("image-preparation-failure") == "failed"
        run = ScoresheetRecognitionRun.objects.get(scoresheet=scoresheet)

    assert run.status == ScoresheetRecognitionRun.Status.FAILED
    assert run.attempt_count == 1
    assert run.max_attempts == 4
    assert run.next_attempt_at is None
    assert run.last_error_code == "IMAGE_DATA_URI_TOO_LARGE"
    assert run.provider_usage == {}
    assert provider_constructions == 0


def test_worker_streams_remote_storage_source_to_temporary_file(
    tmp_path,
    monkeypatch,
):
    from core.services import scoresheet_recognition as recognition

    with override_settings(MEDIA_ROOT=tmp_path):
        _, _, _, _, scoresheet = create_scoresheet()
        run = ScoresheetRecognitionRun.objects.get(scoresheet=scoresheet)
    raw = image_file("remote-source.jpg").read()
    read_sizes: list[int] = []

    class TrackingStream(io.BytesIO):
        def read(self, size=-1):
            read_sizes.append(size)
            return super().read(size)

    class RemoteStorage:
        @staticmethod
        def exists(key):
            del key
            return True

        @staticmethod
        def path(key):
            del key
            raise NotImplementedError

        @staticmethod
        def open(key, mode):
            del key, mode
            return TrackingStream(raw)

    monkeypatch.setattr(recognition, "default_storage", RemoteStorage())
    with recognition._source_image_path(run) as source_path:
        temporary_path = source_path
        assert source_path.read_bytes() == raw

    assert read_sizes
    assert set(read_sizes) == {recognition.SOURCE_CHUNK_BYTES}
    assert not temporary_path.exists()


def test_default_image_contract_uses_decimal_eight_megapixel_upscale_target():
    from django.conf import settings

    assert settings.SCORESHEET_RECOGNITION_UPSCALE_TARGET_PIXELS == 8_000_000
    assert settings.SCORESHEET_RECOGNITION_TIMEOUT_SECONDS == 180
    assert QWEN_DATA_URI_MAX_BYTES == 20_000_000


def test_admin_pdf_uses_current_correction_draft_not_old_publication(tmp_path, monkeypatch):
    from core import api_scoresheets

    captured: dict[str, object] = {}

    def fake_render(document):
        captured["document"] = copy.deepcopy(document)
        return b"%PDF-1.4\n%%EOF\n"

    with override_settings(MEDIA_ROOT=tmp_path):
        setup, _, _, _, scoresheet = create_scoresheet()
        token = obtain_lease(scoresheet, setup["admin"])
        scoresheet = make_ready(scoresheet, setup["admin"], token)
        publish_scoresheet(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=token,
            client_id="web-1",
            surface=ScoresheetEditLease.Surface.WEB,
        )
        scoresheet.refresh_from_db()
        correction_token = obtain_lease(scoresheet, setup["admin"], "correction-tab")
        save_draft_changes(
            scoresheet_id=scoresheet.id,
            actor=setup["admin"],
            expected_version=scoresheet.draft_version,
            lease_token=correction_token,
            client_id="correction-tab",
            surface=ScoresheetEditLease.Surface.WEB,
            changes=[{"path": "/officials/0/name", "value": "纠错后的记录员"}],
        )
        monkeypatch.setattr(api_scoresheets, "render_scoresheet_pdf", fake_render)
        client = Client()
        client.force_login(setup["admin"])
        response = client.get(f"/api/v1/scoresheets/{scoresheet.id}/exports/pdf")

    assert response.status_code == 200
    assert captured["document"]["officials"][0]["name"] == "纠错后的记录员"
