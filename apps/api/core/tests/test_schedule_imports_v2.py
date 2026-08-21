from __future__ import annotations

from datetime import time, timedelta
from io import BytesIO

import pytest
from django.test import override_settings
from django.utils import timezone
from openpyxl import load_workbook

from core.models import (
    Account,
    AdminAuditLog,
    CompetitionGroup,
    Division,
    DrawAssignment,
    Game,
    GameMediaAsset,
    ImportIssue,
    ParticipantSlot,
    Period,
    PeriodCapacity,
    RescheduleRequest,
    ScheduleImportBatch,
    Season,
    SlotReservation,
    Team,
    Venue,
)
from core.services.schedule_imports_v2 import (
    EXPECTED_SHEETS,
    GAME_START_ROW,
    GRID_HEADER_ROW,
    GRID_START_ROW,
    SPECIAL_START_ROW,
    STRUCTURE_START_ROW,
    VENUE_START_COLUMN,
    ScheduleImportError,
    confirm_schedule_import,
    generate_schedule_template,
    reset_schedule_imports,
    schedule_import_reset_preview,
    validate_schedule_upload,
)

pytestmark = [
    pytest.mark.django_db,
    pytest.mark.skip(reason="V2 四页赛程协议已退役；当前服务器仅接受 V3。"),
]


def import_setup():
    starts_on = timezone.localdate() + timedelta(days=30)
    season = Season.objects.create(
        name="2027 北大杯导入测试",
        competition_type=Season.CompetitionType.PKU_CUP,
        year=starts_on.year,
        status=Season.Status.SETUP,
        starts_on=starts_on,
        ends_on=starts_on + timedelta(days=13),
    )
    division = Division.objects.create(
        season=season,
        code="men",
        name="男子组",
        sort_order=1,
    )
    teams = [
        Team.objects.create(
            season=season,
            division=division,
            name=f"测试球队 {index}",
        )
        for index in range(1, 5)
    ]
    period = Period.objects.create(
        season=season,
        code="P1",
        name="第一时段",
        start_time=time(12, 50),
        sort_order=1,
    )
    for day_type in PeriodCapacity.DayType.values:
        PeriodCapacity.objects.create(
            season=season,
            day_type=day_type,
            period=period,
            capacity=2,
        )
    venues = [
        Venue.objects.create(
            season=season,
            name=f"五四东{index}",
            sort_order=index,
        )
        for index in range(1, 3)
    ]
    superadmin = Account.objects.create_user(
        username="schedule-v2-superadmin",
        password="test-password",
        role=Account.Role.SUPERADMIN,
    )
    return {
        "season": season,
        "division": division,
        "teams": teams,
        "period": period,
        "venues": venues,
        "superadmin": superadmin,
    }


def workbook_bytes(workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def grid_cell(workbook, *, target_date, period_code: str, venue_name: str):
    sheet = workbook["赛程网格"]
    column = next(
        current
        for current in range(VENUE_START_COLUMN, sheet.max_column + 1)
        if sheet.cell(GRID_HEADER_ROW, current).value == venue_name
    )
    row = next(
        current
        for current in range(GRID_START_ROW, sheet.max_row + 1)
        if (
            sheet.cell(current, 1).value.date()
            if hasattr(sheet.cell(current, 1).value, "date")
            else sheet.cell(current, 1).value
        )
        == target_date
        and sheet.cell(current, 3).value == period_code
    )
    return sheet.cell(row, column)


def make_workbook(setup, *, code="G001", target_date=None, venue_index=0):
    target_date = target_date or setup["season"].starts_on
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))
    structure = workbook["赛制定义"]
    structure.cell(STRUCTURE_START_ROW, 1, "men")
    structure.cell(STRUCTURE_START_ROW, 2, "A")
    structure.cell(STRUCTURE_START_ROW, 3, "A 组")
    structure.cell(STRUCTURE_START_ROW, 4, 1)
    structure.cell(STRUCTURE_START_ROW, 5, "A1")
    structure.cell(STRUCTURE_START_ROW, 6, "A 组 1 号签")
    structure.cell(STRUCTURE_START_ROW, 7, 1)
    structure.cell(STRUCTURE_START_ROW + 1, 1, "men")
    structure.cell(STRUCTURE_START_ROW + 1, 2, "A")
    structure.cell(STRUCTURE_START_ROW + 1, 3, "A 组")
    structure.cell(STRUCTURE_START_ROW + 1, 4, 1)
    structure.cell(STRUCTURE_START_ROW + 1, 5, "A2")
    structure.cell(STRUCTURE_START_ROW + 1, 6, "A 组 2 号签")
    structure.cell(STRUCTURE_START_ROW + 1, 7, 2)

    games = workbook["比赛清单"]
    for column, value in enumerate(
        [code, "men", "A", "小组赛", 1, "A1", "A2"],
        start=1,
    ):
        games.cell(GAME_START_ROW, column, value)
    grid_cell(
        workbook,
        target_date=target_date,
        period_code=setup["period"].code,
        venue_name=setup["venues"][venue_index].name,
    ).value = code
    return workbook


def validate_workbook(setup, workbook, tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        return validate_schedule_upload(
            actor=setup["superadmin"],
            season=setup["season"],
            content=workbook_bytes(workbook),
            source_name="schedule.xlsx",
        )


def confirm_workbook(setup, workbook, tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path):
        batch = validate_schedule_upload(
            actor=setup["superadmin"],
            season=setup["season"],
            content=workbook_bytes(workbook),
            source_name="schedule.xlsx",
        )
        assert batch.summary["error_count"] == 0
        batch = confirm_schedule_import(
            actor=setup["superadmin"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )
    setup["season"].refresh_from_db()
    return batch


def test_template_has_five_visible_sheets():
    setup = import_setup()
    workbook = load_workbook(BytesIO(generate_schedule_template(setup["season"])))

    assert workbook.sheetnames == EXPECTED_SHEETS
    assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)
    assert workbook["填写说明"]["B4"].value == "2.1.0"
    assert not workbook.defined_names


def test_validate_then_confirm_only_creates_new_records(tmp_path):
    setup = import_setup()
    batch = validate_workbook(setup, make_workbook(setup), tmp_path)

    assert batch.status == ScheduleImportBatch.Status.VALIDATED
    assert batch.summary["error_count"] == 0
    assert batch.summary["new_group_count"] == 1
    assert batch.summary["new_slot_count"] == 2
    assert batch.summary["new_game_count"] == 1
    assert CompetitionGroup.objects.count() == 0
    assert ParticipantSlot.objects.count() == 0
    assert Game.objects.count() == 0

    with override_settings(MEDIA_ROOT=tmp_path):
        confirmed = confirm_schedule_import(
            actor=setup["superadmin"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )

    game = Game.objects.get(code="G001")
    assert confirmed.status == ScheduleImportBatch.Status.CONFIRMED
    assert game.leader_adjustable is True
    assert game.created_by_import_batch_id == batch.id
    assert game.home_team_id is None and game.home_slot.code == "A1"
    assert CompetitionGroup.objects.get().created_by_import_batch_id == batch.id
    assert set(
        ParticipantSlot.objects.values_list("created_by_import_batch_id", flat=True)
    ) == {batch.id}
    assert AdminAuditLog.objects.filter(
        action="SCHEDULE_IMPORT_CONFIRMED",
        object_id=batch.id,
    ).exists()


def test_second_import_references_structure_and_adds_only_new_game(tmp_path):
    setup = import_setup()
    confirm_workbook(setup, make_workbook(setup), tmp_path)
    second = make_workbook(
        setup,
        code="G002",
        target_date=setup["season"].starts_on + timedelta(days=1),
        venue_index=1,
    )

    batch = validate_workbook(setup, second, tmp_path)
    assert batch.summary["error_count"] == 0
    assert batch.summary["new_group_count"] == 0
    assert batch.summary["referenced_group_count"] == 1
    assert batch.summary["new_slot_count"] == 0
    assert batch.summary["referenced_slot_count"] == 2
    assert batch.summary["new_game_count"] == 1

    with override_settings(MEDIA_ROOT=tmp_path):
        confirm_schedule_import(
            actor=setup["superadmin"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )
    assert set(Game.objects.values_list("code", flat=True)) == {"G001", "G002"}
    assert CompetitionGroup.objects.count() == 1
    assert ParticipantSlot.objects.count() == 2


def test_existing_game_code_and_structure_conflict_block_confirmation(tmp_path):
    setup = import_setup()
    confirm_workbook(setup, make_workbook(setup), tmp_path)
    duplicate = make_workbook(setup)
    duplicate["赛制定义"].cell(STRUCTURE_START_ROW, 3, "被篡改的小组名")
    duplicate["赛制定义"].cell(STRUCTURE_START_ROW + 1, 3, "被篡改的小组名")

    batch = validate_workbook(setup, duplicate, tmp_path)
    codes = set(batch.issues.values_list("code", flat=True))
    assert {"GAME_CODE_ALREADY_EXISTS", "EXISTING_GROUP_CONFLICT"} <= codes
    with override_settings(MEDIA_ROOT=tmp_path), pytest.raises(ScheduleImportError) as denied:
        confirm_schedule_import(
            actor=setup["superadmin"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )
    assert denied.value.code == "BATCH_HAS_ERRORS"
    assert Game.objects.count() == 1


def test_formula_and_capacity_errors_are_reported_together(tmp_path):
    setup = import_setup()
    PeriodCapacity.objects.filter(season=setup["season"]).update(capacity=1)
    workbook = make_workbook(setup)
    workbook["填写说明"]["B20"] = "=1+1"
    grid_cell(
        workbook,
        target_date=setup["season"].starts_on,
        period_code=setup["period"].code,
        venue_name=setup["venues"][1].name,
    ).value = "G002"
    for column, value in enumerate(
        ["G002", "men", "A", "小组赛", 1, "A2", "A1"],
        start=1,
    ):
        workbook["比赛清单"].cell(GAME_START_ROW + 1, column, value)

    batch = validate_workbook(setup, workbook, tmp_path)
    codes = set(batch.issues.values_list("code", flat=True))
    assert "FORMULA_FORBIDDEN" in codes
    assert "CAPACITY_EXCEEDED" in codes


def test_game_list_and_grid_consistency_rules_are_reported(tmp_path):
    setup = import_setup()
    workbook = make_workbook(setup)
    games = workbook["比赛清单"]
    games.cell(GAME_START_ROW, 7, "A1")
    original = grid_cell(
        workbook,
        target_date=setup["season"].starts_on,
        period_code=setup["period"].code,
        venue_name=setup["venues"][0].name,
    )
    duplicate = grid_cell(
        workbook,
        target_date=setup["season"].starts_on,
        period_code=setup["period"].code,
        venue_name=setup["venues"][1].name,
    )
    duplicate.value = original.value

    batch = validate_workbook(setup, workbook, tmp_path)
    codes = set(batch.issues.values_list("code", flat=True))
    assert {"SAME_PARTICIPANT", "DUPLICATE_GAME_PLACEMENT"} <= codes


def test_same_slot_cannot_play_two_new_games_at_the_same_time(tmp_path):
    setup = import_setup()
    workbook = make_workbook(setup)
    for column, value in enumerate(
        ["G002", "men", "A", "小组赛", 2, "A2", "A1"],
        start=1,
    ):
        workbook["比赛清单"].cell(GAME_START_ROW + 1, column, value)
    grid_cell(
        workbook,
        target_date=setup["season"].starts_on,
        period_code=setup["period"].code,
        venue_name=setup["venues"][1].name,
    ).value = "G002"

    batch = validate_workbook(setup, workbook, tmp_path)
    assert batch.issues.filter(code="PARTICIPANT_TIME_CONFLICT").exists()


def test_confirm_revalidates_and_rolls_back_after_concurrent_game_creation(tmp_path):
    setup = import_setup()
    workbook = make_workbook(setup)
    batch = validate_workbook(setup, workbook, tmp_path)
    manual_slot_home = ParticipantSlot.objects.create(
        division=setup["division"],
        code="manual-home",
        label="手工主方",
    )
    manual_slot_away = ParticipantSlot.objects.create(
        division=setup["division"],
        code="manual-away",
        label="手工客方",
    )
    Game.objects.create(
        season=setup["season"],
        division=setup["division"],
        code="G001",
        date=setup["season"].starts_on + timedelta(days=2),
        period=setup["period"],
        start_time=setup["period"].start_time,
        venue_name=setup["venues"][1].name,
        home_slot=manual_slot_home,
        away_slot=manual_slot_away,
    )

    with override_settings(MEDIA_ROOT=tmp_path), pytest.raises(ScheduleImportError) as denied:
        confirm_schedule_import(
            actor=setup["superadmin"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
        )
    assert denied.value.code == "REVALIDATION_FAILED"
    assert not CompetitionGroup.objects.filter(code="A").exists()
    assert not ParticipantSlot.objects.filter(code__in=["A1", "A2"]).exists()
    batch.refresh_from_db()
    assert batch.status == ScheduleImportBatch.Status.VALIDATED


def test_reset_removes_only_confirmed_import_objects(tmp_path):
    setup = import_setup()
    batch = confirm_workbook(setup, make_workbook(setup), tmp_path)
    manual_group = CompetitionGroup.objects.create(
        division=setup["division"],
        code="manual",
        name="手工小组",
        sort_order=9,
    )
    preview = schedule_import_reset_preview(
        actor=setup["superadmin"],
        season=setup["season"],
    )
    assert preview["eligible"] is True
    assert preview["game_count"] == 1

    result = reset_schedule_imports(
        actor=setup["superadmin"],
        season_id=setup["season"].id,
        expected_season_version=setup["season"].version,
        season_name=setup["season"].name,
    )

    batch.refresh_from_db()
    assert result["game_count"] == 1
    assert batch.status == ScheduleImportBatch.Status.ROLLED_BACK
    assert not Game.objects.exists()
    assert not ParticipantSlot.objects.exists()
    assert CompetitionGroup.objects.filter(id=manual_group.id).exists()
    assert AdminAuditLog.objects.filter(action="SCHEDULE_IMPORTS_ROLLED_BACK").exists()


def test_reset_is_blocked_by_draw_assignment_without_partial_changes(tmp_path):
    setup = import_setup()
    batch = confirm_workbook(setup, make_workbook(setup), tmp_path)
    slot = ParticipantSlot.objects.get(code="A1")
    DrawAssignment.objects.create(
        season=setup["season"],
        slot=slot,
        team=setup["teams"][0],
        assigned_by=setup["superadmin"],
    )

    preview = schedule_import_reset_preview(
        actor=setup["superadmin"],
        season=setup["season"],
    )
    assert preview["eligible"] is False
    assert "DRAW_ASSIGNMENTS_EXIST" in {
        blocker["code"] for blocker in preview["blockers"]
    }
    with pytest.raises(ScheduleImportError) as denied:
        reset_schedule_imports(
            actor=setup["superadmin"],
            season_id=setup["season"].id,
            expected_season_version=setup["season"].version,
            season_name=setup["season"].name,
        )
    assert denied.value.code == "RESET_BLOCKED"
    assert Game.objects.filter(created_by_import_batch=batch).count() == 1
    assert batch.created_slots.count() == 2
    assert batch.created_groups.count() == 1


def test_reset_requires_setup_state_and_exact_season_name(tmp_path):
    setup = import_setup()
    confirm_workbook(setup, make_workbook(setup), tmp_path)
    with pytest.raises(ScheduleImportError) as mismatch:
        reset_schedule_imports(
            actor=setup["superadmin"],
            season_id=setup["season"].id,
            expected_season_version=setup["season"].version,
            season_name="错误赛季",
        )
    assert mismatch.value.code == "SEASON_NAME_MISMATCH"

    setup["season"].status = Season.Status.PRE_DRAW_PUBLIC
    setup["season"].save()
    with pytest.raises(ScheduleImportError) as wrong_state:
        reset_schedule_imports(
            actor=setup["superadmin"],
            season_id=setup["season"].id,
            expected_season_version=setup["season"].version,
            season_name=setup["season"].name,
        )
    assert wrong_state.value.code == "RESET_BLOCKED"
    assert Game.objects.count() == 1


@pytest.mark.parametrize(
    ("scenario", "expected_code"),
    [
        ("started", "IMPORTED_GAMES_STARTED"),
        ("result", "IMPORTED_GAMES_HAVE_RESULTS"),
        ("reservation", "RESERVATIONS_EXIST"),
        ("request", "RESCHEDULE_REQUESTS_EXIST"),
        ("media", "GAME_MEDIA_EXIST"),
    ],
)
def test_reset_blockers_leave_imported_objects_unchanged(
    tmp_path,
    scenario,
    expected_code,
):
    setup = import_setup()
    batch = confirm_workbook(setup, make_workbook(setup), tmp_path)
    game = Game.objects.get(code="G001")
    now = timezone.now()
    if scenario == "started":
        game.date = timezone.localdate() - timedelta(days=1)
        game.save(update_fields=["date", "updated_at"])
    elif scenario == "result":
        game.home_score = 32
        game.away_score = 28
        game.status = Game.Status.COMPLETED
        game.save(update_fields=["home_score", "away_score", "status", "updated_at"])
    elif scenario in {"reservation", "request"}:
        reservation = SlotReservation.objects.create(
            season=setup["season"],
            date=setup["season"].starts_on + timedelta(days=4),
            period=setup["period"],
            venue=setup["venues"][1],
            venue_name=setup["venues"][1].name,
        )
        if scenario == "request":
            request = RescheduleRequest.objects.create(
                game=game,
                requester_team=setup["teams"][0],
                requester=setup["superadmin"],
                request_type=RescheduleRequest.RequestType.SAME_WEEK,
                target_date=reservation.date,
                target_period=reservation.period,
                target_start_time=reservation.period.start_time,
                target_venue_name=reservation.venue_name,
                reservation=reservation,
                original_game_snapshot={},
                game_version_at_submit=game.version,
                submit_deadline=now + timedelta(days=1),
                confirmation_deadline=now + timedelta(days=2),
            )
            game.active_reschedule_request = request
            game.save(update_fields=["active_reschedule_request", "updated_at"])
    elif scenario == "media":
        GameMediaAsset.objects.create(
            game=game,
            kind=GameMediaAsset.Kind.GAME_PHOTO,
            file_key="tests/import-reset/game-photo.jpg",
            original_filename="game-photo.jpg",
            mime_type="image/jpeg",
            file_sha256="b" * 64,
            byte_size=128,
            width=640,
            height=480,
            uploaded_by=setup["superadmin"],
        )

    preview = schedule_import_reset_preview(
        actor=setup["superadmin"],
        season=setup["season"],
        now=now,
    )
    assert expected_code in {item["code"] for item in preview["blockers"]}
    with pytest.raises(ScheduleImportError) as denied:
        reset_schedule_imports(
            actor=setup["superadmin"],
            season_id=setup["season"].id,
            expected_season_version=setup["season"].version,
            season_name=setup["season"].name,
            now=now,
        )
    assert denied.value.code == "RESET_BLOCKED"
    assert Game.objects.filter(created_by_import_batch=batch).count() == 1
    assert ParticipantSlot.objects.filter(created_by_import_batch=batch).count() == 2
    assert CompetitionGroup.objects.filter(created_by_import_batch=batch).count() == 1


def test_non_superadmin_cannot_upload_or_reset(tmp_path):
    setup = import_setup()
    ordinary_admin = Account.objects.create_user(
        username="schedule-v2-admin",
        password="test-password",
        role=Account.Role.ADMIN,
    )
    workbook = make_workbook(setup)
    with pytest.raises(ScheduleImportError) as upload_denied:
        with override_settings(MEDIA_ROOT=tmp_path):
            validate_schedule_upload(
                actor=ordinary_admin,
                season=setup["season"],
                content=workbook_bytes(workbook),
                source_name="schedule.xlsx",
            )
    assert upload_denied.value.code == "PERMISSION_DENIED"
    assert not ScheduleImportBatch.objects.exists()
    with pytest.raises(ScheduleImportError) as reset_denied:
        schedule_import_reset_preview(actor=ordinary_admin, season=setup["season"])
    assert reset_denied.value.code == "PERMISSION_DENIED"
    assert ImportIssue.objects.count() == 0


def test_special_arrangement_keeps_slot_but_accepts_actual_time_and_free_venue(tmp_path):
    setup = import_setup()
    workbook = make_workbook(setup)
    standard_cell = grid_cell(
        workbook,
        target_date=setup["season"].starts_on,
        period_code=setup["period"].code,
        venue_name=setup["venues"][0].name,
    )
    standard_cell.value = None
    special = workbook["特殊安排"]
    for column, value in enumerate(
        [
            "G001",
            setup["season"].starts_on,
            setup["period"].code,
            "13:05",
            "校外协商场地",
        ],
        start=1,
    ):
        special.cell(SPECIAL_START_ROW, column, value)

    batch = validate_workbook(setup, workbook, tmp_path)

    assert batch.summary["error_count"] == 0
    preview = batch.summary["games"][0]
    assert preview["period_code"] == "P1"
    assert preview["nominal_start_time"] == "12:50"
    assert preview["start_time"] == "13:05"
    assert preview["venue_name"] == "校外协商场地"
    assert preview["standard_venue_id"] is None


def test_game_must_appear_exactly_once_across_grid_and_special_sheet(tmp_path):
    setup = import_setup()
    workbook = make_workbook(setup)
    special = workbook["特殊安排"]
    for column, value in enumerate(
        [
            "G001",
            setup["season"].starts_on,
            setup["period"].code,
            "13:05",
            "校外协商场地",
        ],
        start=1,
    ):
        special.cell(SPECIAL_START_ROW, column, value)

    batch = validate_workbook(setup, workbook, tmp_path)

    assert batch.issues.filter(code="DUPLICATE_GAME_PLACEMENT").exists()
