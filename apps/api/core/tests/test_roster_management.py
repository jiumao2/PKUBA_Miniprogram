from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import pytest
from django.db import IntegrityError, transaction
from django.test import override_settings
from django.utils import timezone
from openpyxl import load_workbook

from core.models import (
    Account,
    AdminAuditLog,
    Division,
    RosterImportBatch,
    RosterImportIssue,
    RosterPlayer,
    Season,
    SeasonLeaderBinding,
    Team,
)
from core.services.roster_management import (
    HEADERS,
    RosterManagementError,
    confirm_roster_import,
    create_team_with_roster,
    generate_roster_template,
    preview_team_change,
    resolve_roster_import,
    roster_import_readiness,
    save_team_roster,
    validate_roster_upload,
)
from core.services.schedule_imports_v3 import schedule_import_readiness

pytestmark = pytest.mark.django_db


def _setup(*, four_divisions: bool = False):
    today = timezone.localdate()
    season = Season.objects.create(
        name="名单测试赛季",
        competition_type=Season.CompetitionType.PKU_CUP,
        year=today.year,
        status=Season.Status.SETUP,
        starts_on=today,
        ends_on=today.replace(year=today.year + 1),
    )
    divisions = [
        Division.objects.create(
            season=season,
            code="men-a",
            name="男甲",
            gender=Division.Gender.MEN,
            sort_order=1,
        ),
        Division.objects.create(
            season=season,
            code="women-a",
            name="女甲",
            gender=Division.Gender.WOMEN,
            sort_order=3,
        ),
    ]
    if four_divisions:
        divisions.insert(
            1,
            Division.objects.create(
                season=season,
                code="men-b",
                name="男乙",
                gender=Division.Gender.MEN,
                sort_order=2,
            ),
        )
        divisions.append(
            Division.objects.create(
                season=season,
                code="women-b",
                name="女乙",
                gender=Division.Gender.WOMEN,
                sort_order=4,
            )
        )
    actor = Account.objects.create_user(
        username="roster-superadmin",
        password="test-password",
        role=Account.Role.SUPERADMIN,
    )
    return {"season": season, "divisions": divisions, "actor": actor}


def _workbook(setup, rows_by_sheet: dict[str, list[tuple[str, str, object]]]) -> bytes:
    workbook = load_workbook(BytesIO(generate_roster_template(setup["season"])))
    for sheet_name, rows in rows_by_sheet.items():
        sheet = workbook[sheet_name]
        for row_number, values in enumerate(rows, start=2):
            for column, value in enumerate(values, start=1):
                sheet.cell(row_number, column, value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _upload(setup, tmp_path, content: bytes):
    with override_settings(MEDIA_ROOT=tmp_path):
        return validate_roster_upload(
            actor=setup["actor"],
            season=setup["season"],
            content=content,
            source_name="roster.xlsx",
        )


def test_dynamic_roster_template_has_instruction_and_text_number_columns():
    setup = _setup(four_divisions=True)
    workbook = load_workbook(BytesIO(generate_roster_template(setup["season"])))

    assert workbook.sheetnames == ["填写说明", "男甲", "男乙", "女甲", "女乙"]
    assert workbook["填写说明"]["B4"].value == "1.0.0"
    for title in workbook.sheetnames[1:]:
        sheet = workbook[title]
        assert [sheet.cell(1, column).value for column in range(1, 4)] == HEADERS
        assert sheet.sheet_view.showGridLines is False
        assert sheet.freeze_panes == "A2"
        assert sheet["C2"].number_format == "@"
        assert all(sheet.cell(2, column).value is None for column in range(1, 4))
        assert len(sheet.data_validations.dataValidation) == 1


def test_upload_audits_required_duplicates_jersey_near_names_and_cross_division(tmp_path):
    setup = _setup(four_divisions=True)
    content = _workbook(
        setup,
        {
            "男甲": [
                ("信息科学技术学院", "张三", "00"),
                ("信息科学技术学院", "李四", "00"),
                ("信息科学技术学院队", "王五", "-121"),
                ("信息科学技术学院队", "王五", "8"),
                ("缺球员队", "", "9"),
            ],
            "男乙": [("信息科学技术学院", "赵六", "12")],
        },
    )
    batch = _upload(setup, tmp_path, content)
    codes = set(batch.issues.values_list("code", flat=True))

    assert batch.status == RosterImportBatch.Status.VALIDATED
    assert {
        "INVALID_JERSEY_NUMBER",
        "DUPLICATE_JERSEY_NUMBER",
        "DUPLICATE_PLAYER_NAME",
        "PLAYER_REQUIRED",
        "SIMILAR_TEAM_NAMES",
        "TEAM_DUPLICATE_ACROSS_DIVISIONS",
    } <= codes
    duplicate_jersey = batch.issues.get(code="DUPLICATE_JERSEY_NUMBER")
    assert duplicate_jersey.severity == RosterImportIssue.Severity.ERROR
    assert batch.summary["error_count"] >= 1


def test_duplicate_active_jersey_blocks_roster_confirmation_without_writes(tmp_path):
    setup = _setup()
    content = _workbook(
        setup,
        {"男甲": [("号码冲突球队", "张三", "7"), ("号码冲突球队", "李四", "7")]},
    )
    with override_settings(MEDIA_ROOT=tmp_path):
        batch = _upload(setup, tmp_path, content)
        with pytest.raises(RosterManagementError) as blocked:
            confirm_roster_import(
                actor=setup["actor"],
                batch_id=batch.id,
                expected_season_version=setup["season"].version,
                warnings_acknowledged=True,
            )

    assert blocked.value.code == "DUPLICATE_JERSEY_NUMBER"
    assert not Team.objects.filter(season=setup["season"]).exists()
    assert not RosterPlayer.objects.filter(team__season=setup["season"]).exists()
    setup["season"].refresh_from_db()
    assert setup["season"].version == 1


def test_name_resolution_reaudits_without_silent_merge(tmp_path):
    setup = _setup()
    content = _workbook(
        setup,
        {
            "男甲": [
                ("信息科学技术学院", "张三", "1"),
                ("信息学院", "李四", "2"),
            ]
        },
    )
    with override_settings(MEDIA_ROOT=tmp_path):
        batch = _upload(setup, tmp_path, content)
        options = batch.summary["name_resolutions"]
        target = next(item for item in options if item["source_name"] == "信息学院")
        batch = resolve_roster_import(
            actor=setup["actor"],
            batch_id=batch.id,
            resolutions={target["key"]: "信息科学技术学院"},
        )

    assert batch.summary["team_count"] == 1
    assert batch.summary["teams"][0]["player_count"] == 2
    assert not batch.issues.filter(code="SIMILAR_TEAM_NAMES").exists()


def test_confirm_replaces_unreferenced_setup_snapshot_preserves_00_and_closes_reimport(tmp_path):
    setup = _setup()
    old_team = Team.objects.create(
        season=setup["season"], division=setup["divisions"][0], name="待替换旧队"
    )
    RosterPlayer.objects.create(team=old_team, name="旧球员", jersey_number="9")
    content = _workbook(
        setup,
        {
            "男甲": [("新球队", "张三", "00"), ("新球队", "李四", "")],
            "女甲": [("新女队", "王五", "7")],
        },
    )
    with override_settings(MEDIA_ROOT=tmp_path):
        batch = _upload(setup, tmp_path, content)
        confirmed = confirm_roster_import(
            actor=setup["actor"],
            batch_id=batch.id,
            expected_season_version=setup["season"].version,
            warnings_acknowledged=True,
        )

    assert confirmed.status == RosterImportBatch.Status.CONFIRMED
    assert not Team.objects.filter(id=old_team.id).exists()
    assert set(Team.objects.filter(season=setup["season"]).values_list("name", flat=True)) == {
        "新球队",
        "新女队",
    }
    assert RosterPlayer.objects.get(name="张三").jersey_number == "00"
    assert RosterPlayer.objects.get(name="李四").jersey_number == ""
    setup["season"].refresh_from_db()
    assert setup["season"].version == 2
    assert roster_import_readiness(setup["season"])["ready"] is False
    assert AdminAuditLog.objects.filter(action="roster.import.confirm").exists()
    assert schedule_import_readiness(setup["season"])["team_count"] == 2
    with override_settings(MEDIA_ROOT=tmp_path), pytest.raises(
        RosterManagementError, match="已经确认"
    ):
        second = _upload(setup, tmp_path, content)
        confirm_roster_import(
            actor=setup["actor"],
            batch_id=second.id,
            expected_season_version=setup["season"].version,
            warnings_acknowledged=True,
        )


def test_confirm_rolls_back_when_team_has_downstream_reference(tmp_path):
    setup = _setup()
    team = Team.objects.create(
        season=setup["season"], division=setup["divisions"][0], name="已绑定球队"
    )
    leader = Account.objects.create_user(username="leader", password="test-password")
    SeasonLeaderBinding.objects.create(season=setup["season"], team=team, account=leader)
    content = _workbook(setup, {"男甲": [("新球队", "张三", "1")]})
    with override_settings(MEDIA_ROOT=tmp_path):
        batch = _upload(setup, tmp_path, content)
        with pytest.raises(RosterManagementError, match="领队绑定"):
            confirm_roster_import(
                actor=setup["actor"],
                batch_id=batch.id,
                expected_season_version=setup["season"].version,
                warnings_acknowledged=True,
            )
    assert Team.objects.filter(id=team.id).exists()
    assert not Team.objects.filter(name="新球队").exists()


def test_online_add_and_published_maintenance_use_stable_ids_versions_and_preview():
    setup = _setup()
    team = create_team_with_roster(
        actor=setup["actor"],
        season=setup["season"],
        division_id=setup["divisions"][0].id,
        name="稳定球队",
        players=[{"name": "张三", "jersey_number": "00"}],
        expected_season_version=setup["season"].version,
    )
    player = team.roster.get()
    setup["season"].status = Season.Status.PUBLISHED
    setup["season"].save()
    payload = {
        "expected_team_version": team.version,
        "name": "稳定球队新标准名",
        "active": True,
        "players": [
            {
                "id": str(player.id),
                "expected_version": player.version,
                "name": "张三",
                "jersey_number": "00",
                "eligible": True,
                "active": True,
            },
            {"name": "李四", "jersey_number": "8", "eligible": True, "active": True},
        ],
    }
    preview = preview_team_change(actor=setup["actor"], team=team, payload=payload)
    assert preview["requires_confirmation"] is True
    changed_after_preview = {
        **payload,
        "players": [*payload["players"][:-1], {**payload["players"][-1], "jersey_number": "9"}],
    }
    with pytest.raises(RosterManagementError, match="重新预览"):
        save_team_roster(
            actor=setup["actor"],
            team_id=team.id,
            payload=changed_after_preview,
            maintenance_token=preview["maintenance_token"],
        )
    with pytest.raises(RosterManagementError, match="重新预览"):
        save_team_roster(actor=setup["actor"], team_id=team.id, payload=payload)
    updated = save_team_roster(
        actor=setup["actor"],
        team_id=team.id,
        payload=payload,
        maintenance_token=preview["maintenance_token"],
    )
    assert updated.id == team.id
    assert updated.name == "稳定球队新标准名"
    assert updated.version == team.version + 1
    assert updated.roster.filter(name="李四", jersey_number="8").exists()
    assert AdminAuditLog.objects.filter(action="roster.team.save", object_id=team.id).exists()


def test_active_jersey_numbers_are_unique_in_service_and_database():
    setup = _setup()
    with pytest.raises(RosterManagementError) as duplicate:
        create_team_with_roster(
            actor=setup["actor"],
            season=setup["season"],
            division_id=setup["divisions"][0].id,
            name="号码冲突球队",
            players=[
                {"name": "张三", "jersey_number": "7", "active": True},
                {"name": "李四", "jersey_number": "7", "active": True},
            ],
            expected_season_version=setup["season"].version,
        )
    assert duplicate.value.code == "DUPLICATE_JERSEY_NUMBER"

    team = Team.objects.create(
        season=setup["season"],
        division=setup["divisions"][0],
        name="数据库约束球队",
    )
    RosterPlayer.objects.create(team=team, name="王一", jersey_number="8", active=True)
    RosterPlayer.objects.create(team=team, name="王二", jersey_number="8", active=False)
    with pytest.raises(IntegrityError), transaction.atomic():
        RosterPlayer.objects.create(team=team, name="王三", jersey_number="8", active=True)


REFERENCE_PATH = Path(
    os.environ.get(
        "PKUBA_ROSTER_REFERENCE_XLSX",
        r"C:\Users\jiumao\Desktop\ScoresheetReader\test\球员名单.xlsx",
    )
)


@pytest.mark.skipif(not REFERENCE_PATH.exists(), reason="local reference workbook unavailable")
def test_local_reference_workbook_is_audited_without_confirmation(tmp_path):
    setup = _setup(four_divisions=True)
    source = load_workbook(REFERENCE_PATH, data_only=False)
    template = load_workbook(BytesIO(generate_roster_template(setup["season"])))
    for source_title, target_title in zip(
        source.sheetnames, ["男甲", "男乙", "女甲", "女乙"], strict=True
    ):
        source_sheet = source[source_title]
        target_sheet = template[target_title]
        for row in source_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            target_sheet.append(row)
    output = BytesIO()
    template.save(output)
    batch = _upload(setup, tmp_path, output.getvalue())
    codes = set(batch.issues.values_list("code", flat=True))

    assert "TEAM_DUPLICATE_ACROSS_DIVISIONS" in codes
    assert "INVALID_JERSEY_NUMBER" in codes
    assert "DUPLICATE_JERSEY_NUMBER" in codes
    assert "SIMILAR_TEAM_NAMES" in codes
    assert batch.status == RosterImportBatch.Status.VALIDATED
    assert not Team.objects.filter(created_by_roster_import_batch=batch).exists()
