from __future__ import annotations

import hashlib
import hmac
import json
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import PurePath
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import IntegrityError, transaction
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from core.models import (
    Account,
    AdminAuditLog,
    Game,
    ImportIssue,
    Period,
    PeriodCapacity,
    ScheduleImportBatch,
    ScheduleSlotLock,
    Season,
    SlotReservation,
    Venue,
)

TEMPLATE_VERSION = "1.0.0"
EXPECTED_SHEETS = ["填写说明", "赛程网格", "模板元数据"]
GRID_HEADER_ROW = 6
GRID_START_ROW = 7
VENUE_START_COLUMN = 5
METADATA_KEY_COLUMN = 10
METADATA_VALUE_COLUMN = 11
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBERS = 2_000


class ScheduleImportError(Exception):
    def __init__(self, message: str, code: str = "SCHEDULE_IMPORT_INVALID"):
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class ParsedIssue:
    severity: str
    code: str
    message: str
    cell: str = ""
    context: dict[str, object] | None = None

    def model_kwargs(self) -> dict[str, object]:
        values = asdict(self)
        values["context"] = values["context"] or {}
        return values


@dataclass(frozen=True)
class GridRow:
    date: date
    period_id: str
    period_code: str
    period_name: str
    capacity: int


def _require_superadmin(actor: Account) -> None:
    if not actor.is_pkuba_superadmin:
        raise ScheduleImportError("只有超级管理员可以导入赛程。", "PERMISSION_DENIED")


def _games(season: Season):
    return (
        Game.objects.filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related(
            "division",
            "group",
            "period",
            "venue",
            "home_team",
            "away_team",
            "home_slot",
            "away_slot",
            "active_reschedule_request",
        )
        .order_by("code")
    )


def _venues(season: Season) -> list[Venue]:
    return list(Venue.objects.filter(season=season, active=True).order_by("sort_order", "name"))


def _periods(season: Season) -> list[Period]:
    return list(Period.objects.filter(season=season).order_by("sort_order", "start_time"))


def _grid_rows(season: Season) -> list[GridRow]:
    periods = _periods(season)
    capacities = {
        (item.weekday, str(item.period_id)): item.capacity
        for item in PeriodCapacity.objects.filter(season=season).select_related("period")
    }
    rows: list[GridRow] = []
    current = season.starts_on
    while current <= season.ends_on:
        for period in periods:
            capacity = capacities.get((current.weekday(), str(period.id)), 0)
            if capacity > 0:
                rows.append(
                    GridRow(
                        date=current,
                        period_id=str(period.id),
                        period_code=period.code,
                        period_name=period.name,
                        capacity=capacity,
                    )
                )
        current += timedelta(days=1)
    return rows


def _participant_reference(game: Game, side: str) -> str:
    team_id = getattr(game, f"{side}_team_id")
    slot_id = getattr(game, f"{side}_slot_id")
    if team_id:
        return f"team:{team_id}"
    return f"slot:{slot_id}"


def _template_payload(season: Season) -> dict[str, object]:
    games = list(_games(season))
    venues = _venues(season)
    periods = _periods(season)
    capacities = list(
        PeriodCapacity.objects.filter(season=season)
        .select_related("period")
        .order_by("weekday", "period__sort_order", "period__code")
    )
    return {
        "template_version": TEMPLATE_VERSION,
        "season": {
            "id": str(season.id),
            "version": season.version,
            "starts_on": season.starts_on.isoformat(),
            "ends_on": season.ends_on.isoformat(),
        },
        "games": [
            {
                "id": str(game.id),
                "version": game.version,
                "code": game.code,
                "division": game.division.code,
                "stage": game.stage,
                "round": game.round_number,
                "home": _participant_reference(game, "home"),
                "away": _participant_reference(game, "away"),
            }
            for game in games
        ],
        "periods": [
            {
                "id": str(period.id),
                "code": period.code,
                "start_time": period.start_time.isoformat(),
                "sort_order": period.sort_order,
            }
            for period in periods
        ],
        "venues": [
            {
                "id": str(venue.id),
                "code": venue.code,
                "sort_order": venue.sort_order,
            }
            for venue in venues
        ],
        "capacities": [
            {
                "weekday": item.weekday,
                "period_id": str(item.period_id),
                "capacity": item.capacity,
            }
            for item in capacities
        ],
    }


def _sign_payload(payload: dict[str, object]) -> str:
    signing_key = getattr(settings, "SCHEDULE_TEMPLATE_SIGNING_KEY", settings.SECRET_KEY)
    message = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode()
    return hmac.new(str(signing_key).encode(), message, hashlib.sha256).hexdigest()


def _set_metadata(workbook: Workbook, values: dict[str, object]) -> None:
    sheet = workbook["模板元数据"]
    sheet.cell(1, METADATA_KEY_COLUMN, "PKUBA_KEY")
    sheet.cell(1, METADATA_VALUE_COLUMN, "PKUBA_VALUE")
    for index, (key, value) in enumerate(values.items(), start=2):
        sheet.cell(index, METADATA_KEY_COLUMN, key)
        sheet.cell(index, METADATA_VALUE_COLUMN, value)


def _worksheet_styles(workbook: Workbook) -> None:
    red = "C91F26"
    ink = "171614"
    canvas = "F2F0EB"
    white = "FFFFFF"
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
    for sheet_name in ("填写说明", "赛程网格"):
        sheet = workbook[sheet_name]
        sheet.freeze_panes = "A3" if sheet_name == "填写说明" else "E7"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
    instructions = workbook["填写说明"]
    instructions.merge_cells("A1:H2")
    instructions["A1"] = "PKUBA 赛程离线填写模板"
    instructions["A1"].font = Font(color=white, bold=True, size=20)
    instructions["A1"].fill = PatternFill("solid", fgColor=red)
    instructions["A1"].alignment = Alignment(vertical="center")
    for row in instructions["A1:H2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=red)
    schedule = workbook["赛程网格"]
    schedule.merge_cells("A1:K2")
    schedule["A1"] = "PKUBA 赛程网格 · 只填写场地列"
    schedule["A1"].font = Font(color=white, bold=True, size=18)
    for row in schedule["A1:K2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=red)
    for cell in schedule[GRID_HEADER_ROW]:
        if cell.value:
            cell.fill = PatternFill("solid", fgColor=ink)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in range(1, schedule.max_column + 1):
        schedule.column_dimensions[get_column_letter(column)].width = 16
    schedule.column_dimensions["A"].width = 13
    schedule.column_dimensions["B"].width = 9
    for row in range(GRID_START_ROW, schedule.max_row + 1):
        for column in range(1, 5):
            schedule.cell(row, column).fill = PatternFill("solid", fgColor=canvas)
    metadata = workbook["模板元数据"]
    for cell in metadata[1]:
        if cell.value:
            cell.fill = PatternFill("solid", fgColor=ink)
            cell.font = Font(color=white, bold=True)


def generate_schedule_template(season: Season) -> bytes:
    games = list(_games(season))
    venues = _venues(season)
    rows = _grid_rows(season)
    if not games:
        raise ScheduleImportError("赛季尚无待排比赛，不能生成模板。", "NO_GAMES")
    if not venues:
        raise ScheduleImportError("赛季尚无可用场地，不能生成模板。", "NO_VENUES")
    if not rows:
        raise ScheduleImportError("赛季尚无开放容量，不能生成模板。", "NO_CAPACITY")
    if any(row.capacity > len(venues) for row in rows):
        raise ScheduleImportError("时段容量大于可用场地数量。", "CAPACITY_EXCEEDS_VENUES")

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    schedule = workbook.create_sheet("赛程网格")
    metadata = workbook.create_sheet("模板元数据")

    instructions_rows = [
        ("模板版本", TEMPLATE_VERSION),
        ("赛季标识", str(season.id)),
        ("赛季名称", season.name),
        ("上传原则", "上传只进入暂存批次；服务器复核并确认差异后才写入正式赛程。"),
        ("填写规则", "只在赛程网格的场地列选择比赛代码，不要填写球队名称。"),
        ("完整性", "每个比赛代码必须且只能出现一次；漏填、重复或额外代码均阻止确认。"),
        ("锁定", "上传后由管理员显式决定每场 leaderAdjustable；末轮和决赛会强提醒。"),
        ("安全", "禁止宏、外部链接和用户区域公式；本地自查不替代服务端校验。"),
    ]
    for row_index, values in enumerate(instructions_rows, start=4):
        instructions.cell(row_index, 1, values[0])
        instructions.cell(row_index, 2, values[1])
        instructions.cell(row_index, 1).font = Font(bold=True)
        instructions.cell(row_index, 2).alignment = Alignment(wrap_text=True)
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 72

    schedule["A4"] = "模板版本"
    schedule["B4"] = TEMPLATE_VERSION
    schedule["C4"] = "赛季标识"
    schedule["D4"] = str(season.id)
    headers = ["日期", "星期", "时段代码", "时段名称", *[venue.name for venue in venues]]
    for column, value in enumerate(headers, start=1):
        schedule.cell(GRID_HEADER_ROW, column, value)
    for offset, venue in enumerate(venues):
        schedule.cell(GRID_HEADER_ROW - 1, VENUE_START_COLUMN + offset, venue.code)
    schedule.row_dimensions[GRID_HEADER_ROW - 1].hidden = True

    weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    row_lookup: dict[tuple[date, str], int] = {}
    for row_index, row in enumerate(rows, start=GRID_START_ROW):
        schedule.cell(row_index, 1, row.date)
        schedule.cell(row_index, 1).number_format = "yyyy-mm-dd"
        schedule.cell(row_index, 2, weekday_names[row.date.weekday()])
        schedule.cell(row_index, 3, row.period_code)
        schedule.cell(row_index, 4, row.period_name)
        row_lookup[(row.date, row.period_code)] = row_index

    metadata_headers = [
        "比赛代码",
        "组别",
        "阶段",
        "轮次",
        "主队/签位",
        "客队/签位",
        "预期次数",
        "网格出现次数",
        "自查状态",
    ]
    for column, value in enumerate(metadata_headers, start=1):
        metadata.cell(1, column, value)
    grid_end_column = VENUE_START_COLUMN + len(venues) - 1
    grid_end_letter = get_column_letter(grid_end_column)
    for row_index, game in enumerate(games, start=2):
        metadata.cell(row_index, 1, game.code)
        metadata.cell(row_index, 2, game.division.name)
        metadata.cell(row_index, 3, game.get_stage_display())
        metadata.cell(row_index, 4, game.round_number)
        metadata.cell(row_index, 5, game.home_display)
        metadata.cell(row_index, 6, game.away_display)
        metadata.cell(row_index, 7, 1)
        metadata.cell(
            row_index,
            8,
            "=COUNTIF("
            f"'赛程网格'!$E$7:${grid_end_letter}${GRID_START_ROW + len(rows) - 1},"
            f"A{row_index})",
        )
        metadata.cell(
            row_index,
            9,
            f'=IF(H{row_index}=1,"正常",IF(H{row_index}=0,"漏填","重复"))',
        )

    game_list_end = len(games) + 1
    validation = DataValidation(
        type="list",
        formula1=f"'模板元数据'!$A$2:$A${game_list_end}",
        allow_blank=True,
    )
    validation.error = "请选择系统提供的比赛代码。"
    validation.errorTitle = "无效比赛代码"
    validation.prompt = "每场比赛必须且只能出现一次。"
    validation.promptTitle = "填写比赛代码"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    schedule.add_data_validation(validation)
    editable_range = (
        f"E{GRID_START_ROW}:{grid_end_letter}{GRID_START_ROW + len(rows) - 1}"
    )
    validation.add(editable_range)
    for row in schedule[editable_range]:
        for cell in row:
            cell.protection = Protection(locked=False)
    schedule.conditional_formatting.add(
        editable_range,
        FormulaRule(
            formula=[
                f"COUNTIF(${get_column_letter(VENUE_START_COLUMN)}$7:"
                f"${grid_end_letter}${GRID_START_ROW + len(rows) - 1},E7)>1"
            ],
            fill=PatternFill("solid", fgColor="F8D9D7"),
        ),
    )

    venue_by_id = {venue.id: index for index, venue in enumerate(venues)}
    for game in games:
        row_index = row_lookup.get((game.date, game.period.code))
        venue_offset = venue_by_id.get(game.venue_id)
        if row_index is None or venue_offset is None:
            raise ScheduleImportError(
                f"比赛 {game.code} 位于未开放日期/时段或已停用场地。",
                "GAME_OUTSIDE_GRID",
            )
        schedule.cell(row_index, VENUE_START_COLUMN + venue_offset, game.code)

    payload = _template_payload(season)
    _set_metadata(
        workbook,
        {
            "TEMPLATE_VERSION": TEMPLATE_VERSION,
            "SEASON_ID": str(season.id),
            "SIGNATURE": _sign_payload(payload),
            "GRID_ROW_COUNT": len(rows),
            "VENUE_COUNT": len(venues),
            "GAME_COUNT": len(games),
        },
    )
    _worksheet_styles(workbook)
    schedule.protection.sheet = True
    schedule.protection.password = "PKUBA"
    metadata.protection.sheet = True
    metadata.protection.password = "PKUBA"
    metadata.sheet_state = "veryHidden"
    workbook.security.lockStructure = True
    workbook.security.workbookPassword = "PKUBA"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _preflight_xlsx(content: bytes) -> None:
    if not content:
        raise ScheduleImportError("上传文件为空。", "EMPTY_FILE")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ScheduleImportError("上传文件超过 10 MB。", "FILE_TOO_LARGE")
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise ScheduleImportError("工作簿包含过多内部文件。", "XLSX_ZIP_BOMB")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise ScheduleImportError("工作簿解压后体积异常。", "XLSX_ZIP_BOMB")
            lower_names = [member.filename.lower() for member in members]
            if any("vbaproject.bin" in name for name in lower_names):
                raise ScheduleImportError("不允许上传含宏的工作簿。", "MACRO_FORBIDDEN")
            if any(name.startswith("xl/externallinks/") for name in lower_names):
                raise ScheduleImportError(
                    "不允许上传含外部链接的工作簿。",
                    "EXTERNAL_LINK_FORBIDDEN",
                )
            for member in members:
                if not member.filename.lower().endswith(".rels"):
                    continue
                relationships = archive.read(member).lower()
                has_external_target = (
                    b'targetmode="external"' in relationships
                    or b"targetmode='external'" in relationships
                )
                if has_external_target:
                    raise ScheduleImportError(
                        "不允许上传含外部链接的工作簿。",
                        "EXTERNAL_LINK_FORBIDDEN",
                    )
    except BadZipFile as error:
        raise ScheduleImportError("文件不是有效的 XLSX 工作簿。", "INVALID_XLSX") from error


def _metadata_values(workbook: Workbook) -> dict[str, object]:
    sheet = workbook["模板元数据"]
    values: dict[str, object] = {}
    for row in range(2, 20):
        key = sheet.cell(row, METADATA_KEY_COLUMN).value
        if isinstance(key, str) and key:
            values[key] = sheet.cell(row, METADATA_VALUE_COLUMN).value
    return values


def _cell_date(cell: Cell) -> date | None:
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _issue(
    code: str,
    message: str,
    *,
    severity: str = ImportIssue.Severity.ERROR,
    cell: str = "",
    context: dict[str, object] | None = None,
) -> ParsedIssue:
    return ParsedIssue(severity=severity, code=code, message=message, cell=cell, context=context)


def _slot_start(target_date: date, period: Period, timezone_name: str) -> datetime:
    return datetime.combine(target_date, period.start_time, tzinfo=ZoneInfo(timezone_name))


def _analyze_workbook(
    season: Season,
    content: bytes,
    *,
    now: datetime,
) -> tuple[list[ParsedIssue], dict[str, object]]:
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as error:
        raise ScheduleImportError("无法读取 XLSX 工作簿。", "INVALID_XLSX") from error
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ScheduleImportError("工作表名称、数量或顺序已被修改。", "SHEET_STRUCTURE_CHANGED")

    metadata_values = _metadata_values(workbook)
    expected_signature = _sign_payload(_template_payload(season))
    required_metadata = {
        "TEMPLATE_VERSION": TEMPLATE_VERSION,
        "SEASON_ID": str(season.id),
        "SIGNATURE": expected_signature,
    }
    for key, expected in required_metadata.items():
        if str(metadata_values.get(key, "")) != str(expected):
            raise ScheduleImportError(
                "模板版本、赛季标识或签名无效；请重新下载模板。",
                "TEMPLATE_SIGNATURE_INVALID",
            )
    schedule = workbook["赛程网格"]
    if schedule["B4"].value != TEMPLATE_VERSION or str(schedule["D4"].value) != str(
        season.id
    ):
        raise ScheduleImportError(
            "模板页显示的版本或赛季标识已被修改。",
            "TEMPLATE_SIGNATURE_INVALID",
        )

    games = list(_games(season))
    games_by_code = {game.code: game for game in games}
    venues = _venues(season)
    periods = {period.code: period for period in _periods(season)}
    periods_by_id = {str(period.id): period for period in periods.values()}
    rows = _grid_rows(season)
    issues: list[ParsedIssue] = []
    occurrences: dict[str, list[tuple[str, date, Period, Venue]]] = defaultdict(list)

    if metadata_values.get("GRID_ROW_COUNT") != len(rows) or metadata_values.get(
        "VENUE_COUNT"
    ) != len(venues):
        raise ScheduleImportError("模板网格规模已改变。", "GRID_STRUCTURE_CHANGED")

    for offset, venue in enumerate(venues):
        cell = schedule.cell(GRID_HEADER_ROW - 1, VENUE_START_COLUMN + offset)
        if cell.value != venue.code:
            raise ScheduleImportError("场地列已被修改或重排。", "VENUE_COLUMNS_CHANGED")

    for offset, expected_row in enumerate(rows):
        row_index = GRID_START_ROW + offset
        actual_date = _cell_date(schedule.cell(row_index, 1))
        period_code = schedule.cell(row_index, 3).value
        if actual_date != expected_row.date or period_code != expected_row.period_code:
            raise ScheduleImportError("日期或时段网格已被修改。", "GRID_ROWS_CHANGED")
        period = periods[expected_row.period_code]
        for venue_offset, venue in enumerate(venues):
            cell = schedule.cell(row_index, VENUE_START_COLUMN + venue_offset)
            if cell.data_type == "f":
                issues.append(
                    _issue(
                        "FORMULA_FORBIDDEN",
                        "场地填写区域不允许公式。",
                        cell=cell.coordinate,
                    )
                )
                continue
            if cell.value in (None, ""):
                continue
            if not isinstance(cell.value, str):
                issues.append(
                    _issue(
                        "INVALID_GAME_CODE",
                        "比赛代码必须是文本。",
                        cell=cell.coordinate,
                    )
                )
                continue
            code = cell.value.strip()
            if code not in games_by_code:
                issues.append(
                    _issue(
                        "UNKNOWN_GAME_CODE",
                        f"未知比赛代码：{code}",
                        cell=cell.coordinate,
                        context={"game_code": code},
                    )
                )
                continue
            occurrences[code].append((cell.coordinate, expected_row.date, period, venue))

    placement: dict[str, dict[str, object]] = {}
    for code in games_by_code:
        found = occurrences.get(code, [])
        if not found:
            issues.append(
                _issue(
                    "MISSING_GAME",
                    f"比赛 {code} 未填写。",
                    context={"game_code": code},
                )
            )
            continue
        if len(found) > 1:
            issues.append(
                _issue(
                    "DUPLICATE_GAME",
                    f"比赛 {code} 出现 {len(found)} 次。",
                    cell=", ".join(item[0] for item in found[:3]),
                    context={"game_code": code, "count": len(found)},
                )
            )
            continue
        cell, target_date, period, venue = found[0]
        placement[code] = {
            "date": target_date.isoformat(),
            "period_id": str(period.id),
            "period_code": period.code,
            "venue_id": str(venue.id),
            "venue_code": venue.code,
            "cell": cell,
        }

    active_reservations = list(
        SlotReservation.objects.filter(season=season, status=SlotReservation.Status.ACTIVE)
        .select_related(
            "request__game__home_team",
            "request__game__away_team",
            "request__game__home_slot",
            "request__game__away_slot",
        )
    )
    reservation_keys = {
        (item.date, str(item.period_id), str(item.venue_id)): item for item in active_reservations
    }
    slot_counts: Counter[tuple[date, str]] = Counter()
    participants_by_slot: dict[tuple[date, str], dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for code, item in placement.items():
        game = games_by_code[code]
        target_date = date.fromisoformat(str(item["date"]))
        period_id = str(item["period_id"])
        venue_id = str(item["venue_id"])
        key = (target_date, period_id)
        slot_counts[key] += 1
        reservation = reservation_keys.get((target_date, period_id, venue_id))
        if reservation:
            issues.append(
                _issue(
                    "VENUE_RESERVED",
                    f"{code} 的目标场地已被活动调赛申请预留。",
                    cell=str(item["cell"]),
                    context={"reservation_id": str(reservation.id)},
                )
            )
        references = (
            _participant_reference(game, "home"),
            _participant_reference(game, "away"),
        )
        for reference in references:
            participants_by_slot[key][reference].append(code)

    for reservation in active_reservations:
        key = (reservation.date, str(reservation.period_id))
        slot_counts[key] += 1
        reserved_game = reservation.request.game
        references = (
            _participant_reference(reserved_game, "home"),
            _participant_reference(reserved_game, "away"),
        )
        for reference in references:
            participants_by_slot[key][reference].append(f"预留:{reservation.id}")
    capacity_map = {
        (row.date, row.period_id): row.capacity
        for row in rows
    }
    for key, used in slot_counts.items():
        capacity = capacity_map.get(key, 0)
        if used > capacity:
            period_name = periods_by_id[key[1]].name
            issues.append(
                _issue(
                    "CAPACITY_EXCEEDED",
                    f"{key[0].isoformat()} / {period_name} 共占用 {used} 场，"
                    f"容量为 {capacity}。",
                    context={
                        "date": key[0].isoformat(),
                        "period_id": key[1],
                        "used": used,
                        "capacity": capacity,
                    },
                )
            )
    for key, participants in participants_by_slot.items():
        for reference, codes in participants.items():
            if len(codes) > 1:
                issues.append(
                    _issue(
                        "PARTICIPANT_TIME_CONFLICT",
                        f"同一球队或签位在同一时段参加多场比赛：{', '.join(codes)}。",
                        context={
                            "date": key[0].isoformat(),
                            "period_id": key[1],
                            "participant": reference,
                            "games": codes,
                        },
                    )
                )

    modified: list[str] = []
    unchanged: list[str] = []
    for code, item in placement.items():
        game = games_by_code[code]
        proposed = (item["date"], item["period_id"], item["venue_id"])
        current = (game.date.isoformat(), str(game.period_id), str(game.venue_id))
        if proposed == current:
            unchanged.append(code)
            continue
        modified.append(code)
        if game.active_reschedule_request_id:
            issues.append(
                _issue(
                    "ACTIVE_REQUEST_BLOCKS_CHANGE",
                    f"比赛 {code} 存在活动调赛申请，普通导入不能修改。",
                    cell=str(item["cell"]),
                )
            )
        if game.status != Game.Status.SCHEDULED:
            issues.append(
                _issue(
                    "RESULT_BLOCKS_CHANGE",
                    f"比赛 {code} 已有赛果或状态不允许导入修改。",
                    cell=str(item["cell"]),
                )
            )
        if _slot_start(game.date, game.period, season.timezone) <= now:
            issues.append(
                _issue(
                    "STARTED_GAME_BLOCKS_CHANGE",
                    f"比赛 {code} 已开始，普通导入不能修改。",
                    cell=str(item["cell"]),
                )
            )

    group_rounds: dict[str, int] = defaultdict(int)
    for game in games:
        if game.group_id and game.stage == Game.Stage.GROUP:
            group_id = str(game.group_id)
            group_rounds[group_id] = max(group_rounds[group_id], game.round_number)
    for game in games:
        if game.leader_adjustable and (
            game.stage == Game.Stage.FINAL
            or (
                game.group_id
                and game.stage == Game.Stage.GROUP
                and game.round_number == group_rounds[str(game.group_id)]
            )
        ):
            issues.append(
                _issue(
                    "REVIEW_LEADER_ADJUSTABLE",
                    f"{game.code} 是末轮或决赛，请显式确认是否禁止领队调赛。",
                    severity=ImportIssue.Severity.WARNING,
                    context={"game_code": game.code},
                )
            )

    summary: dict[str, object] = {
        "expected": len(games),
        "actual_unique": len(placement),
        "modified": sorted(modified),
        "unchanged": sorted(unchanged),
        "added": [],
        "removed": sorted(code for code in games_by_code if code not in placement),
        "assignments": placement,
    }
    return issues, summary


def validate_schedule_upload(
    *,
    actor: Account,
    season: Season,
    content: bytes,
    source_name: str,
    now: datetime | None = None,
) -> ScheduleImportBatch:
    _require_superadmin(actor)
    now = now or timezone.now()
    digest = hashlib.sha256(content).hexdigest()
    safe_name = get_valid_filename(PurePath(source_name).name) or "schedule.xlsx"
    batch = ScheduleImportBatch.objects.create(
        season=season,
        template_version=TEMPLATE_VERSION,
        file_key="",
        file_sha256=digest,
        uploaded_by=actor,
    )
    key = f"schedule-imports/{season.id}/{batch.id}/{safe_name}"
    batch.file_key = default_storage.save(key, ContentFile(content))
    try:
        issues, summary = _analyze_workbook(season, content, now=now)
    except ScheduleImportError as error:
        issues = [_issue(error.code, str(error))]
        summary = {"expected": _games(season).count(), "actual_unique": 0, "assignments": {}}
    ImportIssue.objects.bulk_create(
        [ImportIssue(batch=batch, **issue.model_kwargs()) for issue in issues]
    )
    summary["error_count"] = sum(
        issue.severity == ImportIssue.Severity.ERROR for issue in issues
    )
    summary["warning_count"] = sum(
        issue.severity == ImportIssue.Severity.WARNING for issue in issues
    )
    batch.status = ScheduleImportBatch.Status.VALIDATED
    batch.summary = summary
    batch.save(update_fields=["file_key", "status", "summary", "updated_at"])
    return batch


def _lock_schedule_slots(season: Season, assignments: dict[str, dict[str, object]]) -> None:
    keys = sorted({(item["date"], item["period_id"]) for item in assignments.values()})
    for date_text, period_id in keys:
        target_date = date.fromisoformat(str(date_text))
        try:
            with transaction.atomic():
                ScheduleSlotLock.objects.get_or_create(
                    season=season,
                    date=target_date,
                    period_id=period_id,
                )
        except IntegrityError:
            # Another transaction may create the serialization row first.
            pass
        ScheduleSlotLock.objects.select_for_update().get(
            season=season,
            date=target_date,
            period_id=period_id,
        )


@transaction.atomic
def confirm_schedule_import(
    *,
    actor: Account,
    batch_id: object,
    expected_season_version: int,
    leader_adjustable_by_game: dict[str, bool],
    now: datetime | None = None,
) -> ScheduleImportBatch:
    _require_superadmin(actor)
    now = now or timezone.now()
    batch = (
        ScheduleImportBatch.objects.select_for_update()
        .select_related("season")
        .get(id=batch_id)
    )
    if batch.status == ScheduleImportBatch.Status.CONFIRMED:
        return batch
    if batch.status != ScheduleImportBatch.Status.VALIDATED:
        raise ScheduleImportError("导入批次尚未完成校验。", "BATCH_NOT_VALIDATED")
    if batch.issues.filter(severity=ImportIssue.Severity.ERROR).exists():
        raise ScheduleImportError("导入批次仍有错误，不能确认。", "BATCH_HAS_ERRORS")

    season = Season.objects.select_for_update().get(id=batch.season_id)
    if season.version != expected_season_version:
        raise ScheduleImportError("赛季已被其他操作修改，请重新下载模板。", "VERSION_CONFLICT")
    with default_storage.open(batch.file_key, "rb") as source:
        content = source.read()
    if hashlib.sha256(content).hexdigest() != batch.file_sha256:
        raise ScheduleImportError("上传原文件校验值不一致。", "FILE_HASH_MISMATCH")

    assignments = batch.summary.get("assignments", {})
    if set(leader_adjustable_by_game) != set(assignments):
        raise ScheduleImportError(
            "必须为每场比赛显式确认是否允许领队调赛。",
            "LEADER_POLICY_INCOMPLETE",
        )
    if any(type(value) is not bool for value in leader_adjustable_by_game.values()):
        raise ScheduleImportError(
            "leaderAdjustable 必须是明确的布尔值。",
            "LEADER_POLICY_INVALID",
        )
    games = {
        game.code: game
        for game in Game.objects.select_for_update(of=("self",))
        .filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related(
            "division",
            "group",
            "period",
            "venue",
            "home_team",
            "away_team",
            "home_slot",
            "away_slot",
        )
    }
    slot_assignments = dict(assignments)
    slot_assignments.update(
        {
            f"current:{game.code}": {
                "date": game.date.isoformat(),
                "period_id": str(game.period_id),
            }
            for game in games.values()
        }
    )
    _lock_schedule_slots(season, slot_assignments)
    issues, current_summary = _analyze_workbook(season, content, now=now)
    errors = [issue for issue in issues if issue.severity == ImportIssue.Severity.ERROR]
    if errors:
        raise ScheduleImportError(
            f"确认前复核失败：{errors[0].message}",
            "REVALIDATION_FAILED",
        )
    assignments = current_summary["assignments"]
    before: dict[str, dict[str, object]] = {}
    changed_games: list[Game] = []
    changed_codes: set[str] = set()
    for code, item in assignments.items():
        game = games[code]
        before[code] = {
            "date": game.date.isoformat(),
            "period_id": str(game.period_id),
            "venue_id": str(game.venue_id),
            "leader_adjustable": game.leader_adjustable,
            "version": game.version,
        }
        if game.active_reschedule_request_id and (
            leader_adjustable_by_game[code] != game.leader_adjustable
        ):
            raise ScheduleImportError(
                f"比赛 {code} 存在活动申请，不能改变调赛政策。",
                "ACTIVE_REQUEST_BLOCKS_POLICY_CHANGE",
            )
        schedule_changed = (
            game.date.isoformat(),
            str(game.period_id),
            str(game.venue_id),
        ) != (item["date"], item["period_id"], item["venue_id"])
        policy_changed = leader_adjustable_by_game[code] != game.leader_adjustable
        if policy_changed and (
            game.status != Game.Status.SCHEDULED
            or _slot_start(game.date, game.period, season.timezone) <= now
        ):
            raise ScheduleImportError(
                f"比赛 {code} 已开始或已有赛果，普通导入不能改变调赛政策。",
                "STARTED_GAME_BLOCKS_POLICY_CHANGE",
            )
        if schedule_changed:
            changed_games.append(game)
        if schedule_changed or policy_changed:
            changed_codes.add(code)

    for game in changed_games:
        game.status = Game.Status.VOID
        game.save(update_fields=["status", "updated_at"])

    periods = {str(item.id): item for item in _periods(season)}
    venues = {str(item.id): item for item in _venues(season)}
    after: dict[str, dict[str, object]] = {}
    for code, item in assignments.items():
        game = games[code]
        if code not in changed_codes:
            after[code] = before[code]
            continue
        game.date = date.fromisoformat(str(item["date"]))
        game.period = periods[str(item["period_id"])]
        game.venue = venues[str(item["venue_id"])]
        game.leader_adjustable = leader_adjustable_by_game[code]
        game.status = Game.Status.SCHEDULED
        game.version += 1
        game.full_clean()
        game.save(
            update_fields=[
                "date",
                "period",
                "venue",
                "leader_adjustable",
                "status",
                "version",
                "updated_at",
            ]
        )
        after[code] = {
            "date": game.date.isoformat(),
            "period_id": str(game.period_id),
            "venue_id": str(game.venue_id),
            "leader_adjustable": game.leader_adjustable,
            "version": game.version,
        }

    if changed_codes:
        season.version += 1
        season.save(update_fields=["version", "updated_at"])
    batch.status = ScheduleImportBatch.Status.CONFIRMED
    batch.confirmed_at = now
    batch.summary = {
        **current_summary,
        "error_count": 0,
        "warning_count": sum(
            issue.severity == ImportIssue.Severity.WARNING for issue in issues
        ),
        "confirmed_season_version": season.version,
    }
    batch.save(update_fields=["status", "confirmed_at", "summary", "updated_at"])
    AdminAuditLog.objects.create(
        actor=actor,
        action="SCHEDULE_IMPORT_CONFIRMED",
        object_type="ScheduleImportBatch",
        object_id=batch.id,
        before=before,
        after=after,
        metadata={"file_sha256": batch.file_sha256},
    )
    return batch
