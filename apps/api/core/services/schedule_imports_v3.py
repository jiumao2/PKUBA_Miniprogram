from __future__ import annotations

import hashlib
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import PurePath
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import (
    Account,
    AdminAuditLog,
    CompetitionGroup,
    Division,
    DrawAssignment,
    Game,
    GameMediaAsset,
    ImportIssue,
    ParticipantSlot,
    Period,
    RescheduleRequest,
    ScheduleGridDraft,
    ScheduleImportBatch,
    ScheduleSlotFamily,
    ScheduleSlotLock,
    Season,
    SlotReservation,
    Team,
    Venue,
)
from core.services.schedule_capacity import effective_capacity_map

TEMPLATE_VERSION = "3.3.0"
EXPECTED_SHEETS = ["填写说明", "签位定义", "赛程网格"]
SLOT_HEADER_ROW = 5
SLOT_START_ROW = 6
GRID_TIME_ROW = 5
GRID_VENUE_ROW = 6
GRID_START_ROW = 7
GRID_WEEKDAY_COLUMN = 2
GRID_START_COLUMN = 3
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBERS = 2_000
MAX_DATA_ROWS = 5_000
MAX_GRID_COLUMNS = 64

# 这些组合只用于新空白草稿/模板的初始排版；上传和在线校验从工作簿表头识别，
# 不会要求管理员保持这 16 列，也不会把它们写成赛季资源配置。
DEFAULT_GRID_COLUMNS = (
    ("p1", "五四东一", False),
    ("p1", "五四东二", False),
    ("p1", "五四东三", False),
    ("p2", "五四东一", False),
    ("p2", "五四东二", False),
    ("p2", "五四东三", False),
    ("p3", "五四东一", False),
    ("p3", "五四东二", False),
    ("p3", "五四东三", False),
    ("p4", "邱德拔", True),
    ("p5", "五四东一", False),
    ("p5", "五四东二", False),
    ("p6", "五四东一", False),
    ("p6", "五四东二", False),
    ("p7", "邱德拔", True),
    ("p8", "五四东一", False),
)

SLOT_HEADERS = ["组别", "场次信息", "轮次", "签位代码"]
STAGE_BY_LABEL = {label: value for value, label in Game.Stage.choices}
STAGE_BY_LABEL.update({value: value for value, _label in Game.Stage.choices})
STAGE_LABEL_BY_VALUE = dict(Game.Stage.choices)
STAGE_ORDER = {value: index for index, (value, _label) in enumerate(Game.Stage.choices)}
STAGE_CODE = {
    Game.Stage.GROUP: "GRP",
    Game.Stage.ROUND_ROBIN: "RR",
    Game.Stage.KNOCKOUT: "KO",
    Game.Stage.SEMIFINAL: "SF",
    Game.Stage.FINAL: "F",
    Game.Stage.RELEGATION: "REL",
}
SLOT_PATTERN = re.compile(r"^([A-Za-z])([1-9][0-9]*)$")
MATCHUP_PATTERN = re.compile(
    r"^([A-Za-z][1-9][0-9]*)\s*[vV][sS]\s*([A-Za-z][1-9][0-9]*)"
    r"\s*(?:[（(]\s*女\s*[）)])?$"
)
WOMEN_SUFFIX_PATTERN = re.compile(r"[（(]\s*女\s*[）)]\s*$")
FINAL_ONLY_SUFFIX_PATTERN = re.compile(r"\s*[（(]\s*仅决赛\s*[）)]\s*$")
GROUP_STAGES = {Game.Stage.GROUP, Game.Stage.ROUND_ROBIN}
WEEKDAY_LABELS = ("周一", "周二", "周三", "周四", "周五", "周六", "周日")
GRID_SIDE = Side(style="thin", color="C8D5CD")
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)


class ScheduleImportError(Exception):
    def __init__(self, message: str, code: str = "SCHEDULE_IMPORT_INVALID"):
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class ParsedIssue:
    severity: str
    code: str
    message: str
    cell: str = ""
    context: dict[str, object] | None = None

    def model_kwargs(self) -> dict[str, object]:
        values = asdict(self)
        values["cell"] = str(values["cell"])[:32]
        values["context"] = values["context"] or {}
        return values


@dataclass(frozen=True)
class ParsedFamily:
    division_id: str
    division_code: str
    division_name: str
    gender: str
    stage: str
    round_number: int
    prefix: str
    slot_count: int
    sort_order: int

    @property
    def key(self) -> tuple[str, str, int, str]:
        return (self.division_code, self.stage, self.round_number, self.prefix)

    @property
    def group_code(self) -> str | None:
        return self.prefix if self.stage in GROUP_STAGES else None


@dataclass(frozen=True)
class ParsedGroup:
    division_code: str
    code: str
    name: str
    sort_order: int


@dataclass(frozen=True)
class ParsedSlot:
    division_code: str
    division_name: str
    gender: str
    stage: str
    round_number: int
    prefix: str
    number: int
    code: str
    group_code: str | None
    cell: str

    @property
    def family_key(self) -> tuple[str, str, int, str]:
        return (self.division_code, self.stage, self.round_number, self.prefix)


@dataclass(frozen=True)
class ParsedGame:
    code: str
    semantic_key: str
    division_code: str
    division_name: str
    group_code: str | None
    stage: str
    round_number: int
    home_slot_code: str
    away_slot_code: str
    target_date: date
    period_id: str
    period_code: str
    period_name: str
    start_time: str
    venue_name: str
    final_only: bool
    leader_adjustable: bool
    cell: str


@dataclass(frozen=True)
class ParsedGridColumn:
    period_id: str
    period_code: str
    period_name: str
    start_time: str
    venue_name: str
    final_only: bool
    sort_order: int
    source_column: int


@dataclass
class WorkbookAnalysis:
    families: dict[tuple[str, str, int, str], ParsedFamily]
    groups: dict[tuple[str, str], ParsedGroup]
    slots: dict[tuple[str, str], ParsedSlot]
    games: dict[str, ParsedGame]
    columns: list[ParsedGridColumn]
    summary: dict[str, object]


def _issue(
    code: str,
    message: str,
    *,
    severity: str = ImportIssue.Severity.ERROR,
    cell: str = "",
    context: dict[str, object] | None = None,
) -> ParsedIssue:
    return ParsedIssue(severity=severity, code=code, message=message, cell=cell, context=context)


def _require_superadmin(actor: Account) -> None:
    if not actor.is_pkuba_superadmin:
        raise ScheduleImportError("只有超级管理员可以导入赛程。", "PERMISSION_DENIED")


def _periods(season: Season) -> list[Period]:
    return list(Period.objects.filter(season=season).order_by("sort_order", "start_time"))


def _normalize_venue_name(value: object) -> tuple[str, bool]:
    raw = " ".join(str(value or "").strip().split())
    final_only = bool(FINAL_ONLY_SUFFIX_PATTERN.search(raw))
    venue_name = FINAL_ONLY_SUFFIX_PATTERN.sub("", raw).strip()
    return venue_name, final_only


def _default_grid_columns(season: Season) -> list[ParsedGridColumn]:
    periods = {item.code.lower(): item for item in _periods(season)}
    columns: list[ParsedGridColumn] = []
    for period_code, venue_name, final_only in DEFAULT_GRID_COLUMNS:
        period = periods.get(period_code)
        if period is None:
            continue
        sort_order = len(columns) + 1
        columns.append(
            ParsedGridColumn(
                period_id=str(period.id),
                period_code=period.code,
                period_name=period.name,
                start_time=period.start_time.strftime("%H:%M"),
                venue_name=venue_name,
                final_only=final_only,
                sort_order=sort_order,
                source_column=GRID_START_COLUMN + sort_order - 1,
            )
        )
    return columns


def _slot_families(season: Season) -> list[ScheduleSlotFamily]:
    return list(
        ScheduleSlotFamily.objects.filter(season=season)
        .select_related("division")
        .order_by("sort_order")
    )


def _family_expected_games(stage: str, slot_count: int) -> int:
    if stage in GROUP_STAGES:
        return slot_count * (slot_count - 1) // 2
    return slot_count // 2


def _calendar_dates(season: Season) -> list[date]:
    count = (season.ends_on - season.starts_on).days + 1
    return [season.starts_on + timedelta(days=offset) for offset in range(count)]


def schedule_import_readiness(season: Season) -> dict[str, object]:
    divisions = list(Division.objects.filter(season=season).order_by("sort_order", "name"))
    families = _slot_families(season)
    default_columns = _default_grid_columns(season)
    division_count = len(divisions)
    team_count = Team.objects.filter(season=season, active=True).count()
    period_count = Period.objects.filter(season=season).count()
    venue_count = Venue.objects.filter(season=season, active=True).count()
    calendar_day_count = len(_calendar_dates(season))
    expected_game_count = sum(
        _family_expected_games(item.stage, item.slot_count) for item in families
    )
    template_blockers: list[dict[str, object]] = []
    blockers: list[dict[str, object]] = []

    def block(code: str, message: str, count: int = 1) -> None:
        template_blockers.append({"code": code, "message": message, "count": count})

    if season.status != Season.Status.SETUP:
        blockers.append(
            {
                "code": "SEASON_NOT_SETUP",
                "message": "当前赛季可以下载模板，但只有准备中的赛季可以上传并确认新赛程。",
                "count": 1,
            }
        )
    for count, code, message in (
        (division_count, "NO_DIVISIONS", "赛季尚未配置组别。"),
        (team_count, "NO_TEAMS", "赛季尚未配置球队。"),
        (period_count, "NO_PERIODS", "赛季尚未配置比赛时段。"),
        (len(families), "NO_SLOT_FAMILIES", "尚未在第一步配置签位方案。"),
    ):
        if count == 0:
            block(code, message)

    prefix_owner: dict[tuple[str, str], tuple[str, str]] = {}
    group_counts: Counter[str] = Counter()
    for family in families:
        if family.division.season_id != season.id:
            block("FOREIGN_SLOT_FAMILY", "签位方案引用了其他赛季的组别。")
            continue
        if not re.fullmatch(r"[A-Za-z]", family.prefix):
            block("INVALID_SLOT_PREFIX", "签位字母必须是一个大小写敏感英文字母。")
        owner_key = (family.division.gender, family.prefix)
        owner = (family.division.code, family.stage)
        if owner_key in prefix_owner and prefix_owner[owner_key] != owner:
            block(
                "DUPLICATE_GENDER_PREFIX",
                f"同一性别内签位字母 {family.prefix} 被多个组别或阶段使用。",
            )
        prefix_owner[owner_key] = owner
        if family.stage in GROUP_STAGES:
            group_counts[family.division.code] += family.slot_count
        if family.stage == Game.Stage.SEMIFINAL and family.slot_count != 4:
            block("INVALID_SLOT_COUNT", "半决赛签位数固定为 4。")
        if family.stage == Game.Stage.FINAL and family.slot_count != 2:
            block("INVALID_SLOT_COUNT", "决赛签位数固定为 2。")
        if family.stage in {Game.Stage.KNOCKOUT, Game.Stage.RELEGATION} and (
            family.slot_count < 2 or family.slot_count % 2
        ):
            block("INVALID_SLOT_COUNT", "淘汰赛和保级赛签位数必须是不少于 2 的偶数。")
    for division in divisions:
        if division.code not in group_counts:
            continue
        active_teams = division.teams.filter(active=True).count()
        if group_counts[division.code] != active_teams:
            block(
                "SLOT_COUNT_TEAM_MISMATCH",
                f"{division.name}的小组赛/循环赛签位共 {group_counts[division.code]} 个，"
                f"必须等于当前启用球队数 {active_teams}。",
            )

    blockers.extend(template_blockers)
    return {
        "season_id": str(season.id),
        "season_version": season.version,
        "ready": not blockers,
        "template_ready": not template_blockers,
        "division_count": division_count,
        "team_count": team_count,
        "period_count": period_count,
        "venue_count": venue_count,
        "slot_family_count": len(families),
        "grid_column_count": len(default_columns),
        "calendar_day_count": calendar_day_count,
        "expected_game_count": expected_game_count,
        "existing_game_count": Game.objects.filter(season=season).count(),
        "blockers": blockers,
        "template_blockers": template_blockers,
    }


def _require_ready(season: Season) -> None:
    readiness = schedule_import_readiness(season)
    if not readiness["ready"]:
        first = readiness["blockers"][0]
        raise ScheduleImportError(str(first["message"]), str(first["code"]))


def _require_template_ready(season: Season) -> None:
    readiness = schedule_import_readiness(season)
    if not readiness["template_ready"]:
        first = readiness["template_blockers"][0]
        raise ScheduleImportError(str(first["message"]), str(first["code"]))


def _style_title(sheet, title: str, end_column: int) -> None:
    end_column = max(3, end_column)
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(color="FFFFFF", bold=True, size=18)
    cell.alignment = Alignment(vertical="center")
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=end_column):
        for item in row:
            item.fill = PatternFill("solid", fgColor="087A45")


def _style_headers(sheet, row: int, end_column: int) -> None:
    for cell in sheet[row][:end_column]:
        cell.fill = PatternFill("solid", fgColor="21543C")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID_BORDER


def generate_schedule_template(
    season: Season,
    *,
    columns: list[ParsedGridColumn] | None = None,
    cells: dict[tuple[date, int], tuple[str, bool]] | None = None,
) -> bytes:
    _require_template_ready(season)
    divisions = list(season.divisions.order_by("sort_order", "name"))
    families = _slot_families(season)
    columns = _default_grid_columns(season) if columns is None else columns
    cells = cells or {}
    if not columns:
        raise ScheduleImportError(
            "当前赛季没有可用于空白模板的默认时段，请先检查赛季时段。",
            "NO_DEFAULT_GRID_COLUMNS",
        )
    dates = _calendar_dates(season)

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    slots_sheet = workbook.create_sheet("签位定义")
    grid_sheet = workbook.create_sheet("赛程网格")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = True

    instruction_rows = [
        ("格式版本", TEMPLATE_VERSION),
        ("赛季名称", season.name),
        ("日期范围", f"{season.starts_on.isoformat()} 至 {season.ends_on.isoformat()}"),
        ("最简填写", "只需在第三页赛程网格填写对阵；第二页仅解释签位字母，上传时不检查其内容。"),
        ("女子比赛", "在对阵末尾加（女），例如 A1vsA2（女）；男、女可以复用同一字母。"),
        ("签位提示", "第二页仅解释第一步配置中的签位字母含义，无需填写或修改。"),
        ("完整性", "小组/循环必须填满全部两两对阵；其他阶段每个签位必须且只能出现一次。"),
        (
            "表头规则",
            "第三页 C 列起可增删或调整列：第 5 行时间必须匹配赛季时段，"
            "第 6 行填写场地；仅决赛列写“场地（仅决赛）”。",
        ),
        (
            "确认规则",
            "上传仅覆盖服务器草稿；显式核对并确认前不会写正式赛程，"
            "已有比赛不会更新或删除。",
        ),
        ("安全限制", "禁止宏、外部链接和单元格公式；服务器保留上传文件 SHA-256。"),
    ]
    for row_index, (label, value) in enumerate(instruction_rows, start=4):
        label_cell = instructions.cell(row_index, 1, label)
        value_cell = instructions.cell(row_index, 2, value)
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="E4F1E8")
        label_cell.alignment = Alignment(vertical="top")
        value_cell.fill = PatternFill("solid", fgColor="F8FBF9")
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in (label_cell, value_cell):
            cell.border = GRID_BORDER
    summary_start = 15
    instructions.cell(summary_start, 1, "组别").font = Font(bold=True)
    instructions.cell(summary_start, 2, "启用球队").font = Font(bold=True)
    instructions.cell(summary_start, 3, "计划签位").font = Font(bold=True)
    instructions.cell(summary_start, 4, "计划比赛").font = Font(bold=True)
    _style_headers(instructions, summary_start, 4)
    for offset, division in enumerate(divisions, start=summary_start + 1):
        related = [item for item in families if item.division_id == division.id]
        instructions.cell(offset, 1, division.name)
        instructions.cell(offset, 2, division.teams.filter(active=True).count())
        instructions.cell(offset, 3, sum(item.slot_count for item in related))
        instructions.cell(
            offset,
            4,
            sum(_family_expected_games(item.stage, item.slot_count) for item in related),
        )
        for column in range(1, 5):
            cell = instructions.cell(offset, column)
            cell.fill = PatternFill("solid", fgColor="F8FBF9")
            cell.border = GRID_BORDER
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "center",
                vertical="center",
            )
    _style_title(instructions, "PKUBA 赛程编排 V3.3 · 填写说明", 7)
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 76
    for column in ("C", "D"):
        instructions.column_dimensions[column].width = 14

    for column, header in enumerate(SLOT_HEADERS, start=1):
        slots_sheet.cell(SLOT_HEADER_ROW, column, header)
    row_index = SLOT_START_ROW
    for family in families:
        for number in range(1, family.slot_count + 1):
            slots_sheet.cell(row_index, 1, family.division.name)
            slots_sheet.cell(row_index, 2, family.get_stage_display())
            slots_sheet.cell(row_index, 3, family.round_number)
            slots_sheet.cell(row_index, 4, f"{family.prefix}{number}")
            for column in range(1, 5):
                slots_sheet.cell(row_index, column).fill = PatternFill(
                    "solid", fgColor="F2F5F3"
                )
                slots_sheet.cell(row_index, column).border = GRID_BORDER
            row_index += 1
    slots_sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    slots_sheet.cell(
        3,
        1,
        "本页由“赛季和组别”配置自动生成，仅用于说明字母含义；无需修改，上传时不读取本页业务内容。",
    )
    slots_sheet.cell(3, 1).font = Font(color="66766F", italic=True)
    slots_sheet.cell(3, 1).alignment = Alignment(wrap_text=True)
    _style_title(slots_sheet, "签位提示 · 自动生成，无需填写", 3)
    _style_headers(slots_sheet, SLOT_HEADER_ROW, 4)
    slots_sheet.freeze_panes = "A6"
    slots_sheet.column_dimensions["A"].width = 18
    slots_sheet.column_dimensions["B"].width = 18
    slots_sheet.column_dimensions["C"].width = 12
    slots_sheet.column_dimensions["D"].width = 18

    grid_end_column = GRID_START_COLUMN + len(columns) - 1
    grid_sheet.cell(GRID_TIME_ROW, 1, "日期")
    grid_sheet.cell(GRID_VENUE_ROW, 1, "yyyy-mm-dd")
    grid_sheet.cell(GRID_TIME_ROW, GRID_WEEKDAY_COLUMN, "星期")
    grid_sheet.cell(GRID_VENUE_ROW, GRID_WEEKDAY_COLUMN, "自动生成")
    for offset, column in enumerate(columns, start=GRID_START_COLUMN):
        grid_sheet.cell(GRID_TIME_ROW, offset, column.start_time)
        venue_header = (
            f"{column.venue_name}（仅决赛）" if column.final_only else column.venue_name
        )
        grid_sheet.cell(GRID_VENUE_ROW, offset, venue_header)
        fill = "F4D89B" if column.final_only else "21543C"
        font_color = "5B3A00" if column.final_only else "FFFFFF"
        for header_row in (GRID_TIME_ROW, GRID_VENUE_ROW):
            grid_sheet.cell(header_row, offset).fill = PatternFill("solid", fgColor=fill)
            grid_sheet.cell(header_row, offset).font = Font(color=font_color, bold=True)
            grid_sheet.cell(header_row, offset).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            grid_sheet.cell(header_row, offset).border = GRID_BORDER
    _style_headers(grid_sheet, GRID_TIME_ROW, GRID_WEEKDAY_COLUMN)
    _style_headers(grid_sheet, GRID_VENUE_ROW, GRID_WEEKDAY_COLUMN)
    for row_offset, target_date in enumerate(dates, start=GRID_START_ROW):
        grid_sheet.cell(row_offset, 1, target_date)
        grid_sheet.cell(row_offset, 1).number_format = "yyyy-mm-dd"
        grid_sheet.cell(row_offset, GRID_WEEKDAY_COLUMN, WEEKDAY_LABELS[target_date.weekday()])
        date_fill = "E4F1E8" if target_date.weekday() >= 5 else "F2F5F3"
        for column_index in (1, GRID_WEEKDAY_COLUMN):
            cell = grid_sheet.cell(row_offset, column_index)
            cell.fill = PatternFill("solid", fgColor=date_fill)
            cell.border = GRID_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for grid_column_index, configured in enumerate(columns, start=GRID_START_COLUMN):
            cell = grid_sheet.cell(row_offset, grid_column_index)
            body_fill = "FFF2D8" if configured.final_only else "F8FBF9"
            cell.fill = PatternFill("solid", fgColor=body_fill)
            cell.border = GRID_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            draft_value = cells.get((target_date, configured.sort_order))
            if draft_value:
                cell.value = draft_value[0]
                if not draft_value[1]:
                    cell.comment = None
    grid_end_row = GRID_START_ROW + len(dates) - 1
    editable_range = (
        f"{get_column_letter(GRID_START_COLUMN)}{GRID_START_ROW}:"
        f"{get_column_letter(grid_end_column)}{grid_end_row}"
    )
    grid_sheet.conditional_formatting.add(
        editable_range,
        FormulaRule(
            formula=[
                f'ISNUMBER(SEARCH("女",{get_column_letter(GRID_START_COLUMN)}{GRID_START_ROW}))'
            ],
            font=Font(color="C62828", bold=True),
        ),
    )
    grid_sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=grid_end_column)
    grid_sheet.cell(
        3,
        1,
        "日期和星期已自动生成；在网格中填写 A1vsA2，女子比赛写 A1vsA2（女）。"
        "时间与场地表头可以编辑，场地末尾加（仅决赛）会把该列限制为决赛。",
    )
    grid_sheet.cell(3, 1).font = Font(color="66766F", italic=True)
    grid_sheet.cell(3, 1).alignment = Alignment(wrap_text=True)
    _style_title(grid_sheet, "赛程网格 · 一天一行", grid_end_column)
    grid_sheet.freeze_panes = "C7"
    grid_sheet.column_dimensions["A"].width = 13
    grid_sheet.column_dimensions["B"].width = 9
    for column in range(GRID_START_COLUMN, grid_end_column + 1):
        grid_sheet.column_dimensions[get_column_letter(column)].width = 17
    grid_sheet.auto_filter.ref = f"A{GRID_TIME_ROW}:B{grid_end_row}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _preflight_xlsx(content: bytes) -> None:
    if not content:
        raise ScheduleImportError("上传文件为空。", "EMPTY_FILE")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ScheduleImportError("上传文件超过 10 MB。", "FILE_TOO_LARGE")
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise ScheduleImportError("工作簿包含过多内部文件。", "XLSX_ZIP_BOMB")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise ScheduleImportError("工作簿解压后体积异常。", "XLSX_ZIP_BOMB")
            lower_names = [member.filename.lower() for member in members]
            if any("vbaproject.bin" in name for name in lower_names):
                raise ScheduleImportError("不允许上传含宏的工作簿。", "MACRO_FORBIDDEN")
            if any(name.startswith("xl/externallinks/") for name in lower_names):
                raise ScheduleImportError(
                    "不允许上传含外部链接的工作簿。", "EXTERNAL_LINK_FORBIDDEN"
                )
            for member in members:
                if not member.filename.lower().endswith(".rels"):
                    continue
                relationships = archive.read(member).lower()
                if (
                    b'targetmode="external"' in relationships
                    or b"targetmode='external'" in relationships
                ):
                    raise ScheduleImportError(
                        "不允许上传含外部链接的工作簿。", "EXTERNAL_LINK_FORBIDDEN"
                    )
    except BadZipFile as error:
        raise ScheduleImportError("文件不是有效的 XLSX 工作簿。", "INVALID_XLSX") from error


def _scan_formulas(workbook: Workbook, issues: list[ParsedIssue]) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    issues.append(
                        _issue(
                            "FORMULA_FORBIDDEN",
                            f"工作表“{sheet.title}”不允许单元格公式。",
                            cell=f"{sheet.title}!{cell.coordinate}",
                        )
                    )


def _validate_headers(sheet, row: int, expected: list[str]) -> None:
    actual = [sheet.cell(row, column).value for column in range(1, len(expected) + 1)]
    if actual != expected:
        raise ScheduleImportError(
            f"工作表“{sheet.title}”第 {row} 行字段名或顺序不正确。",
            "COLUMN_STRUCTURE_CHANGED",
        )


def _read_text(cell: Cell, label: str, issues: list[ParsedIssue]) -> str | None:
    if cell.data_type == "f":
        return None
    if cell.value in (None, ""):
        issues.append(_issue("REQUIRED_FIELD_MISSING", f"{label}不能为空。", cell=cell.coordinate))
        return None
    if not isinstance(cell.value, str):
        issues.append(_issue("TEXT_FIELD_REQUIRED", f"{label}必须是文本。", cell=cell.coordinate))
        return None
    value = cell.value.strip()
    if not value:
        issues.append(_issue("REQUIRED_FIELD_MISSING", f"{label}不能为空。", cell=cell.coordinate))
        return None
    return value


def _cell_date(cell: Cell) -> date | None:
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _parse_slot_definitions_from_sheet(
    season: Season,
    sheet,
    issues: list[ParsedIssue],
) -> tuple[
    dict[tuple[str, str, int, str], ParsedFamily],
    dict[tuple[str, str], ParsedGroup],
    dict[tuple[str, str], ParsedSlot],
]:
    _validate_headers(sheet, SLOT_HEADER_ROW, SLOT_HEADERS)
    divisions = list(Division.objects.filter(season=season).order_by("sort_order", "name"))
    divisions_by_name = {item.name: item for item in divisions}
    raw_slots: list[tuple[Division, str, int, str, int, str]] = []
    seen_slot_keys: set[tuple[str, str]] = set()
    prefix_owner: dict[tuple[str, str], tuple[str, str, int]] = {}

    if sheet.max_row - SLOT_START_ROW + 1 > MAX_DATA_ROWS:
        issues.append(_issue("TOO_MANY_ROWS", "签位定义超过 5000 行。"))
    last_row = min(sheet.max_row, SLOT_START_ROW + MAX_DATA_ROWS - 1)
    for row in range(SLOT_START_ROW, last_row + 1):
        cells = [sheet.cell(row, column) for column in range(1, 5)]
        if all(cell.value in (None, "") for cell in cells):
            continue
        division_name = _read_text(cells[0], "组别", issues)
        stage_label = _read_text(cells[1], "场次信息", issues)
        try:
            round_number = int(cells[2].value)
        except (TypeError, ValueError):
            round_number = 0
            issues.append(
                _issue(
                    "INVALID_ROUND_NUMBER",
                    "轮次必须是大于等于 1 的整数。",
                    cell=cells[2].coordinate,
                )
            )
        slot_code = _read_text(cells[3], "签位代码", issues)
        if not division_name or not stage_label or round_number < 1 or not slot_code:
            continue
        division = divisions_by_name.get(division_name)
        if division is None:
            issues.append(
                _issue(
                    "UNKNOWN_DIVISION",
                    f"数据库中不存在组别“{division_name}”。",
                    cell=cells[0].coordinate,
                )
            )
            continue
        if not division.teams.filter(active=True).exists():
            issues.append(
                _issue(
                    "DIVISION_HAS_NO_TEAMS",
                    f"组别 {division.name} 尚无启用球队。",
                    cell=cells[0].coordinate,
                )
            )
        stage = STAGE_BY_LABEL.get(stage_label)
        if stage is None:
            issues.append(
                _issue(
                    "INVALID_STAGE",
                    f"未知比赛阶段：{stage_label}。",
                    cell=cells[1].coordinate,
                )
            )
            continue
        match = SLOT_PATTERN.fullmatch(slot_code)
        if match is None:
            issues.append(
                _issue(
                    "INVALID_SLOT_CODE",
                    "签位代码必须是一个大小写敏感英文字母加正整数，例如 A1 或 a1。",
                    cell=cells[3].coordinate,
                )
            )
            continue
        prefix, number_text = match.groups()
        number = int(number_text)
        namespace_key = (division.gender, prefix)
        owner = (division.code, stage, round_number)
        if namespace_key in prefix_owner and prefix_owner[namespace_key] != owner:
            issues.append(
                _issue(
                    "DUPLICATE_GENDER_PREFIX",
                    f"同一性别内签位字母 {prefix} 只能属于一个组别和阶段。",
                    cell=cells[3].coordinate,
                )
            )
        prefix_owner[namespace_key] = owner
        slot_key = (division.code, slot_code)
        if slot_key in seen_slot_keys:
            issues.append(
                _issue(
                    "DUPLICATE_SLOT_DEFINITION",
                    f"签位 {division.name}/{slot_code} 在表中重复。",
                    cell=cells[3].coordinate,
                )
            )
            continue
        seen_slot_keys.add(slot_key)
        raw_slots.append((division, stage, round_number, prefix, number, cells[3].coordinate))

    family_numbers: dict[tuple[str, str, int, str], set[int]] = defaultdict(set)
    family_divisions: dict[tuple[str, str, int, str], Division] = {}
    family_cells: dict[tuple[str, str, int, str], dict[int, str]] = defaultdict(dict)
    for division, stage, round_number, prefix, number, cell in raw_slots:
        key = (division.code, stage, round_number, prefix)
        family_numbers[key].add(number)
        family_divisions[key] = division
        family_cells[key][number] = cell

    sorted_keys = sorted(
        family_numbers,
        key=lambda key: (
            family_divisions[key].sort_order,
            STAGE_ORDER[key[1]],
            key[2],
            key[3],
        ),
    )
    families: dict[tuple[str, str, int, str], ParsedFamily] = {}
    groups: dict[tuple[str, str], ParsedGroup] = {}
    slots: dict[tuple[str, str], ParsedSlot] = {}
    group_totals: Counter[str] = Counter()
    for sort_order, key in enumerate(sorted_keys, start=1):
        division = family_divisions[key]
        _division_code, stage, round_number, prefix = key
        numbers = family_numbers[key]
        slot_count = max(numbers)
        expected_numbers = set(range(1, slot_count + 1))
        if numbers != expected_numbers:
            missing = sorted(expected_numbers - numbers)
            issues.append(
                _issue(
                    "NON_CONTIGUOUS_SLOT_NUMBERS",
                    f"{division.name}/{STAGE_LABEL_BY_VALUE[stage]}/{prefix} 的编号必须从 1 连续，"
                    f"缺少 {', '.join(map(str, missing[:20]))}。",
                )
            )
        if slot_count < 2:
            issues.append(_issue("INVALID_SLOT_COUNT", "每个签位族至少需要 2 个签位。"))
        if stage == Game.Stage.SEMIFINAL and slot_count != 4:
            issues.append(_issue("INVALID_SLOT_COUNT", "半决赛签位数固定为 4。"))
        if stage == Game.Stage.FINAL and slot_count != 2:
            issues.append(_issue("INVALID_SLOT_COUNT", "决赛签位数固定为 2。"))
        if stage in {Game.Stage.KNOCKOUT, Game.Stage.RELEGATION} and (
            slot_count < 2 or slot_count % 2
        ):
            issues.append(
                _issue("INVALID_SLOT_COUNT", "淘汰赛和保级赛签位数必须是不少于 2 的偶数。")
            )
        family = ParsedFamily(
            division_id=str(division.id),
            division_code=division.code,
            division_name=division.name,
            gender=division.gender,
            stage=stage,
            round_number=round_number,
            prefix=prefix,
            slot_count=slot_count,
            sort_order=sort_order,
        )
        families[key] = family
        if stage in GROUP_STAGES:
            group_totals[division.code] += slot_count
            groups[(division.code, prefix)] = ParsedGroup(
                division_code=division.code,
                code=prefix,
                name=f"{prefix}组",
                sort_order=sort_order,
            )
        for number in sorted(numbers):
            code = f"{prefix}{number}"
            slots[(division.code, code)] = ParsedSlot(
                division_code=division.code,
                division_name=division.name,
                gender=division.gender,
                stage=stage,
                round_number=round_number,
                prefix=prefix,
                number=number,
                code=code,
                group_code=prefix if stage in GROUP_STAGES else None,
                cell=family_cells[key][number],
            )

    for division in divisions:
        if division.code not in group_totals:
            continue
        team_count = division.teams.filter(active=True).count()
        if group_totals[division.code] != team_count:
            issues.append(
                _issue(
                    "SLOT_COUNT_TEAM_MISMATCH",
                    f"{division.name}的小组赛/循环赛签位共 {group_totals[division.code]} 个，"
                    f"必须等于当前启用球队数 {team_count}。",
                )
            )

    existing_groups = {
        (item.division.code, item.code): item
        for item in CompetitionGroup.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    for key, parsed in groups.items():
        existing = existing_groups.get(key)
        if existing and existing.name != parsed.name:
            issues.append(
                _issue(
                    "EXISTING_GROUP_CONFLICT",
                    f"小组 {parsed.division_code}/{parsed.code} 与数据库现有定义不一致。",
                )
            )
    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division", "group"
        )
    }
    for key, parsed in slots.items():
        existing = existing_slots.get(key)
        if existing is None:
            continue
        existing_group = existing.group.code if existing.group_id else None
        if (existing_group, existing.label, existing.seed) != (
            parsed.group_code,
            parsed.code,
            parsed.number,
        ):
            issues.append(
                _issue(
                    "EXISTING_SLOT_CONFLICT",
                    f"签位 {parsed.division_name}/{parsed.code} 与数据库现有定义不一致。",
                    cell=parsed.cell,
                )
            )
    if not slots:
        issues.append(_issue("NO_SLOTS_IN_WORKBOOK", "签位定义中没有可用签位。"))
    return families, groups, slots


def _configured_slot_definitions(
    season: Season,
    issues: list[ParsedIssue],
) -> tuple[
    dict[tuple[str, str, int, str], ParsedFamily],
    dict[tuple[str, str], ParsedGroup],
    dict[tuple[str, str], ParsedSlot],
]:
    """Build the import namespace from authoritative season configuration.

    The workbook's second sheet is deliberately informational. Upload validation
    never reads its business values, so changing or clearing that sheet cannot
    redefine divisions, stages, groups, or participant slots.
    """

    families: dict[tuple[str, str, int, str], ParsedFamily] = {}
    groups: dict[tuple[str, str], ParsedGroup] = {}
    slots: dict[tuple[str, str], ParsedSlot] = {}
    for configured in _slot_families(season):
        division = configured.division
        parsed_family = ParsedFamily(
            division_id=str(division.id),
            division_code=division.code,
            division_name=division.name,
            gender=division.gender,
            stage=configured.stage,
            round_number=configured.round_number,
            prefix=configured.prefix,
            slot_count=configured.slot_count,
            sort_order=configured.sort_order,
        )
        families[parsed_family.key] = parsed_family
        if configured.stage in GROUP_STAGES:
            groups[(division.code, configured.prefix)] = ParsedGroup(
                division_code=division.code,
                code=configured.prefix,
                name=f"{configured.prefix}组",
                sort_order=configured.sort_order,
            )
        for number in range(1, configured.slot_count + 1):
            code = f"{configured.prefix}{number}"
            slots[(division.code, code)] = ParsedSlot(
                division_code=division.code,
                division_name=division.name,
                gender=division.gender,
                stage=configured.stage,
                round_number=configured.round_number,
                prefix=configured.prefix,
                number=number,
                code=code,
                group_code=configured.prefix if configured.stage in GROUP_STAGES else None,
                cell="",
            )

    existing_groups = {
        (item.division.code, item.code): item
        for item in CompetitionGroup.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    for key, parsed in groups.items():
        existing = existing_groups.get(key)
        if existing and existing.name != parsed.name:
            issues.append(
                _issue(
                    "EXISTING_GROUP_CONFLICT",
                    f"小组 {parsed.division_code}/{parsed.code} 与数据库现有定义不一致。",
                )
            )

    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division", "group"
        )
    }
    for key, parsed in slots.items():
        existing = existing_slots.get(key)
        if existing is None:
            continue
        existing_group = existing.group.code if existing.group_id else None
        if (existing_group, existing.label, existing.seed) != (
            parsed.group_code,
            parsed.code,
            parsed.number,
        ):
            issues.append(
                _issue(
                    "EXISTING_SLOT_CONFLICT",
                    f"签位 {parsed.division_name}/{parsed.code} 与数据库现有定义不一致。",
                )
            )
    return families, groups, slots


def _time_text(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, str):
        normalized = value.strip()
        try:
            return time.fromisoformat(normalized).strftime("%H:%M")
        except ValueError:
            return None
    return None


def _game_identity(
    division_code: str, stage: str, first_slot: str, second_slot: str
) -> tuple[str, str]:
    left, right = sorted((first_slot, second_slot))
    semantic = f"{division_code}|{stage}|{left}|{right}"
    readable = f"{division_code}-{STAGE_CODE[stage]}-{left}-{right}"
    if len(readable) <= 40:
        return readable, semantic
    digest = hashlib.sha256(semantic.encode("utf-8")).hexdigest()[:10]
    return f"{readable[:29]}-{digest}", semantic


def _parse_grid_columns(season: Season, sheet) -> list[ParsedGridColumn]:
    if sheet.cell(GRID_TIME_ROW, 1).value != "日期":
        raise ScheduleImportError("赛程网格 A5 必须为“日期”。", "COLUMN_STRUCTURE_CHANGED")
    if sheet.cell(GRID_VENUE_ROW, 1).value != "yyyy-mm-dd":
        raise ScheduleImportError(
            "赛程网格 A6 必须为“yyyy-mm-dd”。", "COLUMN_STRUCTURE_CHANGED"
        )
    if sheet.cell(GRID_TIME_ROW, GRID_WEEKDAY_COLUMN).value != "星期":
        raise ScheduleImportError("赛程网格 B5 必须为“星期”。", "COLUMN_STRUCTURE_CHANGED")
    if sheet.cell(GRID_VENUE_ROW, GRID_WEEKDAY_COLUMN).value != "自动生成":
        raise ScheduleImportError(
            "赛程网格 B6 必须为“自动生成”。", "COLUMN_STRUCTURE_CHANGED"
        )

    header_columns = [
        column_index
        for column_index in range(GRID_START_COLUMN, sheet.max_column + 1)
        if sheet.cell(GRID_TIME_ROW, column_index).value not in (None, "")
        or sheet.cell(GRID_VENUE_ROW, column_index).value not in (None, "")
    ]
    if not header_columns:
        raise ScheduleImportError("赛程网格至少需要一个时段+场地列。", "NO_GRID_COLUMNS")
    last_column = max(header_columns)
    column_count = last_column - GRID_START_COLUMN + 1
    if column_count > MAX_GRID_COLUMNS:
        raise ScheduleImportError(
            f"赛程网格最多支持 {MAX_GRID_COLUMNS} 个连续排期列。",
            "TOO_MANY_GRID_COLUMNS",
        )

    periods_by_time: dict[str, list[Period]] = defaultdict(list)
    for period in _periods(season):
        periods_by_time[period.start_time.strftime("%H:%M")].append(period)
    columns: list[ParsedGridColumn] = []
    seen: set[tuple[str, str]] = set()
    for sort_order, column_index in enumerate(
        range(GRID_START_COLUMN, last_column + 1), start=1
    ):
        time_cell = sheet.cell(GRID_TIME_ROW, column_index)
        venue_cell = sheet.cell(GRID_VENUE_ROW, column_index)
        if time_cell.value in (None, "") or venue_cell.value in (None, ""):
            raise ScheduleImportError(
                f"赛程网格 {get_column_letter(column_index)} 列的时间和场地必须同时填写，"
                "且列之间不能留空。",
                "GRID_COLUMN_GAP",
            )
        time_text = _time_text(time_cell.value)
        matching_periods = periods_by_time.get(time_text or "", [])
        if not matching_periods:
            allowed = "、".join(sorted(periods_by_time))
            raise ScheduleImportError(
                f"赛程网格 {get_column_letter(column_index)}5 的时间必须匹配赛季时段：{allowed}。",
                "UNKNOWN_GRID_PERIOD",
            )
        if len(matching_periods) > 1:
            raise ScheduleImportError(
                f"赛季内有多个时段使用 {time_text}，无法识别 "
                f"{get_column_letter(column_index)} 列。",
                "AMBIGUOUS_GRID_PERIOD",
            )
        venue_name, final_only = _normalize_venue_name(venue_cell.value)
        if not venue_name:
            raise ScheduleImportError(
                f"赛程网格 {get_column_letter(column_index)}6 的场地不能为空。",
                "EMPTY_GRID_VENUE",
            )
        if len(venue_name) > 120:
            raise ScheduleImportError(
                f"赛程网格 {get_column_letter(column_index)}6 的场地名称不能超过 120 个字符。",
                "GRID_VENUE_TOO_LONG",
            )
        period = matching_periods[0]
        duplicate_key = (str(period.id), venue_name.casefold())
        if duplicate_key in seen:
            raise ScheduleImportError(
                f"{time_text} + {venue_name} 在赛程网格表头中重复。",
                "DUPLICATE_GRID_COLUMN",
            )
        seen.add(duplicate_key)
        columns.append(
            ParsedGridColumn(
                period_id=str(period.id),
                period_code=period.code,
                period_name=period.name,
                start_time=time_text or "",
                venue_name=venue_name,
                final_only=final_only,
                sort_order=sort_order,
                source_column=column_index,
            )
        )

    for column_index in range(last_column + 1, sheet.max_column + 1):
        if any(
            sheet.cell(row_index, column_index).value not in (None, "")
            for row_index in range(GRID_START_ROW, min(sheet.max_row, MAX_DATA_ROWS) + 1)
        ):
            raise ScheduleImportError(
                "赛程网格存在没有时间/场地表头的数据列。", "GRID_COLUMN_WITHOUT_HEADER"
            )
    return columns


def _parse_grid(
    season: Season,
    sheet,
    slots: dict[tuple[str, str], ParsedSlot],
    issues: list[ParsedIssue],
    leader_adjustable_by_cell: dict[str, bool] | None = None,
) -> tuple[dict[str, ParsedGame], list[ParsedGridColumn]]:
    leader_adjustable_by_cell = leader_adjustable_by_cell or {}
    configured_columns = _parse_grid_columns(season, sheet)
    expected_end_column = GRID_START_COLUMN + len(configured_columns) - 1

    expected_dates = _calendar_dates(season)
    for offset, expected_date in enumerate(expected_dates, start=GRID_START_ROW):
        actual_date = _cell_date(sheet.cell(offset, 1))
        if actual_date != expected_date:
            issues.append(
                _issue(
                    "GRID_DATES_CHANGED",
                    f"第 {offset} 行日期应为 {expected_date.isoformat()}。",
                    cell=sheet.cell(offset, 1).coordinate,
                )
            )
    expected_end_row = GRID_START_ROW + len(expected_dates) - 1
    for row in range(expected_end_row + 1, min(sheet.max_row, MAX_DATA_ROWS) + 1):
        if any(
            sheet.cell(row, column).value not in (None, "")
            for column in range(1, expected_end_column + 1)
        ):
            issues.append(
                _issue(
                    "EXTRA_GRID_ROW",
                    "赛程网格包含赛季日期范围外的额外数据行。",
                    cell=sheet.cell(row, 1).coordinate,
                )
            )

    namespace_slots: dict[tuple[str, str], ParsedSlot] = {}
    for slot in slots.values():
        namespace_slots[(slot.gender, slot.code)] = slot
    games: dict[str, ParsedGame] = {}
    semantic_cells: dict[str, str] = {}
    for row_offset, target_date in enumerate(expected_dates, start=GRID_START_ROW):
        for column_index, configured in enumerate(
            configured_columns, start=GRID_START_COLUMN
        ):
            cell = sheet.cell(row_offset, column_index)
            if cell.data_type == "f" or cell.value in (None, ""):
                continue
            if not isinstance(cell.value, str):
                issues.append(
                    _issue("INVALID_MATCHUP", "对阵必须是文本，例如 A1vsA2。", cell=cell.coordinate)
                )
                continue
            raw = cell.value.strip()
            match = MATCHUP_PATTERN.fullmatch(raw)
            if match is None:
                issues.append(
                    _issue(
                        "INVALID_MATCHUP",
                        "对阵格式应为 A1vsA2；女子比赛末尾加（女）。",
                        cell=cell.coordinate,
                    )
                )
                continue
            home_code, away_code = match.groups()
            gender = (
                Division.Gender.WOMEN
                if WOMEN_SUFFIX_PATTERN.search(raw)
                else Division.Gender.MEN
            )
            home = namespace_slots.get((gender, home_code))
            away = namespace_slots.get((gender, away_code))
            if home is None or away is None:
                suffix_hint = "并检查女子比赛是否带（女）" if gender == Division.Gender.MEN else ""
                issues.append(
                    _issue(
                        "UNKNOWN_SLOT",
                        f"找不到对阵签位 {home_code} 或 {away_code}{suffix_hint}。",
                        cell=cell.coordinate,
                    )
                )
                continue
            if home.code == away.code:
                issues.append(
                    _issue("SAME_PARTICIPANT", "双方签位不能相同。", cell=cell.coordinate)
                )
                continue
            if home.family_key != away.family_key:
                issues.append(
                    _issue(
                        "SLOT_FAMILY_MISMATCH",
                        "双方必须属于同一组别、阶段和签位字母。",
                        cell=cell.coordinate,
                    )
                )
                continue
            if configured.final_only and home.stage != Game.Stage.FINAL:
                issues.append(
                    _issue(
                        "FINAL_ONLY_COLUMN",
                        f"{configured.start_time} + {configured.venue_name} 只允许决赛。",
                        cell=cell.coordinate,
                    )
                )
            code, semantic = _game_identity(
                home.division_code, home.stage, home.code, away.code
            )
            if semantic in semantic_cells:
                issues.append(
                    _issue(
                        "DUPLICATE_MATCHUP",
                        f"该对阵已在 {semantic_cells[semantic]} 出现；主客颠倒仍视为同一场。",
                        cell=cell.coordinate,
                    )
                )
                continue
            semantic_cells[semantic] = cell.coordinate
            if code in games:
                issues.append(
                    _issue(
                        "GAME_CODE_COLLISION",
                        "自动生成的比赛编号发生冲突。",
                        cell=cell.coordinate,
                    )
                )
                continue
            games[code] = ParsedGame(
                code=code,
                semantic_key=semantic,
                division_code=home.division_code,
                division_name=home.division_name,
                group_code=home.group_code,
                stage=home.stage,
                round_number=home.round_number,
                home_slot_code=home.code,
                away_slot_code=away.code,
                target_date=target_date,
                period_id=configured.period_id,
                period_code=configured.period_code,
                period_name=configured.period_name,
                start_time=configured.start_time,
                venue_name=configured.venue_name,
                final_only=configured.final_only,
                leader_adjustable=leader_adjustable_by_cell.get(
                    f"赛程网格!{cell.coordinate}",
                    leader_adjustable_by_cell.get(cell.coordinate, True),
                ),
                cell=f"赛程网格!{cell.coordinate}",
            )
    if not games:
        issues.append(_issue("NO_GAMES_IN_WORKBOOK", "赛程网格中没有可新增比赛。"))
    return games, configured_columns


def _validate_existing_matchups(
    season: Season,
    games: dict[str, ParsedGame],
    issues: list[ParsedIssue],
) -> None:
    existing_codes = set(Game.objects.filter(season=season).values_list("code", flat=True))
    existing_semantics: dict[str, str] = {}
    for game in (
        Game.objects.filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related("division", "home_slot", "away_slot")
    ):
        if not game.home_slot_id or not game.away_slot_id:
            continue
        _code, semantic = _game_identity(
            game.division.code,
            game.stage,
            game.home_slot.code,
            game.away_slot.code,
        )
        existing_semantics[semantic] = game.code
    for game in games.values():
        if game.code in existing_codes:
            issues.append(
                _issue(
                    "GAME_CODE_ALREADY_EXISTS",
                    f"自动比赛编号 {game.code} 已存在；新增式导入不会更新比赛。",
                    cell=game.cell,
                )
            )
        if game.semantic_key in existing_semantics:
            issues.append(
                _issue(
                    "MATCHUP_ALREADY_EXISTS",
                    f"该对阵已存在于比赛 {existing_semantics[game.semantic_key]}。",
                    cell=game.cell,
                )
            )


def _validate_completeness(
    season: Season,
    families: dict[tuple[str, str, int, str], ParsedFamily],
    slots: dict[tuple[str, str], ParsedSlot],
    games: dict[str, ParsedGame],
    issues: list[ParsedIssue],
) -> int:
    pair_counts: dict[tuple[str, str, int, str], Counter[tuple[str, str]]] = defaultdict(Counter)
    usage_counts: dict[tuple[str, str, int, str], Counter[str]] = defaultdict(Counter)
    covered_matchups: set[
        tuple[tuple[str, str, int, str], tuple[str, str]]
    ] = set()

    def add_game(division_code: str, stage: str, home: str, away: str) -> None:
        first = slots.get((division_code, home))
        second = slots.get((division_code, away))
        if first is None or second is None or first.family_key != second.family_key:
            return
        if first.stage != stage:
            return
        key = first.family_key
        matchup = tuple(sorted((home, away)))
        pair_counts[key][matchup] += 1
        covered_matchups.add((key, matchup))
        usage_counts[key][home] += 1
        usage_counts[key][away] += 1

    for existing in (
        Game.objects.filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related("division", "home_slot", "away_slot")
    ):
        if existing.home_slot_id and existing.away_slot_id:
            add_game(
                existing.division.code,
                existing.stage,
                existing.home_slot.code,
                existing.away_slot.code,
            )
    for game in games.values():
        add_game(
            game.division_code,
            game.stage,
            game.home_slot_code,
            game.away_slot_code,
        )

    for key, family in families.items():
        family_slots = [
            slot.code for slot in slots.values() if slot.family_key == key
        ]
        label = f"{family.division_name}/{STAGE_LABEL_BY_VALUE[family.stage]}/{family.prefix}"
        if family.stage in GROUP_STAGES:
            expected = {
                tuple(sorted((family_slots[left], family_slots[right])))
                for left in range(len(family_slots))
                for right in range(left + 1, len(family_slots))
            }
            missing = sorted(expected - set(pair_counts[key]))
            duplicates = sorted(pair for pair, count in pair_counts[key].items() if count > 1)
            if missing:
                preview = ", ".join(f"{a}vs{b}" for a, b in missing[:12])
                issues.append(
                    _issue(
                        "MISSING_ROUND_ROBIN_MATCHUPS",
                        f"{label} 缺少 {len(missing)} 场应有对阵：{preview}。",
                        context={"missing_count": len(missing)},
                    )
                )
            if duplicates:
                issues.append(
                    _issue(
                        "DUPLICATE_FAMILY_MATCHUPS",
                        f"{label} 有 {len(duplicates)} 组重复对阵。",
                    )
                )
        else:
            missing_slots = sorted(code for code in family_slots if usage_counts[key][code] == 0)
            repeated_slots = sorted(code for code in family_slots if usage_counts[key][code] > 1)
            if missing_slots:
                issues.append(
                    _issue(
                        "UNUSED_ELIMINATION_SLOTS",
                        f"{label} 有未使用签位：{', '.join(missing_slots[:20])}。",
                        context={"missing_count": len(missing_slots)},
                    )
                )
            if repeated_slots:
                issues.append(
                    _issue(
                        "REUSED_ELIMINATION_SLOTS",
                        f"{label} 有重复使用签位：{', '.join(repeated_slots[:20])}。",
                        context={"repeated_count": len(repeated_slots)},
                    )
                )
    return len(covered_matchups)


def _slot_reference(division_code: str, slot_code: str) -> str:
    return f"slot:{division_code}:{slot_code}"


def _validate_schedule_conflicts(
    season: Season,
    games: dict[str, ParsedGame],
    issues: list[ParsedIssue],
) -> None:
    existing_games = list(
        Game.objects.filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related("division", "period", "home_slot", "away_slot")
    )
    reservations = list(
        SlotReservation.objects.filter(
            season=season, status=SlotReservation.Status.ACTIVE
        ).select_related(
            "period",
            "venue",
            "request__game__division",
            "request__game__home_slot",
            "request__game__away_slot",
        )
    )
    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    assignment_by_slot = {
        item.slot_id: item.team_id
        for item in DrawAssignment.objects.filter(season=season)
    }
    occupancy: dict[tuple[date, object, str], list[tuple[str, str, str]]] = defaultdict(list)
    participants: dict[tuple[date, object], dict[str, list[tuple[str, str]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    counts: Counter[tuple[date, object]] = Counter()
    keys_with_new_games: set[tuple[date, object]] = set()

    def add_existing_participants(game: Game, key: tuple[date, object], label: str) -> None:
        references: set[str] = set()
        for side in ("home", "away"):
            team_id = getattr(game, f"{side}_team_id")
            slot = getattr(game, f"{side}_slot")
            if team_id:
                references.add(f"team:{team_id}")
            if slot:
                references.add(_slot_reference(game.division.code, slot.code))
                assigned_team_id = assignment_by_slot.get(slot.id)
                if assigned_team_id:
                    references.add(f"team:{assigned_team_id}")
        for reference in references:
            participants[key][reference].append(("existing", label))

    for game in existing_games:
        key = (game.date, game.period_id)
        counts[key] += 1
        occupancy[(game.date, game.period_id, game.venue_name)].append(
            ("existing", game.code, "")
        )
        add_existing_participants(game, key, game.code)
    for reservation in reservations:
        key = (reservation.date, reservation.period_id)
        counts[key] += 1
        occupancy[(reservation.date, reservation.period_id, reservation.venue_name)].append(
            ("reservation", str(reservation.id), "")
        )
        request = getattr(reservation, "request", None)
        if request is not None:
            add_existing_participants(request.game, key, f"预留:{reservation.id}")

    for game in games.values():
        period_id = next(
            item.id for item in _periods(season) if str(item.id) == game.period_id
        )
        key = (game.target_date, period_id)
        keys_with_new_games.add(key)
        counts[key] += 1
        occupancy[(game.target_date, period_id, game.venue_name)].append(
            ("new", game.code, game.cell)
        )
        for slot_code in (game.home_slot_code, game.away_slot_code):
            slot_key = (game.division_code, slot_code)
            participants[key][_slot_reference(*slot_key)].append(("new", game.code))
            existing_slot = existing_slots.get(slot_key)
            assigned_team_id = assignment_by_slot.get(existing_slot.id) if existing_slot else None
            if assigned_team_id:
                participants[key][f"team:{assigned_team_id}"].append(("new", game.code))

    for key, occupants in occupancy.items():
        if len(occupants) <= 1 or not any(kind == "new" for kind, _label, _cell in occupants):
            continue
        new_occupant = next(item for item in occupants if item[0] == "new")
        period = Period.objects.filter(id=key[1]).first()
        period_label = period.code.upper() if period else str(key[1])
        has_reservation = any(kind == "reservation" for kind, _label, _cell in occupants)
        if has_reservation:
            issues.append(
                _issue(
                    "VENUE_OCCUPIED",
                    f"{key[0].isoformat()} / {period_label} 的目标时段与活动调赛资源预留冲突；"
                    "具体场地将在调赛生效后公布。",
                    context={
                        "date": key[0].isoformat(),
                        "period_code": period_label,
                        "venue_hidden_until_reschedule_effective": True,
                    },
                )
            )
            continue
        issues.append(
            _issue(
                "VENUE_OCCUPIED",
                f"{key[0].isoformat()} / {period_label} / {key[2]} 已有其他比赛或有效预留。",
                cell=new_occupant[2],
                context={"occupants": [label for _kind, label, _cell in occupants]},
            )
        )

    periods = _periods(season)
    capacity_map = effective_capacity_map(
        season=season,
        dates=sorted({key[0] for key in keys_with_new_games}),
        periods=periods,
    )
    period_by_id = {item.id: item for item in periods}
    for key in keys_with_new_games:
        capacity = capacity_map.get(key, 0)
        if counts[key] > capacity:
            period = period_by_id.get(key[1])
            period_label = period.code.upper() if period else str(key[1])
            issues.append(
                _issue(
                    "CAPACITY_EXCEEDED",
                    f"{key[0].isoformat()} / {period_label} 共占用 {counts[key]} 场，"
                    f"容量为 {capacity}。",
                    context={
                        "date": key[0].isoformat(),
                        "period_code": period_label,
                        "used": counts[key],
                        "capacity": capacity,
                    },
                )
            )
    for key, references in participants.items():
        for reference, occupants in references.items():
            labels = list(dict.fromkeys(label for _kind, label in occupants))
            if len(labels) > 1 and any(kind == "new" for kind, _label in occupants):
                issues.append(
                    _issue(
                        "PARTICIPANT_TIME_CONFLICT",
                        f"同一球队或签位在同一时段参加多场比赛：{', '.join(labels)}。",
                        context={
                            "date": key[0].isoformat(),
                            "period_id": str(key[1]),
                            "participant": reference,
                            "games": labels,
                        },
                    )
                )


def _build_summary(
    season: Season,
    groups: dict[tuple[str, str], ParsedGroup],
    slots: dict[tuple[str, str], ParsedSlot],
    games: dict[str, ParsedGame],
    issues: list[ParsedIssue],
    covered_game_count: int,
    grid_column_count: int,
) -> dict[str, object]:
    existing_groups = {
        (item.division.code, item.code)
        for item in CompetitionGroup.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    existing_slots = {
        (item.division.code, item.code)
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    existing_codes = set(Game.objects.filter(season=season).values_list("code", flat=True))
    group_rows = [
        {
            "action": "REFERENCE" if key in existing_groups else "CREATE",
            "division_code": item.division_code,
            "division_name": season.divisions.get(code=item.division_code).name,
            "code": item.code,
            "name": item.name,
            "sort_order": item.sort_order,
        }
        for key, item in sorted(groups.items())
    ]
    slot_rows = [
        {
            "action": "REFERENCE" if key in existing_slots else "CREATE",
            "division_code": item.division_code,
            "division_name": item.division_name,
            "group_code": item.group_code,
            "code": item.code,
            "label": item.code,
            "seed": item.number,
        }
        for key, item in sorted(slots.items())
    ]
    game_rows = [
        {
            "action": "CONFLICT" if item.code in existing_codes else "CREATE",
            "code": item.code,
            "division_code": item.division_code,
            "division_name": item.division_name,
            "group_code": item.group_code,
            "stage": item.stage,
            "stage_name": STAGE_LABEL_BY_VALUE[item.stage],
            "round_number": item.round_number,
            "home_slot_code": item.home_slot_code,
            "home_slot_label": item.home_slot_code,
            "away_slot_code": item.away_slot_code,
            "away_slot_label": item.away_slot_code,
            "date": item.target_date.isoformat(),
            "period_code": item.period_code,
            "period_name": item.period_name,
            "start_time": item.start_time,
            "venue_name": item.venue_name,
            "final_only": item.final_only,
            "leader_adjustable": item.leader_adjustable,
            "cell": item.cell,
        }
        for item in sorted(
            games.values(),
            key=lambda row: (row.target_date, row.start_time, row.venue_name, row.code),
        )
    ]
    readiness = schedule_import_readiness(season)
    return {
        "existing_game_count": readiness["existing_game_count"],
        "covered_game_count": covered_game_count,
        "new_group_count": sum(row["action"] == "CREATE" for row in group_rows),
        "referenced_group_count": sum(row["action"] == "REFERENCE" for row in group_rows),
        "new_slot_count": sum(row["action"] == "CREATE" for row in slot_rows),
        "referenced_slot_count": sum(row["action"] == "REFERENCE" for row in slot_rows),
        "new_game_count": sum(row["action"] == "CREATE" for row in game_rows),
        "groups": group_rows,
        "slots": slot_rows,
        "games": game_rows,
        "prerequisites": {
            "division_count": readiness["division_count"],
            "team_count": readiness["team_count"],
            "period_count": readiness["period_count"],
            "venue_count": readiness["venue_count"],
            "slot_family_count": readiness["slot_family_count"],
            "grid_column_count": grid_column_count,
            "calendar_day_count": readiness["calendar_day_count"],
            "expected_game_count": readiness["expected_game_count"],
        },
        "error_count": sum(issue.severity == ImportIssue.Severity.ERROR for issue in issues),
        "warning_count": sum(issue.severity == ImportIssue.Severity.WARNING for issue in issues),
    }


def _empty_summary(season: Season) -> dict[str, object]:
    readiness = schedule_import_readiness(season)
    return {
        "existing_game_count": readiness["existing_game_count"],
        "covered_game_count": 0,
        "new_group_count": 0,
        "referenced_group_count": 0,
        "new_slot_count": 0,
        "referenced_slot_count": 0,
        "new_game_count": 0,
        "groups": [],
        "slots": [],
        "games": [],
        "prerequisites": {
            "division_count": readiness["division_count"],
            "team_count": readiness["team_count"],
            "period_count": readiness["period_count"],
            "venue_count": readiness["venue_count"],
            "slot_family_count": readiness["slot_family_count"],
            "grid_column_count": readiness["grid_column_count"],
            "calendar_day_count": readiness["calendar_day_count"],
            "expected_game_count": readiness["expected_game_count"],
        },
        "error_count": 1,
        "warning_count": 0,
    }


def _analyze_workbook(
    season: Season,
    content: bytes,
    *,
    leader_adjustable_by_cell: dict[str, bool] | None = None,
) -> tuple[list[ParsedIssue], WorkbookAnalysis]:
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as error:
        raise ScheduleImportError("无法读取 XLSX 工作簿。", "INVALID_XLSX") from error
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ScheduleImportError("工作表名称、数量或顺序不正确。", "SHEET_STRUCTURE_CHANGED")
    if any(workbook[name].sheet_state != "visible" for name in EXPECTED_SHEETS):
        raise ScheduleImportError("三个工作表必须全部可见。", "SHEET_VISIBILITY_INVALID")
    if str(workbook["填写说明"]["B4"].value or "").strip() != TEMPLATE_VERSION:
        raise ScheduleImportError(
            f"不支持当前格式版本，请使用 {TEMPLATE_VERSION}。",
            "UNSUPPORTED_TEMPLATE_VERSION",
        )

    issues: list[ParsedIssue] = []
    for blocker in schedule_import_readiness(season)["blockers"]:
        issues.append(_issue(str(blocker["code"]), str(blocker["message"])))
    _scan_formulas(workbook, issues)
    families, groups, slots = _configured_slot_definitions(season, issues)
    games, columns = _parse_grid(
        season,
        workbook["赛程网格"],
        slots,
        issues,
        leader_adjustable_by_cell,
    )
    _validate_existing_matchups(season, games, issues)
    covered_game_count = _validate_completeness(season, families, slots, games, issues)
    expected_game_count = int(schedule_import_readiness(season)["expected_game_count"])
    if covered_game_count != expected_game_count:
        issues.append(
            _issue(
                "GAME_COUNT_MISMATCH",
                f"赛制配置预计 {expected_game_count} 场，现有有效赛程与本次导入合计覆盖 "
                f"{covered_game_count} 场。",
                context={
                    "expected": expected_game_count,
                    "covered": covered_game_count,
                },
            )
        )
    _validate_schedule_conflicts(season, games, issues)
    summary = _build_summary(
        season,
        groups,
        slots,
        games,
        issues,
        covered_game_count,
        len(columns),
    )
    return issues, WorkbookAnalysis(families, groups, slots, games, columns, summary)


def validate_schedule_upload(
    *,
    actor: Account,
    season: Season,
    content: bytes,
    source_name: str,
    source_kind: str = ScheduleImportBatch.SourceKind.XLSX,
    source_draft: ScheduleGridDraft | None = None,
    source_draft_version: int | None = None,
    leader_adjustable_by_cell: dict[str, bool] | None = None,
    now: datetime | None = None,
) -> ScheduleImportBatch:
    del now
    _require_superadmin(actor)
    digest = hashlib.sha256(content).hexdigest()
    safe_name = get_valid_filename(PurePath(source_name).name) or "schedule.xlsx"
    batch = ScheduleImportBatch.objects.create(
        season=season,
        template_version=TEMPLATE_VERSION,
        file_key="",
        file_sha256=digest,
        uploaded_by=actor,
        source_kind=source_kind,
        source_draft=source_draft,
        source_draft_version=source_draft_version,
        source_snapshot={
            "leader_adjustable_by_cell": leader_adjustable_by_cell or {},
        },
    )
    key = f"schedule-imports/{season.id}/{batch.id}/{safe_name}"
    batch.file_key = default_storage.save(key, ContentFile(content))
    try:
        issues, analysis = _analyze_workbook(
            season,
            content,
            leader_adjustable_by_cell=leader_adjustable_by_cell,
        )
        summary = analysis.summary
    except ScheduleImportError as error:
        issues = [_issue(error.code, str(error))]
        summary = _empty_summary(season)
    ImportIssue.objects.bulk_create(
        [ImportIssue(batch=batch, **issue.model_kwargs()) for issue in issues]
    )
    batch.status = ScheduleImportBatch.Status.VALIDATED
    batch.summary = summary
    batch.save(update_fields=["file_key", "status", "summary", "updated_at"])
    return batch


def _lock_schedule_slots(season: Season, games: dict[str, ParsedGame]) -> None:
    keys = sorted({(item.target_date, item.period_id) for item in games.values()})
    periods = {str(item.id): item for item in _periods(season)}
    for target_date, period_id in keys:
        period = periods[period_id]
        try:
            with transaction.atomic():
                ScheduleSlotLock.objects.get_or_create(
                    season=season,
                    date=target_date,
                    period=period,
                )
        except IntegrityError:
            pass
        ScheduleSlotLock.objects.select_for_update().get(
            season=season,
            date=target_date,
            period=period,
        )


def _slot_start(target_date: date, start_time: time, timezone_name: str) -> datetime:
    return datetime.combine(target_date, start_time, tzinfo=ZoneInfo(timezone_name))


def _confirm_schedule_import(
    *,
    actor: Account,
    batch_id: object,
    expected_season_version: int,
    now: datetime,
) -> ScheduleImportBatch:
    with transaction.atomic():
        batch = (
            ScheduleImportBatch.objects.select_for_update()
            .select_related("season")
            .get(id=batch_id)
        )
        if batch.status == ScheduleImportBatch.Status.CONFIRMED:
            return batch
        if batch.status == ScheduleImportBatch.Status.ROLLED_BACK:
            raise ScheduleImportError("该导入批次已回滚，不能再次确认。", "BATCH_ROLLED_BACK")
        if batch.status != ScheduleImportBatch.Status.VALIDATED:
            raise ScheduleImportError("导入批次尚未完成校验。", "BATCH_NOT_VALIDATED")
        if batch.issues.filter(severity=ImportIssue.Severity.ERROR).exists():
            raise ScheduleImportError("导入批次仍有错误，不能确认。", "BATCH_HAS_ERRORS")

        season = Season.objects.select_for_update().get(id=batch.season_id)
        if season.version != expected_season_version:
            raise ScheduleImportError("赛季已被其他操作修改，请重新校验文件。", "VERSION_CONFLICT")
        if season.status != Season.Status.SETUP:
            raise ScheduleImportError("只有准备中的赛季可以确认赛程导入。", "SEASON_NOT_SETUP")
        source_draft = None
        if batch.source_kind == ScheduleImportBatch.SourceKind.ONLINE_DRAFT:
            if not batch.source_draft_id or batch.source_draft_version is None:
                raise ScheduleImportError("在线草稿来源信息不完整。", "DRAFT_SOURCE_MISSING")
            source_draft = ScheduleGridDraft.objects.select_for_update().get(
                id=batch.source_draft_id
            )
            if source_draft.version != batch.source_draft_version:
                raise ScheduleImportError(
                    "在线草稿已在校验后发生变化，请重新核对。", "DRAFT_VERSION_CONFLICT"
                )
        with default_storage.open(batch.file_key, "rb") as source:
            content = source.read()
        if hashlib.sha256(content).hexdigest() != batch.file_sha256:
            raise ScheduleImportError("上传原文件校验值不一致。", "FILE_HASH_MISMATCH")

        list(CompetitionGroup.objects.select_for_update().filter(division__season=season))
        list(ParticipantSlot.objects.select_for_update().filter(division__season=season))
        list(Game.objects.select_for_update().filter(season=season))
        list(ScheduleSlotFamily.objects.select_for_update().filter(season=season))
        list(
            SlotReservation.objects.select_for_update().filter(
                season=season,
                status__in=[SlotReservation.Status.ACTIVE, SlotReservation.Status.CONVERTED],
            )
        )

        leader_adjustable_by_cell = {
            str(key): bool(value)
            for key, value in batch.source_snapshot.get(
                "leader_adjustable_by_cell", {}
            ).items()
        }
        issues, analysis = _analyze_workbook(
            season,
            content,
            leader_adjustable_by_cell=leader_adjustable_by_cell,
        )
        errors = [issue for issue in issues if issue.severity == ImportIssue.Severity.ERROR]
        if errors:
            raise ScheduleImportError(
                f"确认前复核失败：{errors[0].message}",
                "REVALIDATION_FAILED",
            )
        _lock_schedule_slots(season, analysis.games)

        divisions = {item.code: item for item in Division.objects.filter(season=season)}
        group_objects = {
            (item.division.code, item.code): item
            for item in CompetitionGroup.objects.filter(division__season=season).select_related(
                "division"
            )
        }
        created_groups: list[CompetitionGroup] = []
        for key, item in sorted(analysis.groups.items()):
            if key in group_objects:
                continue
            group = CompetitionGroup(
                division=divisions[item.division_code],
                code=item.code,
                name=item.name,
                sort_order=item.sort_order,
                created_by_import_batch=batch,
            )
            group.full_clean()
            group.save()
            group_objects[key] = group
            created_groups.append(group)

        slot_objects = {
            (item.division.code, item.code): item
            for item in ParticipantSlot.objects.filter(division__season=season).select_related(
                "division", "group"
            )
        }
        created_slots: list[ParticipantSlot] = []
        for key, item in sorted(analysis.slots.items()):
            if key in slot_objects:
                continue
            group = (
                group_objects[(item.division_code, item.group_code)]
                if item.group_code
                else None
            )
            slot = ParticipantSlot(
                division=divisions[item.division_code],
                group=group,
                code=item.code,
                label=item.code,
                seed=item.number,
                created_by_import_batch=batch,
            )
            slot.full_clean()
            slot.save()
            slot_objects[key] = slot
            created_slots.append(slot)

        periods = {str(item.id): item for item in _periods(season)}
        created_games: list[Game] = []
        for item in sorted(analysis.games.values(), key=lambda row: row.code):
            game = Game(
                season=season,
                division=divisions[item.division_code],
                group=(
                    group_objects[(item.division_code, item.group_code)]
                    if item.group_code
                    else None
                ),
                code=item.code,
                stage=item.stage,
                round_number=item.round_number,
                date=item.target_date,
                period=periods[item.period_id],
                start_time=time.fromisoformat(item.start_time),
                venue_name=item.venue_name,
                home_slot=slot_objects[(item.division_code, item.home_slot_code)],
                away_slot=slot_objects[(item.division_code, item.away_slot_code)],
                leader_adjustable=item.leader_adjustable,
                status=Game.Status.SCHEDULED,
                created_by_import_batch=batch,
            )
            game.full_clean()
            game.save()
            created_games.append(game)

        before = {
            "season_version": season.version,
            "existing_game_count": analysis.summary["existing_game_count"],
        }
        season.version += 1
        season.save(update_fields=["version", "updated_at"])
        batch.status = ScheduleImportBatch.Status.CONFIRMED
        batch.confirmed_at = now
        batch.summary = {
            **analysis.summary,
            "error_count": 0,
            "warning_count": sum(
                issue.severity == ImportIssue.Severity.WARNING for issue in issues
            ),
            "confirmed_season_version": season.version,
            "created_group_ids": [str(item.id) for item in created_groups],
            "created_slot_ids": [str(item.id) for item in created_slots],
            "created_game_ids": [str(item.id) for item in created_games],
        }
        batch.save(update_fields=["status", "confirmed_at", "summary", "updated_at"])
        if source_draft is not None:
            source_draft.cells.all().delete()
            source_draft.version += 1
            source_draft.updated_by = actor
            source_draft.save(update_fields=["version", "updated_by", "updated_at"])
        AdminAuditLog.objects.create(
            actor=actor,
            action="SCHEDULE_IMPORT_CONFIRMED",
            object_type="ScheduleImportBatch",
            object_id=batch.id,
            before=before,
            after={
                "season_version": season.version,
                "created_group_ids": batch.summary["created_group_ids"],
                "created_slot_ids": batch.summary["created_slot_ids"],
                "created_game_ids": batch.summary["created_game_ids"],
            },
            metadata={
                "file_sha256": batch.file_sha256,
                "template_version": TEMPLATE_VERSION,
                "source_kind": batch.source_kind,
                "source_draft_version": batch.source_draft_version,
            },
        )
        return batch


def confirm_schedule_import(
    *,
    actor: Account,
    batch_id: object,
    expected_season_version: int,
    now: datetime | None = None,
) -> ScheduleImportBatch:
    _require_superadmin(actor)
    try:
        return _confirm_schedule_import(
            actor=actor,
            batch_id=batch_id,
            expected_season_version=expected_season_version,
            now=now or timezone.now(),
        )
    except IntegrityError as error:
        raise ScheduleImportError(
            "确认时检测到并发写入冲突，未创建任何赛程，请重新校验。",
            "CONCURRENT_CONFLICT",
        ) from error


def _blocker(code: str, message: str, count: int = 1) -> dict[str, object]:
    return {"code": code, "message": message, "count": count}


def _reset_preview(season: Season, *, now: datetime) -> dict[str, object]:
    batches = ScheduleImportBatch.objects.filter(
        season=season,
        status=ScheduleImportBatch.Status.CONFIRMED,
    ).order_by("created_at")
    games = Game.objects.filter(created_by_import_batch__in=batches)
    slots = ParticipantSlot.objects.filter(created_by_import_batch__in=batches)
    groups = CompetitionGroup.objects.filter(created_by_import_batch__in=batches)
    game_ids = games.values_list("id", flat=True)
    slot_ids = slots.values_list("id", flat=True)
    group_ids = groups.values_list("id", flat=True)
    blockers: list[dict[str, object]] = []

    if season.status != Season.Status.SETUP:
        blockers.append(_blocker("SEASON_NOT_SETUP", "只有准备中的赛季可以重置导入。"))
    batch_count = batches.count()
    if batch_count == 0:
        blockers.append(_blocker("NO_CONFIRMED_IMPORTS", "本赛季没有可重置的已确认导入批次。"))

    draw_count = DrawAssignment.objects.filter(season=season).count()
    if draw_count:
        blockers.append(_blocker("DRAW_ASSIGNMENTS_EXIST", "赛季已有抽签映射。", draw_count))
    started_count = sum(
        _slot_start(game.date, game.start_time, season.timezone) <= now
        for game in games.only("date", "start_time")
    )
    if started_count:
        blockers.append(
            _blocker("IMPORTED_GAMES_STARTED", "已有导入比赛到达开赛时间。", started_count)
        )
    result_count = games.filter(
        Q(home_score__isnull=False)
        | Q(away_score__isnull=False)
        | ~Q(status=Game.Status.SCHEDULED)
    ).count()
    if result_count:
        blockers.append(
            _blocker(
                "IMPORTED_GAMES_HAVE_RESULTS",
                "已有导入比赛包含赛果或非未赛状态。",
                result_count,
            )
        )
    request_count = RescheduleRequest.objects.filter(game_id__in=game_ids).count()
    if request_count:
        blockers.append(
            _blocker("RESCHEDULE_REQUESTS_EXIST", "导入比赛已有调赛申请。", request_count)
        )
    reservation_count = SlotReservation.objects.filter(
        season=season,
        status__in=[SlotReservation.Status.ACTIVE, SlotReservation.Status.CONVERTED],
    ).count()
    if reservation_count:
        blockers.append(
            _blocker(
                "RESERVATIONS_EXIST",
                "赛季已有有效或已转换的场次预留。",
                reservation_count,
            )
        )
    media_count = GameMediaAsset.objects.filter(game_id__in=game_ids).count()
    if media_count:
        blockers.append(_blocker("GAME_MEDIA_EXIST", "导入比赛已关联媒体文件。", media_count))

    external_game_slot_count = (
        Game.objects.filter(Q(home_slot_id__in=slot_ids) | Q(away_slot_id__in=slot_ids))
        .exclude(id__in=game_ids)
        .count()
    )
    if external_game_slot_count:
        blockers.append(
            _blocker(
                "IMPORTED_SLOTS_IN_USE",
                "导入签位已被非导入比赛引用。",
                external_game_slot_count,
            )
        )
    external_group_game_count = Game.objects.filter(group_id__in=group_ids).exclude(
        id__in=game_ids
    ).count()
    external_group_slot_count = ParticipantSlot.objects.filter(group_id__in=group_ids).exclude(
        id__in=slot_ids
    ).count()
    if external_group_game_count or external_group_slot_count:
        blockers.append(
            _blocker(
                "IMPORTED_GROUPS_IN_USE",
                "导入小组已被非导入签位或比赛引用。",
                external_group_game_count + external_group_slot_count,
            )
        )

    return {
        "season_id": str(season.id),
        "season_name": season.name,
        "season_version": season.version,
        "eligible": not blockers,
        "confirmed_batch_count": batch_count,
        "game_count": games.count(),
        "slot_count": slots.count(),
        "group_count": groups.count(),
        "batch_ids": [str(item) for item in batches.values_list("id", flat=True)],
        "blockers": blockers,
    }


def schedule_import_reset_preview(
    *, actor: Account, season: Season, now: datetime | None = None
) -> dict[str, object]:
    _require_superadmin(actor)
    return _reset_preview(season, now=now or timezone.now())


def _reset_schedule_imports(
    *,
    actor: Account,
    season_id: object,
    expected_season_version: int,
    season_name: str,
    now: datetime,
) -> dict[str, object]:
    with transaction.atomic():
        season = Season.objects.select_for_update().get(id=season_id)
        if season.version != expected_season_version:
            raise ScheduleImportError(
                "赛季已被其他操作修改，请重新执行重置预检。",
                "VERSION_CONFLICT",
            )
        if season_name.strip() != season.name:
            raise ScheduleImportError("输入的赛季名称不匹配。", "SEASON_NAME_MISMATCH")

        batches = list(
            ScheduleImportBatch.objects.select_for_update()
            .filter(season=season, status=ScheduleImportBatch.Status.CONFIRMED)
            .order_by("created_at")
        )
        list(Game.objects.select_for_update().filter(created_by_import_batch__in=batches))
        list(
            ParticipantSlot.objects.select_for_update().filter(
                created_by_import_batch__in=batches
            )
        )
        list(
            CompetitionGroup.objects.select_for_update().filter(
                created_by_import_batch__in=batches
            )
        )
        preview = _reset_preview(season, now=now)
        if preview["blockers"]:
            first = preview["blockers"][0]
            raise ScheduleImportError(str(first["message"]), "RESET_BLOCKED")

        games = Game.objects.filter(created_by_import_batch__in=batches)
        slots = ParticipantSlot.objects.filter(created_by_import_batch__in=batches)
        groups = CompetitionGroup.objects.filter(created_by_import_batch__in=batches)
        deleted = {
            "game_count": games.count(),
            "slot_count": slots.count(),
            "group_count": groups.count(),
            "batch_count": len(batches),
        }
        games.delete()
        slots.delete()
        groups.delete()

        before_version = season.version
        season.version += 1
        season.save(update_fields=["version", "updated_at"])
        for batch in batches:
            batch.status = ScheduleImportBatch.Status.ROLLED_BACK
            batch.summary = {
                **batch.summary,
                "rolled_back_at": now.isoformat(),
                "rolled_back_season_version": season.version,
            }
            batch.save(update_fields=["status", "summary", "updated_at"])
        AdminAuditLog.objects.create(
            actor=actor,
            action="SCHEDULE_IMPORTS_ROLLED_BACK",
            object_type="Season",
            object_id=season.id,
            before={"season_version": before_version, **deleted},
            after={"season_version": season.version, "remaining_imported_game_count": 0},
            metadata={"batch_ids": [str(batch.id) for batch in batches]},
        )
        return {
            "season_id": str(season.id),
            "season_version": season.version,
            "rolled_back_at": now,
            **deleted,
        }


def reset_schedule_imports(
    *,
    actor: Account,
    season_id: object,
    expected_season_version: int,
    season_name: str,
    now: datetime | None = None,
) -> dict[str, object]:
    _require_superadmin(actor)
    try:
        return _reset_schedule_imports(
            actor=actor,
            season_id=season_id,
            expected_season_version=expected_season_version,
            season_name=season_name,
            now=now or timezone.now(),
        )
    except ProtectedError as error:
        raise ScheduleImportError(
            "赛程对象仍被其他记录引用，未删除任何数据。",
            "RESET_PROTECTED",
        ) from error
