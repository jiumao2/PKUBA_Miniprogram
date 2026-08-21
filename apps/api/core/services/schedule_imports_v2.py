from __future__ import annotations

import hashlib
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import PurePath
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.validators import validate_slug
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from core.models import (
    Account,
    AdminAuditLog,
    CompetitionGroup,
    Division,
    DrawAssignment,
    Game,
    GameMediaAsset,
    ImportIssue,
    ParticipantSlot,
    Period,
    RescheduleRequest,
    ScheduleImportBatch,
    ScheduleSlotLock,
    Season,
    SlotReservation,
    Team,
    Venue,
)
from core.services.schedule_capacity import effective_capacity, effective_capacity_map

TEMPLATE_VERSION = "2.1.0"
EXPECTED_SHEETS = ["填写说明", "赛制定义", "比赛清单", "赛程网格", "特殊安排"]
STRUCTURE_HEADER_ROW = 5
STRUCTURE_START_ROW = 6
GAME_HEADER_ROW = 5
GAME_START_ROW = 6
GRID_HEADER_ROW = 6
GRID_START_ROW = 7
VENUE_START_COLUMN = 5
SPECIAL_HEADER_ROW = 5
SPECIAL_START_ROW = 6
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBERS = 2_000
MAX_DATA_ROWS = 5_000

STRUCTURE_HEADERS = [
    "组别代码",
    "小组代码",
    "小组名称",
    "小组排序",
    "签位代码",
    "签位名称",
    "种子序号",
]
GAME_HEADERS = [
    "比赛编号",
    "组别代码",
    "小组代码",
    "阶段",
    "轮次",
    "主方签位代码",
    "客方签位代码",
]
GRID_HEADERS = ["日期", "星期", "时段代码", "时段名称"]
SPECIAL_HEADERS = ["比赛编号", "日期", "时段代码", "实际开赛时间", "场地"]
STAGE_BY_LABEL = {label: value for value, label in Game.Stage.choices}
STAGE_BY_LABEL.update({value: value for value, _label in Game.Stage.choices})
STAGE_LABEL_BY_VALUE = dict(Game.Stage.choices)
WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


class ScheduleImportError(Exception):
    def __init__(self, message: str, code: str = "SCHEDULE_IMPORT_INVALID"):
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class ParsedIssue:
    severity: str
    code: str
    message: str
    cell: str = ""
    context: dict[str, object] | None = None

    def model_kwargs(self) -> dict[str, object]:
        values = asdict(self)
        values["cell"] = str(values["cell"])[:32]
        values["context"] = values["context"] or {}
        return values


@dataclass(frozen=True)
class GridRow:
    date: date
    period_id: str
    period_code: str
    period_name: str
    capacity: int


@dataclass(frozen=True)
class ParsedGroup:
    division_code: str
    code: str
    name: str
    sort_order: int
    cell: str


@dataclass(frozen=True)
class ParsedSlot:
    division_code: str
    group_code: str | None
    code: str
    label: str
    seed: int | None
    cell: str


@dataclass(frozen=True)
class ParsedGame:
    code: str
    division_code: str
    group_code: str | None
    stage: str
    round_number: int
    home_slot_code: str
    away_slot_code: str
    cell: str


@dataclass(frozen=True)
class Placement:
    code: str
    date: date
    period_code: str
    start_time: time
    venue_name: str
    standard_venue_id: str | None
    cell: str


@dataclass
class WorkbookAnalysis:
    groups: dict[tuple[str, str], ParsedGroup]
    slots: dict[tuple[str, str], ParsedSlot]
    games: dict[str, ParsedGame]
    placements: dict[str, Placement]
    summary: dict[str, object]


def _require_superadmin(actor: Account) -> None:
    if not actor.is_pkuba_superadmin:
        raise ScheduleImportError("只有超级管理员可以导入赛程。", "PERMISSION_DENIED")


def _venues(season: Season) -> list[Venue]:
    return list(Venue.objects.filter(season=season, active=True).order_by("sort_order", "name"))


def _periods(season: Season) -> list[Period]:
    return list(Period.objects.filter(season=season).order_by("sort_order", "start_time"))


def _grid_rows(season: Season) -> list[GridRow]:
    periods = _periods(season)
    dates = [
        season.starts_on + timedelta(days=offset)
        for offset in range((season.ends_on - season.starts_on).days + 1)
    ]
    capacities = effective_capacity_map(season=season, dates=dates, periods=periods)
    rows: list[GridRow] = []
    current = season.starts_on
    while current <= season.ends_on:
        for period in periods:
            capacity = capacities.get((current, period.id), 0)
            if capacity > 0:
                rows.append(
                    GridRow(
                        date=current,
                        period_id=str(period.id),
                        period_code=period.code.upper(),
                        period_name=period.name,
                        capacity=capacity,
                    )
                )
        current += timedelta(days=1)
    return rows


def schedule_import_readiness(season: Season) -> dict[str, object]:
    division_count = Division.objects.filter(season=season).count()
    team_count = Team.objects.filter(season=season, active=True).count()
    period_count = Period.objects.filter(season=season).count()
    venue_count = Venue.objects.filter(season=season, active=True).count()
    rows = _grid_rows(season)
    blockers: list[dict[str, object]] = []
    if season.status != Season.Status.SETUP:
        blockers.append(
            {
                "code": "SEASON_NOT_SETUP",
                "message": "只有准备中的赛季可以新增式导入赛程。",
                "count": 1,
            }
        )
    for count, code, message in (
        (division_count, "NO_DIVISIONS", "赛季尚未配置组别。"),
        (team_count, "NO_TEAMS", "赛季尚未配置球队。"),
        (period_count, "NO_PERIODS", "赛季尚未配置比赛时段。"),
        (venue_count, "NO_VENUES", "赛季尚未配置可用场地。"),
        (len(rows), "NO_CAPACITY", "赛季尚未配置任何开放容量。"),
    ):
        if count == 0:
            blockers.append({"code": code, "message": message, "count": 1})
    return {
        "season_id": str(season.id),
        "season_version": season.version,
        "ready": not blockers,
        "division_count": division_count,
        "team_count": team_count,
        "period_count": period_count,
        "venue_count": venue_count,
        "open_grid_row_count": len(rows),
        "existing_game_count": Game.objects.filter(season=season).count(),
        "blockers": blockers,
    }


def _require_ready(season: Season) -> None:
    readiness = schedule_import_readiness(season)
    if not readiness["ready"]:
        first = readiness["blockers"][0]
        raise ScheduleImportError(str(first["message"]), str(first["code"]))


def _style_title(sheet, title: str, end_column: int) -> None:
    end_column = max(7, end_column)
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
    sheet.cell(1, 1, title)
    sheet.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=18)
    sheet.cell(1, 1).alignment = Alignment(vertical="center")
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=end_column):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="087A45")


def _style_headers(sheet, row: int, end_column: int) -> None:
    for cell in sheet[row][:end_column]:
        cell.fill = PatternFill("solid", fgColor="21543C")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_schedule_template(season: Season) -> bytes:
    _require_ready(season)
    divisions = list(Division.objects.filter(season=season).order_by("sort_order", "name"))
    venues = _venues(season)
    rows = _grid_rows(season)
    slots = list(
        ParticipantSlot.objects.filter(division__season=season)
        .select_related("division", "group")
        .order_by("division__sort_order", "group__sort_order", "code")
    )

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    structure = workbook.create_sheet("赛制定义")
    games = workbook.create_sheet("比赛清单")
    schedule = workbook.create_sheet("赛程网格")
    special = workbook.create_sheet("特殊安排")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    instruction_rows = [
        ("格式版本", TEMPLATE_VERSION),
        ("赛季名称", season.name),
        (
            "最简填写",
            "通常只补“比赛清单”，再把比赛编号放入“赛程网格”；本文件已预填可复用的基础信息。",
        ),
        (
            "系统预填",
            "组别、已有签位、日期、星期、时段和场地由系统预填；容量只在赛季设置中配置一次。",
        ),
        (
            "特殊安排",
            "比赛时间微调或使用非标准场地时，不放入主网格，改在“特殊安排”中填写；每场仍只能出现一次。",
        ),
        ("比赛编号", "推荐使用人可读编号，如“男甲·A1vsA2”；同一对阵重复时再追加日期或短序号。"),
        ("新增规则", "本文件只新增比赛；已有比赛编号会阻止确认，未列出的旧比赛保持不变。"),
        ("调赛政策", "XLSX 不设置领队调赛政策；新比赛默认可调，之后在网页赛程编辑器修改。"),
        ("安全限制", "禁止宏、外部链接和单元格公式。上传只进入暂存批次，确认后才写数据库。"),
    ]
    for row_index, (label, value) in enumerate(instruction_rows, start=4):
        instructions.cell(row_index, 1, label).font = Font(bold=True)
        instructions.cell(row_index, 2, value).alignment = Alignment(wrap_text=True)
    instructions.cell(14, 1, "当前赛季组别代码").font = Font(bold=True)
    for column, value in enumerate(("代码", "名称", "球队数"), start=1):
        instructions.cell(15, column, value)
    for offset, division in enumerate(divisions, start=16):
        instructions.cell(offset, 1, division.code)
        instructions.cell(offset, 2, division.name)
        instructions.cell(offset, 3, division.teams.filter(active=True).count())
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 78
    instructions.column_dimensions["C"].width = 12
    _style_title(instructions, "PKUBA 赛季初赛程导入模板", 7)

    for column, header in enumerate(STRUCTURE_HEADERS, start=1):
        structure.cell(STRUCTURE_HEADER_ROW, column, header)
    for row_index, slot in enumerate(slots, start=STRUCTURE_START_ROW):
        values = (
            slot.division.code,
            slot.group.code if slot.group_id else "",
            slot.group.name if slot.group_id else "",
            slot.group.sort_order if slot.group_id else "",
            slot.code,
            slot.label,
            slot.seed if slot.seed is not None else "",
        )
        for column, value in enumerate(values, start=1):
            structure.cell(row_index, column, value).fill = PatternFill("solid", fgColor="F2F5F3")
    structure.freeze_panes = "A6"
    _style_title(structure, "赛制定义 · 小组与签位", len(STRUCTURE_HEADERS))
    _style_headers(structure, STRUCTURE_HEADER_ROW, len(STRUCTURE_HEADERS))

    for column, header in enumerate(GAME_HEADERS, start=1):
        games.cell(GAME_HEADER_ROW, column, header)
    stage_validation = DataValidation(
        type="list",
        formula1='"小组赛,循环赛,淘汰赛,半决赛,决赛,保级赛"',
        allow_blank=False,
    )
    stage_validation.error = "请选择模板提供的比赛阶段。"
    stage_validation.showErrorMessage = True
    games.add_data_validation(stage_validation)
    stage_validation.add(f"D{GAME_START_ROW}:D{GAME_START_ROW + MAX_DATA_ROWS - 1}")
    games.freeze_panes = "A6"
    _style_title(games, "比赛清单 · 定义本次新增比赛", len(GAME_HEADERS))
    _style_headers(games, GAME_HEADER_ROW, len(GAME_HEADERS))

    for column, header in enumerate(GRID_HEADERS, start=1):
        schedule.cell(GRID_HEADER_ROW, column, header)
    for offset, venue in enumerate(venues):
        column = VENUE_START_COLUMN + offset
        schedule.cell(GRID_HEADER_ROW, column, venue.name)
    for row_index, row in enumerate(rows, start=GRID_START_ROW):
        schedule.cell(row_index, 1, row.date)
        schedule.cell(row_index, 1).number_format = "yyyy-mm-dd"
        schedule.cell(row_index, 2, WEEKDAY_NAMES[row.date.weekday()])
        schedule.cell(row_index, 3, row.period_code)
        schedule.cell(row_index, 4, row.period_name)
        for column in range(1, 5):
            schedule.cell(row_index, column).fill = PatternFill("solid", fgColor="F2F5F3")
    grid_end_column = VENUE_START_COLUMN + len(venues) - 1
    grid_end_letter = get_column_letter(grid_end_column)
    grid_end_row = GRID_START_ROW + len(rows) - 1
    game_validation = DataValidation(
        type="list",
        formula1=f"'比赛清单'!$A${GAME_START_ROW}:$A${GAME_START_ROW + MAX_DATA_ROWS - 1}",
        allow_blank=True,
    )
    game_validation.error = "请选择比赛清单中的比赛编号。"
    game_validation.prompt = "每场比赛必须且只能出现一次。"
    game_validation.showErrorMessage = True
    game_validation.showInputMessage = True
    schedule.add_data_validation(game_validation)
    editable_range = f"E{GRID_START_ROW}:{grid_end_letter}{grid_end_row}"
    game_validation.add(editable_range)
    schedule.conditional_formatting.add(
        editable_range,
        FormulaRule(
            formula=[f"COUNTIF($E$7:${grid_end_letter}${grid_end_row},E7)>1"],
            fill=PatternFill("solid", fgColor="FCE8E6"),
        ),
    )
    schedule.freeze_panes = "E7"
    _style_title(schedule, "赛程网格 · 将比赛编号放入场地格", grid_end_column)
    _style_headers(schedule, GRID_HEADER_ROW, grid_end_column)
    for column, header in enumerate(SPECIAL_HEADERS, start=1):
        special.cell(SPECIAL_HEADER_ROW, column, header)
    special.freeze_panes = "A6"
    _style_title(special, "特殊安排 · 实际时间与自由场地", len(SPECIAL_HEADERS))
    _style_headers(special, SPECIAL_HEADER_ROW, len(SPECIAL_HEADERS))
    special_game_validation = DataValidation(
        type="list",
        formula1=f"'比赛清单'!$A${GAME_START_ROW}:$A${GAME_START_ROW + MAX_DATA_ROWS - 1}",
        allow_blank=True,
    )
    special.add_data_validation(special_game_validation)
    special_game_validation.add(
        f"A{SPECIAL_START_ROW}:A{SPECIAL_START_ROW + MAX_DATA_ROWS - 1}"
    )

    hints = (
        (structure, "已有小组和签位会预填；没有新增赛制时无需修改本表。", len(STRUCTURE_HEADERS)),
        (games, "按旧模板习惯使用人可读对阵编号；每场只需在清单定义一次。", len(GAME_HEADERS)),
        (schedule, "日期、时段和场地已预填；通过下拉或粘贴把每个对阵放入一个格。", grid_end_column),
        (special, "只填写不适合主网格的比赛；时段仍必须选择 P1 至 P6。", len(SPECIAL_HEADERS)),
    )
    for sheet, hint, end_column in hints:
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column)
        sheet.cell(3, 1, hint)
        sheet.cell(3, 1).font = Font(color="66766F", italic=True)
        sheet.cell(3, 1).alignment = Alignment(vertical="center", wrap_text=True)
        for cell in sheet[3][:end_column]:
            cell.fill = PatternFill("solid", fgColor="F2F5F3")

    for sheet in (structure, games, special):
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18
    structure.column_dimensions["F"].width = 28
    games.column_dimensions["A"].width = 20
    special.column_dimensions["A"].width = 22
    special.column_dimensions["B"].width = 14
    special.column_dimensions["C"].width = 14
    special.column_dimensions["D"].width = 18
    special.column_dimensions["E"].width = 28
    schedule.column_dimensions["A"].width = 13
    schedule.column_dimensions["B"].width = 9
    schedule.column_dimensions["C"].width = 14
    schedule.column_dimensions["D"].width = 18
    for column in range(VENUE_START_COLUMN, grid_end_column + 1):
        schedule.column_dimensions[get_column_letter(column)].width = 18

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _preflight_xlsx(content: bytes) -> None:
    if not content:
        raise ScheduleImportError("上传文件为空。", "EMPTY_FILE")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ScheduleImportError("上传文件超过 10 MB。", "FILE_TOO_LARGE")
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise ScheduleImportError("工作簿包含过多内部文件。", "XLSX_ZIP_BOMB")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise ScheduleImportError("工作簿解压后体积异常。", "XLSX_ZIP_BOMB")
            lower_names = [member.filename.lower() for member in members]
            if any("vbaproject.bin" in name for name in lower_names):
                raise ScheduleImportError("不允许上传含宏的工作簿。", "MACRO_FORBIDDEN")
            if any(name.startswith("xl/externallinks/") for name in lower_names):
                raise ScheduleImportError(
                    "不允许上传含外部链接的工作簿。",
                    "EXTERNAL_LINK_FORBIDDEN",
                )
            for member in members:
                if not member.filename.lower().endswith(".rels"):
                    continue
                relationships = archive.read(member).lower()
                has_external_target = (
                    b'targetmode="external"' in relationships
                    or b"targetmode='external'" in relationships
                )
                if has_external_target:
                    raise ScheduleImportError(
                        "不允许上传含外部链接的工作簿。",
                        "EXTERNAL_LINK_FORBIDDEN",
                    )
    except BadZipFile as error:
        raise ScheduleImportError("文件不是有效的 XLSX 工作簿。", "INVALID_XLSX") from error


def _issue(
    code: str,
    message: str,
    *,
    severity: str = ImportIssue.Severity.ERROR,
    cell: str = "",
    context: dict[str, object] | None = None,
) -> ParsedIssue:
    return ParsedIssue(severity=severity, code=code, message=message, cell=cell, context=context)


def _cell_date(cell: Cell) -> date | None:
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _cell_time(cell: Cell) -> time | None:
    value = cell.value
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, (int, float)) and 0 <= value < 1:
        total_minutes = round(float(value) * 24 * 60) % (24 * 60)
        return time(total_minutes // 60, total_minutes % 60)
    if isinstance(value, str):
        try:
            return time.fromisoformat(value.strip()).replace(second=0, microsecond=0)
        except ValueError:
            return None
    return None


def _read_text(
    cell: Cell,
    label: str,
    issues: list[ParsedIssue],
    *,
    required: bool = True,
) -> str | None:
    if cell.data_type == "f":
        return None
    if cell.value in (None, ""):
        if required:
            issues.append(
                _issue("REQUIRED_FIELD_MISSING", f"{label}不能为空。", cell=cell.coordinate)
            )
        return None
    if not isinstance(cell.value, str):
        issues.append(_issue("TEXT_FIELD_REQUIRED", f"{label}必须是文本。", cell=cell.coordinate))
        return None
    value = cell.value.strip()
    if required and not value:
        issues.append(_issue("REQUIRED_FIELD_MISSING", f"{label}不能为空。", cell=cell.coordinate))
        return None
    return value or None


def _read_positive_int(
    cell: Cell,
    label: str,
    issues: list[ParsedIssue],
    *,
    required: bool,
) -> int | None:
    if cell.data_type == "f":
        return None
    value = cell.value
    if value in (None, ""):
        if required:
            issues.append(
                _issue("REQUIRED_FIELD_MISSING", f"{label}不能为空。", cell=cell.coordinate)
            )
        return None
    try:
        number = int(value)
    except (TypeError, ValueError):
        number = 0
    if number <= 0 or str(value).strip() not in {str(number), f"{number}.0"}:
        issues.append(
            _issue(
                "POSITIVE_INTEGER_REQUIRED",
                f"{label}必须是正整数。",
                cell=cell.coordinate,
            )
        )
        return None
    return number


def _check_slug(value: str, label: str, cell: str, issues: list[ParsedIssue]) -> None:
    try:
        validate_slug(value)
    except ValidationError:
        issues.append(
            _issue(
                "INVALID_CODE",
                f"{label}只能包含字母、数字、下划线或连字符。",
                cell=cell,
            )
        )


def _validate_headers(sheet, row: int, expected: list[str]) -> None:
    actual = [sheet.cell(row, column).value for column in range(1, len(expected) + 1)]
    if actual != expected:
        raise ScheduleImportError(
            f"工作表“{sheet.title}”第 {row} 行字段名或顺序不正确。",
            "COLUMN_STRUCTURE_CHANGED",
        )


def _scan_formulas(workbook: Workbook, issues: list[ParsedIssue]) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    issues.append(
                        _issue(
                            "FORMULA_FORBIDDEN",
                            f"工作表“{sheet.title}”不允许单元格公式。",
                            cell=f"{sheet.title}!{cell.coordinate}",
                        )
                    )


def _parse_structure(
    season: Season,
    sheet,
    issues: list[ParsedIssue],
) -> tuple[
    dict[tuple[str, str], ParsedGroup],
    dict[tuple[str, str], ParsedSlot],
    dict[tuple[str, str], str],
    dict[tuple[str, str], str],
]:
    _validate_headers(sheet, STRUCTURE_HEADER_ROW, STRUCTURE_HEADERS)
    divisions = {item.code: item for item in Division.objects.filter(season=season)}
    existing_groups = {
        (item.division.code, item.code): item
        for item in CompetitionGroup.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division", "group"
        )
    }
    groups: dict[tuple[str, str], ParsedGroup] = {}
    slots: dict[tuple[str, str], ParsedSlot] = {}
    group_actions: dict[tuple[str, str], str] = {}
    slot_actions: dict[tuple[str, str], str] = {}
    if sheet.max_row - STRUCTURE_START_ROW + 1 > MAX_DATA_ROWS:
        issues.append(_issue("TOO_MANY_ROWS", "赛制定义超过 5000 行。"))
    last_row = min(sheet.max_row, STRUCTURE_START_ROW + MAX_DATA_ROWS - 1)
    for row in range(STRUCTURE_START_ROW, last_row + 1):
        cells = [sheet.cell(row, column) for column in range(1, len(STRUCTURE_HEADERS) + 1)]
        if all(cell.value in (None, "") for cell in cells):
            continue
        division_code = _read_text(cells[0], "组别代码", issues)
        group_code = _read_text(cells[1], "小组代码", issues, required=False)
        group_name = _read_text(cells[2], "小组名称", issues, required=bool(group_code))
        group_order = _read_positive_int(cells[3], "小组排序", issues, required=bool(group_code))
        slot_code = _read_text(cells[4], "签位代码", issues)
        slot_label = _read_text(cells[5], "签位名称", issues)
        seed = _read_positive_int(cells[6], "种子序号", issues, required=False)
        if not division_code or not slot_code or not slot_label:
            continue
        division = divisions.get(division_code)
        if division is None:
            issues.append(
                _issue(
                    "UNKNOWN_DIVISION",
                    f"数据库中不存在组别代码 {division_code}。",
                    cell=cells[0].coordinate,
                    context={"division_code": division_code},
                )
            )
            continue
        if not Team.objects.filter(season=season, division=division, active=True).exists():
            issues.append(
                _issue(
                    "DIVISION_HAS_NO_TEAMS",
                    f"组别 {division_code} 尚无有效球队。",
                    cell=cells[0].coordinate,
                )
            )
        if len(slot_code) > 32 or len(slot_label) > 80:
            issues.append(
                _issue(
                    "FIELD_TOO_LONG",
                    "签位代码或名称超过长度限制。",
                    cell=cells[4].coordinate,
                )
            )
            continue
        if group_code:
            _check_slug(group_code, "小组代码", cells[1].coordinate, issues)
            if len(group_code) > 16 or (group_name and len(group_name) > 40):
                issues.append(
                    _issue(
                        "FIELD_TOO_LONG",
                        "小组代码或名称超过长度限制。",
                        cell=cells[1].coordinate,
                    )
                )
                continue
            if not group_name or group_order is None:
                continue
            group_key = (division_code, group_code)
            parsed_group = ParsedGroup(
                division_code=division_code,
                code=group_code,
                name=group_name,
                sort_order=group_order,
                cell=cells[1].coordinate,
            )
            prior = groups.get(group_key)
            if prior and (prior.name, prior.sort_order) != (group_name, group_order):
                issues.append(
                    _issue(
                        "CONFLICTING_GROUP_DEFINITION",
                        f"小组 {division_code}/{group_code} 在表格中存在不同定义。",
                        cell=cells[1].coordinate,
                    )
                )
            else:
                groups[group_key] = parsed_group
            existing_group = existing_groups.get(group_key)
            if existing_group:
                group_actions[group_key] = "REFERENCE"
                if (existing_group.name, existing_group.sort_order) != (group_name, group_order):
                    issues.append(
                        _issue(
                            "EXISTING_GROUP_CONFLICT",
                            f"小组 {division_code}/{group_code} 与数据库现有定义不一致，不能覆盖。",
                            cell=cells[1].coordinate,
                        )
                    )
            else:
                group_actions[group_key] = "CREATE"
        elif any(cell.value not in (None, "") for cell in cells[2:4]):
            issues.append(
                _issue(
                    "GROUP_FIELDS_WITHOUT_CODE",
                    "填写小组名称或排序时必须同时填写小组代码。",
                    cell=cells[1].coordinate,
                )
            )

        slot_key = (division_code, slot_code)
        parsed_slot = ParsedSlot(
            division_code=division_code,
            group_code=group_code,
            code=slot_code,
            label=slot_label,
            seed=seed,
            cell=cells[4].coordinate,
        )
        if slot_key in slots:
            issues.append(
                _issue(
                    "DUPLICATE_SLOT_DEFINITION",
                    f"签位 {division_code}/{slot_code} 在表格中重复。",
                    cell=cells[4].coordinate,
                )
            )
            continue
        slots[slot_key] = parsed_slot
        existing_slot = existing_slots.get(slot_key)
        if existing_slot:
            slot_actions[slot_key] = "REFERENCE"
            existing_group_code = existing_slot.group.code if existing_slot.group_id else None
            if (
                existing_group_code,
                existing_slot.label,
                existing_slot.seed,
            ) != (group_code, slot_label, seed):
                issues.append(
                    _issue(
                        "EXISTING_SLOT_CONFLICT",
                        f"签位 {division_code}/{slot_code} 与数据库现有定义不一致，不能覆盖。",
                        cell=cells[4].coordinate,
                    )
                )
        else:
            slot_actions[slot_key] = "CREATE"
    return groups, slots, group_actions, slot_actions


def _parse_games(
    season: Season,
    sheet,
    groups: dict[tuple[str, str], ParsedGroup],
    slots: dict[tuple[str, str], ParsedSlot],
    issues: list[ParsedIssue],
) -> dict[str, ParsedGame]:
    _validate_headers(sheet, GAME_HEADER_ROW, GAME_HEADERS)
    divisions = {item.code: item for item in Division.objects.filter(season=season)}
    existing_groups = {
        (item.division.code, item.code): item
        for item in CompetitionGroup.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division", "group"
        )
    }
    available_group_keys = set(groups) | set(existing_groups)
    existing_codes = set(Game.objects.filter(season=season).values_list("code", flat=True))
    parsed: dict[str, ParsedGame] = {}
    if sheet.max_row - GAME_START_ROW + 1 > MAX_DATA_ROWS:
        issues.append(_issue("TOO_MANY_ROWS", "比赛清单超过 5000 行。"))
    last_row = min(sheet.max_row, GAME_START_ROW + MAX_DATA_ROWS - 1)
    for row in range(GAME_START_ROW, last_row + 1):
        cells = [sheet.cell(row, column) for column in range(1, len(GAME_HEADERS) + 1)]
        if all(cell.value in (None, "") for cell in cells):
            continue
        code = _read_text(cells[0], "比赛编号", issues)
        division_code = _read_text(cells[1], "组别代码", issues)
        group_code = _read_text(cells[2], "小组代码", issues, required=False)
        stage_label = _read_text(cells[3], "阶段", issues)
        round_number = _read_positive_int(cells[4], "轮次", issues, required=True)
        home_slot_code = _read_text(cells[5], "主方签位代码", issues)
        away_slot_code = _read_text(cells[6], "客方签位代码", issues)
        required_values = (
            code,
            division_code,
            stage_label,
            round_number,
            home_slot_code,
            away_slot_code,
        )
        if not all(required_values):
            continue
        if len(code) > 40:
            issues.append(
                _issue("FIELD_TOO_LONG", "比赛编号超过 40 个字符。", cell=cells[0].coordinate)
            )
            continue
        if code in parsed:
            issues.append(
                _issue(
                    "DUPLICATE_GAME_DEFINITION",
                    f"比赛编号 {code} 在清单中重复。",
                    cell=cells[0].coordinate,
                )
            )
            continue
        if code in existing_codes:
            issues.append(
                _issue(
                    "GAME_CODE_ALREADY_EXISTS",
                    f"比赛编号 {code} 已存在；新增式导入不允许更新比赛。",
                    cell=cells[0].coordinate,
                    context={"game_code": code},
                )
            )
        division = divisions.get(division_code)
        if division is None:
            issues.append(
                _issue(
                    "UNKNOWN_DIVISION",
                    f"数据库中不存在组别代码 {division_code}。",
                    cell=cells[1].coordinate,
                )
            )
            continue
        if not Team.objects.filter(season=season, division=division, active=True).exists():
            issues.append(
                _issue(
                    "DIVISION_HAS_NO_TEAMS",
                    f"组别 {division_code} 尚无有效球队。",
                    cell=cells[1].coordinate,
                )
            )
        stage = STAGE_BY_LABEL.get(stage_label)
        if stage is None:
            issues.append(
                _issue(
                    "INVALID_STAGE",
                    f"未知比赛阶段：{stage_label}。",
                    cell=cells[3].coordinate,
                )
            )
            continue
        if stage in {Game.Stage.GROUP, Game.Stage.ROUND_ROBIN} and not group_code:
            issues.append(
                _issue(
                    "GROUP_REQUIRED",
                    "小组赛或循环赛必须填写小组代码。",
                    cell=cells[2].coordinate,
                )
            )
        if group_code and (division_code, group_code) not in available_group_keys:
            issues.append(
                _issue(
                    "UNKNOWN_GROUP",
                    f"找不到小组 {division_code}/{group_code}。",
                    cell=cells[2].coordinate,
                )
            )
        if home_slot_code == away_slot_code:
            issues.append(
                _issue("SAME_PARTICIPANT", "主方和客方签位不能相同。", cell=cells[5].coordinate)
            )
        home_key = (division_code, home_slot_code)
        away_key = (division_code, away_slot_code)
        for key, label, cell in (
            (home_key, "主方", cells[5]),
            (away_key, "客方", cells[6]),
        ):
            if key not in slots and key not in existing_slots:
                issues.append(
                    _issue(
                        "UNKNOWN_SLOT",
                        f"找不到{label}签位 {division_code}/{key[1]}。",
                        cell=cell.coordinate,
                    )
                )
        if group_code and stage in {Game.Stage.GROUP, Game.Stage.ROUND_ROBIN}:
            for key, label, cell in (
                (home_key, "主方", cells[5]),
                (away_key, "客方", cells[6]),
            ):
                parsed_slot = slots.get(key)
                existing_slot = existing_slots.get(key)
                slot_group_code = (
                    parsed_slot.group_code
                    if parsed_slot
                    else existing_slot.group.code
                    if existing_slot and existing_slot.group_id
                    else None
                )
                if slot_group_code != group_code:
                    issues.append(
                        _issue(
                            "SLOT_GROUP_MISMATCH",
                            f"{label}签位不属于比赛填写的小组 {group_code}。",
                            cell=cell.coordinate,
                        )
                    )
        parsed[code] = ParsedGame(
            code=code,
            division_code=division_code,
            group_code=group_code,
            stage=stage,
            round_number=round_number,
            home_slot_code=home_slot_code,
            away_slot_code=away_slot_code,
            cell=cells[0].coordinate,
        )
    if not parsed:
        issues.append(_issue("NO_GAMES_IN_WORKBOOK", "比赛清单中没有可导入的比赛。"))
    return parsed


def _parse_grid(
    season: Season,
    sheet,
    special_sheet,
    games: dict[str, ParsedGame],
    issues: list[ParsedIssue],
) -> dict[str, Placement]:
    _validate_headers(sheet, GRID_HEADER_ROW, GRID_HEADERS)
    _validate_headers(special_sheet, SPECIAL_HEADER_ROW, SPECIAL_HEADERS)
    periods = {item.code.lower(): item for item in _periods(season)}
    venues = {item.name: item for item in _venues(season)}
    venue_columns: dict[int, Venue] = {}
    last_column = max(
        [
            column
            for column in range(VENUE_START_COLUMN, sheet.max_column + 1)
            if sheet.cell(GRID_HEADER_ROW, column).value not in (None, "")
        ],
        default=VENUE_START_COLUMN - 1,
    )
    if last_column < VENUE_START_COLUMN:
        issues.append(_issue("NO_VENUE_COLUMNS", "赛程网格没有场地列。"))
    seen_venue_names: set[str] = set()
    for column in range(VENUE_START_COLUMN, last_column + 1):
        name_cell = sheet.cell(GRID_HEADER_ROW, column)
        name = _read_text(name_cell, "场地名称", issues)
        if not name:
            continue
        if name in seen_venue_names:
            issues.append(
                _issue(
                    "DUPLICATE_VENUE_COLUMN",
                    f"场地名称 {name} 重复。",
                    cell=name_cell.coordinate,
                )
            )
            continue
        seen_venue_names.add(name)
        venue = venues.get(name)
        if venue is None:
            issues.append(
                _issue(
                    "UNKNOWN_VENUE",
                    f"数据库中不存在可用场地“{name}”。",
                    cell=name_cell.coordinate,
                )
            )
            continue
        venue_columns[column] = venue

    occurrences: dict[str, list[Placement]] = defaultdict(list)
    seen_rows: set[tuple[date, str]] = set()
    if sheet.max_row - GRID_START_ROW + 1 > MAX_DATA_ROWS:
        issues.append(_issue("TOO_MANY_ROWS", "赛程网格超过 5000 行。"))
    last_row = min(sheet.max_row, GRID_START_ROW + MAX_DATA_ROWS - 1)
    for row in range(GRID_START_ROW, last_row + 1):
        relevant = any(
            sheet.cell(row, column).value not in (None, "")
            for column in range(1, max(last_column, 4) + 1)
        )
        if not relevant:
            continue
        target_date = _cell_date(sheet.cell(row, 1))
        raw_period_code = _read_text(sheet.cell(row, 3), "时段代码", issues)
        period_code = raw_period_code.lower() if raw_period_code else None
        if target_date is None:
            issues.append(
                _issue(
                    "INVALID_DATE",
                    "日期必须是 Excel 日期或 YYYY-MM-DD。",
                    cell=sheet.cell(row, 1).coordinate,
                )
            )
        elif not season.starts_on <= target_date <= season.ends_on:
            issues.append(
                _issue(
                    "DATE_OUTSIDE_SEASON",
                    "比赛日期超出赛季起止日期。",
                    cell=sheet.cell(row, 1).coordinate,
                )
            )
        period = periods.get(period_code or "")
        if period is None and period_code:
            issues.append(
                _issue(
                    "UNKNOWN_PERIOD",
                    f"数据库中不存在时段 {period_code}。",
                    cell=sheet.cell(row, 3).coordinate,
                )
            )
        if target_date and period:
            key = (target_date, period.code)
            if key in seen_rows:
                issues.append(
                    _issue(
                        "DUPLICATE_GRID_ROW",
                        f"日期 {target_date.isoformat()} 与时段 {period.code} 重复。",
                        cell=sheet.cell(row, 1).coordinate,
                    )
                )
            seen_rows.add(key)
            capacity = effective_capacity(
                season_id=season.id,
                target_date=target_date,
                period_id=period.id,
            )
            if capacity <= 0:
                issues.append(
                    _issue(
                        "PERIOD_CLOSED",
                        f"{target_date.isoformat()} / {period.name} 未开放容量。",
                        cell=sheet.cell(row, 3).coordinate,
                    )
                )
        for column, venue in venue_columns.items():
            cell = sheet.cell(row, column)
            if cell.data_type == "f" or cell.value in (None, ""):
                continue
            if not isinstance(cell.value, str):
                issues.append(
                    _issue("INVALID_GAME_CODE", "比赛编号必须是文本。", cell=cell.coordinate)
                )
                continue
            code = cell.value.strip()
            if code not in games:
                issues.append(
                    _issue(
                        "UNKNOWN_GAME_CODE",
                        f"赛程网格中的比赛编号 {code} 不在比赛清单中。",
                        cell=cell.coordinate,
                        context={"game_code": code},
                    )
                )
                continue
            if target_date is None or period is None:
                continue
            occurrences[code].append(
                Placement(
                    code=code,
                    date=target_date,
                    period_code=period.code,
                    start_time=period.start_time,
                    venue_name=venue.name,
                    standard_venue_id=str(venue.id),
                    cell=cell.coordinate,
                )
            )

    special_last_row = min(
        special_sheet.max_row, SPECIAL_START_ROW + MAX_DATA_ROWS - 1
    )
    if special_sheet.max_row - SPECIAL_START_ROW + 1 > MAX_DATA_ROWS:
        issues.append(_issue("TOO_MANY_ROWS", "特殊安排超过 5000 行。"))
    for row in range(SPECIAL_START_ROW, special_last_row + 1):
        if not any(
            special_sheet.cell(row, column).value not in (None, "")
            for column in range(1, len(SPECIAL_HEADERS) + 1)
        ):
            continue
        code = _read_text(special_sheet.cell(row, 1), "比赛编号", issues)
        target_date = _cell_date(special_sheet.cell(row, 2))
        raw_period_code = _read_text(special_sheet.cell(row, 3), "时段代码", issues)
        period_code = raw_period_code.lower() if raw_period_code else None
        actual_time = _cell_time(special_sheet.cell(row, 4))
        venue_name = _read_text(special_sheet.cell(row, 5), "实际场地", issues)
        if code and code not in games:
            issues.append(
                _issue(
                    "UNKNOWN_GAME_CODE",
                    f"特殊安排中的比赛编号 {code} 不在比赛清单中。",
                    cell=special_sheet.cell(row, 1).coordinate,
                )
            )
        if target_date is None:
            issues.append(
                _issue(
                    "INVALID_DATE",
                    "日期必须是 Excel 日期或 YYYY-MM-DD。",
                    cell=special_sheet.cell(row, 2).coordinate,
                )
            )
        elif not season.starts_on <= target_date <= season.ends_on:
            issues.append(
                _issue(
                    "DATE_OUTSIDE_SEASON",
                    "比赛日期超出赛季起止日期。",
                    cell=special_sheet.cell(row, 2).coordinate,
                )
            )
        period = periods.get(period_code or "")
        if period is None and period_code:
            issues.append(
                _issue(
                    "UNKNOWN_PERIOD",
                    f"数据库中不存在时段 {raw_period_code}。",
                    cell=special_sheet.cell(row, 3).coordinate,
                )
            )
        if actual_time is None:
            issues.append(
                _issue(
                    "INVALID_START_TIME",
                    "实际开赛时间必须填写为 HH:MM。",
                    cell=special_sheet.cell(row, 4).coordinate,
                )
            )
        if target_date and period:
            capacity = effective_capacity(
                season_id=season.id,
                target_date=target_date,
                period_id=period.id,
            )
            if capacity <= 0:
                issues.append(
                    _issue(
                        "PERIOD_CLOSED",
                        f"{target_date.isoformat()} / {period.name} 未开放容量。",
                        cell=special_sheet.cell(row, 3).coordinate,
                    )
                )
        if (
            code in games
            and target_date is not None
            and period is not None
            and actual_time is not None
            and venue_name
        ):
            standard_venue = venues.get(venue_name)
            occurrences[code].append(
                Placement(
                    code=code,
                    date=target_date,
                    period_code=period.code,
                    start_time=actual_time,
                    venue_name=venue_name,
                    standard_venue_id=(
                        str(standard_venue.id) if standard_venue is not None else None
                    ),
                    cell=special_sheet.cell(row, 1).coordinate,
                )
            )
    placements: dict[str, Placement] = {}
    for code in games:
        found = occurrences.get(code, [])
        if not found:
            issues.append(
                _issue(
                    "MISSING_GAME_PLACEMENT",
                    f"比赛 {code} 未放入赛程网格。",
                    context={"game_code": code},
                )
            )
        elif len(found) > 1:
            issues.append(
                _issue(
                    "DUPLICATE_GAME_PLACEMENT",
            f"比赛 {code} 在赛程网格和特殊安排中共出现 {len(found)} 次。",
                    cell=found[0].cell,
                    context={"game_code": code, "count": len(found)},
                )
            )
        else:
            placements[code] = found[0]
    return placements


def _slot_reference(division_code: str, slot_code: str) -> str:
    return f"slot:{division_code}:{slot_code}"


def _validate_schedule_conflicts(
    season: Season,
    games: dict[str, ParsedGame],
    placements: dict[str, Placement],
    issues: list[ParsedIssue],
) -> None:
    existing_games = list(
        Game.objects.filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related("division", "period", "home_slot", "away_slot")
    )
    reservations = list(
        SlotReservation.objects.filter(season=season, status=SlotReservation.Status.ACTIVE)
        .select_related(
            "period",
            "venue",
            "request__game__division",
            "request__game__home_slot",
            "request__game__away_slot",
        )
    )
    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division"
        )
    }
    assignment_by_slot = {
        item.slot_id: item.team_id
        for item in DrawAssignment.objects.filter(season=season).select_related("slot", "team")
    }
    occupancy: dict[tuple[date, str, str], list[tuple[str, str, str]]] = defaultdict(list)
    participants: dict[tuple[date, str], dict[str, list[tuple[str, str]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    counts: Counter[tuple[date, str]] = Counter()
    keys_with_new_games: set[tuple[date, str]] = set()

    def add_existing_participants(game: Game, key: tuple[date, str], label: str) -> None:
        references: set[str] = set()
        for side in ("home", "away"):
            team_id = getattr(game, f"{side}_team_id")
            slot = getattr(game, f"{side}_slot")
            if team_id:
                references.add(f"team:{team_id}")
            if slot:
                references.add(_slot_reference(game.division.code, slot.code))
                assigned_team_id = assignment_by_slot.get(slot.id)
                if assigned_team_id:
                    references.add(f"team:{assigned_team_id}")
        for reference in references:
            participants[key][reference].append(("existing", label))

    for game in existing_games:
        key = (game.date, game.period.code)
        counts[key] += 1
        occupancy[(game.date, game.period.code, game.venue_name)].append(
            ("existing", game.code, "")
        )
        add_existing_participants(game, key, game.code)
    for reservation in reservations:
        key = (reservation.date, reservation.period.code)
        counts[key] += 1
        occupancy[(reservation.date, reservation.period.code, reservation.venue_name)].append(
            ("reservation", str(reservation.id), "")
        )
        request = getattr(reservation, "request", None)
        if request is not None:
            add_existing_participants(request.game, key, f"预留:{reservation.id}")

    for code, placement in placements.items():
        game = games[code]
        key = (placement.date, placement.period_code)
        keys_with_new_games.add(key)
        counts[key] += 1
        occupancy[(placement.date, placement.period_code, placement.venue_name)].append(
            ("new", code, placement.cell)
        )
        for slot_code in (game.home_slot_code, game.away_slot_code):
            slot_key = (game.division_code, slot_code)
            participants[key][_slot_reference(*slot_key)].append(("new", code))
            existing_slot = existing_slots.get(slot_key)
            assigned_team_id = assignment_by_slot.get(existing_slot.id) if existing_slot else None
            if assigned_team_id:
                participants[key][f"team:{assigned_team_id}"].append(("new", code))

    for key, occupants in occupancy.items():
        if len(occupants) <= 1 or not any(kind == "new" for kind, _label, _cell in occupants):
            continue
        new_occupant = next(item for item in occupants if item[0] == "new")
        issues.append(
            _issue(
                "VENUE_OCCUPIED",
                f"{key[0].isoformat()} / {key[1]} / {key[2]} 已有其他比赛或有效预留。",
                cell=new_occupant[2],
                context={"occupants": [label for _kind, label, _cell in occupants]},
            )
        )

    periods_by_code = {item.code: item for item in _periods(season)}
    for key in keys_with_new_games:
        capacity = effective_capacity(
            season_id=season.id,
            target_date=key[0],
            period_id=periods_by_code[key[1]].id,
        )
        if counts[key] > capacity:
            issues.append(
                _issue(
                    "CAPACITY_EXCEEDED",
                    f"{key[0].isoformat()} / {key[1]} 共占用 {counts[key]} 场，容量为 {capacity}。",
                    context={
                        "date": key[0].isoformat(),
                        "period_code": key[1],
                        "used": counts[key],
                        "capacity": capacity,
                    },
                )
            )
    for key, references in participants.items():
        for reference, occupants in references.items():
            labels = list(dict.fromkeys(label for _kind, label in occupants))
            if len(labels) > 1 and any(kind == "new" for kind, _label in occupants):
                issues.append(
                    _issue(
                        "PARTICIPANT_TIME_CONFLICT",
                        f"同一球队或签位在同一时段参加多场比赛：{', '.join(labels)}。",
                        context={
                            "date": key[0].isoformat(),
                            "period_code": key[1],
                            "participant": reference,
                            "games": labels,
                        },
                    )
                )


def _build_summary(
    season: Season,
    groups: dict[tuple[str, str], ParsedGroup],
    slots: dict[tuple[str, str], ParsedSlot],
    games: dict[str, ParsedGame],
    placements: dict[str, Placement],
    group_actions: dict[tuple[str, str], str],
    slot_actions: dict[tuple[str, str], str],
    issues: list[ParsedIssue],
) -> dict[str, object]:
    divisions = {item.code: item for item in Division.objects.filter(season=season)}
    existing_slots = {
        (item.division.code, item.code): item
        for item in ParticipantSlot.objects.filter(division__season=season).select_related(
            "division", "group"
        )
    }
    period_by_code = {item.code: item for item in _periods(season)}
    group_rows = [
        {
            "action": group_actions.get(key, "CREATE"),
            "division_code": item.division_code,
            "division_name": divisions[item.division_code].name,
            "code": item.code,
            "name": item.name,
            "sort_order": item.sort_order,
        }
        for key, item in sorted(groups.items())
        if item.division_code in divisions
    ]
    slot_rows = [
        {
            "action": slot_actions.get(key, "CREATE"),
            "division_code": item.division_code,
            "division_name": divisions[item.division_code].name,
            "group_code": item.group_code,
            "code": item.code,
            "label": item.label,
            "seed": item.seed,
        }
        for key, item in sorted(slots.items())
        if item.division_code in divisions
    ]
    existing_codes = set(Game.objects.filter(season=season).values_list("code", flat=True))
    game_rows: list[dict[str, object]] = []
    for code, item in sorted(games.items()):
        placement = placements.get(code)
        home = slots.get((item.division_code, item.home_slot_code)) or existing_slots.get(
            (item.division_code, item.home_slot_code)
        )
        away = slots.get((item.division_code, item.away_slot_code)) or existing_slots.get(
            (item.division_code, item.away_slot_code)
        )
        period = period_by_code.get(placement.period_code) if placement else None
        game_rows.append(
            {
                "action": "CONFLICT" if code in existing_codes else "CREATE",
                "code": code,
                "division_code": item.division_code,
                "division_name": divisions[item.division_code].name
                if item.division_code in divisions
                else "",
                "group_code": item.group_code,
                "stage": item.stage,
                "stage_name": STAGE_LABEL_BY_VALUE[item.stage],
                "round_number": item.round_number,
                "home_slot_code": item.home_slot_code,
                "home_slot_label": home.label if home else "",
                "away_slot_code": item.away_slot_code,
                "away_slot_label": away.label if away else "",
                "date": placement.date.isoformat() if placement else None,
                "period_code": placement.period_code if placement else None,
                "period_name": period.name if period else None,
                "nominal_start_time": (
                    period.start_time.isoformat(timespec="minutes") if period else None
                ),
                "start_time": (
                    placement.start_time.isoformat(timespec="minutes") if placement else None
                ),
                "venue_name": placement.venue_name if placement else None,
                "standard_venue_id": (
                    placement.standard_venue_id if placement else None
                ),
                "cell": placement.cell if placement else "",
            }
        )
    readiness = schedule_import_readiness(season)
    return {
        "existing_game_count": Game.objects.filter(season=season).count(),
        "new_group_count": sum(row["action"] == "CREATE" for row in group_rows),
        "referenced_group_count": sum(row["action"] == "REFERENCE" for row in group_rows),
        "new_slot_count": sum(row["action"] == "CREATE" for row in slot_rows),
        "referenced_slot_count": sum(row["action"] == "REFERENCE" for row in slot_rows),
        "new_game_count": sum(row["action"] == "CREATE" for row in game_rows),
        "groups": group_rows,
        "slots": slot_rows,
        "games": game_rows,
        "prerequisites": {
            "division_count": readiness["division_count"],
            "team_count": readiness["team_count"],
            "period_count": readiness["period_count"],
            "venue_count": readiness["venue_count"],
            "open_grid_row_count": readiness["open_grid_row_count"],
        },
        "error_count": sum(issue.severity == ImportIssue.Severity.ERROR for issue in issues),
        "warning_count": sum(issue.severity == ImportIssue.Severity.WARNING for issue in issues),
    }


def _analyze_workbook(
    season: Season,
    content: bytes,
) -> tuple[list[ParsedIssue], WorkbookAnalysis]:
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as error:
        raise ScheduleImportError("无法读取 XLSX 工作簿。", "INVALID_XLSX") from error
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ScheduleImportError("工作表名称、数量或顺序不正确。", "SHEET_STRUCTURE_CHANGED")
    if any(workbook[name].sheet_state != "visible" for name in EXPECTED_SHEETS):
        raise ScheduleImportError("五个工作表必须全部可见。", "SHEET_VISIBILITY_INVALID")
    if str(workbook["填写说明"]["B4"].value or "").strip() != TEMPLATE_VERSION:
        raise ScheduleImportError(
            f"不支持当前格式版本，请使用 {TEMPLATE_VERSION}。",
            "UNSUPPORTED_TEMPLATE_VERSION",
        )

    issues: list[ParsedIssue] = []
    readiness = schedule_import_readiness(season)
    for blocker in readiness["blockers"]:
        issues.append(_issue(str(blocker["code"]), str(blocker["message"])))
    _scan_formulas(workbook, issues)
    groups, slots, group_actions, slot_actions = _parse_structure(
        season, workbook["赛制定义"], issues
    )
    games = _parse_games(season, workbook["比赛清单"], groups, slots, issues)
    placements = _parse_grid(
        season,
        workbook["赛程网格"],
        workbook["特殊安排"],
        games,
        issues,
    )
    _validate_schedule_conflicts(season, games, placements, issues)
    summary = _build_summary(
        season,
        groups,
        slots,
        games,
        placements,
        group_actions,
        slot_actions,
        issues,
    )
    return issues, WorkbookAnalysis(groups, slots, games, placements, summary)


def validate_schedule_upload(
    *,
    actor: Account,
    season: Season,
    content: bytes,
    source_name: str,
    now: datetime | None = None,
) -> ScheduleImportBatch:
    del now
    _require_superadmin(actor)
    digest = hashlib.sha256(content).hexdigest()
    safe_name = get_valid_filename(PurePath(source_name).name) or "schedule.xlsx"
    batch = ScheduleImportBatch.objects.create(
        season=season,
        template_version=TEMPLATE_VERSION,
        file_key="",
        file_sha256=digest,
        uploaded_by=actor,
    )
    key = f"schedule-imports/{season.id}/{batch.id}/{safe_name}"
    batch.file_key = default_storage.save(key, ContentFile(content))
    try:
        issues, analysis = _analyze_workbook(season, content)
        summary = analysis.summary
    except ScheduleImportError as error:
        issues = [_issue(error.code, str(error))]
        readiness = schedule_import_readiness(season)
        summary = {
            "existing_game_count": readiness["existing_game_count"],
            "new_group_count": 0,
            "referenced_group_count": 0,
            "new_slot_count": 0,
            "referenced_slot_count": 0,
            "new_game_count": 0,
            "groups": [],
            "slots": [],
            "games": [],
            "prerequisites": {
                "division_count": readiness["division_count"],
                "team_count": readiness["team_count"],
                "period_count": readiness["period_count"],
                "venue_count": readiness["venue_count"],
                "open_grid_row_count": readiness["open_grid_row_count"],
            },
            "error_count": 1,
            "warning_count": 0,
        }
    ImportIssue.objects.bulk_create(
        [ImportIssue(batch=batch, **issue.model_kwargs()) for issue in issues]
    )
    batch.status = ScheduleImportBatch.Status.VALIDATED
    batch.summary = summary
    batch.save(update_fields=["file_key", "status", "summary", "updated_at"])
    return batch


def _lock_schedule_slots(season: Season, placements: dict[str, Placement]) -> None:
    keys = sorted({(item.date, item.period_code) for item in placements.values()})
    periods = {item.code: item for item in _periods(season)}
    for target_date, period_code in keys:
        period = periods[period_code]
        try:
            with transaction.atomic():
                ScheduleSlotLock.objects.get_or_create(
                    season=season,
                    date=target_date,
                    period=period,
                )
        except IntegrityError:
            pass
        ScheduleSlotLock.objects.select_for_update().get(
            season=season,
            date=target_date,
            period=period,
        )


def _game_start(game: Game, timezone_name: str) -> datetime:
    return datetime.combine(game.date, game.start_time, tzinfo=ZoneInfo(timezone_name))


def _confirm_schedule_import(
    *,
    actor: Account,
    batch_id: object,
    expected_season_version: int,
    now: datetime,
) -> ScheduleImportBatch:
    with transaction.atomic():
        batch = (
            ScheduleImportBatch.objects.select_for_update()
            .select_related("season")
            .get(id=batch_id)
        )
        if batch.status == ScheduleImportBatch.Status.CONFIRMED:
            return batch
        if batch.status == ScheduleImportBatch.Status.ROLLED_BACK:
            raise ScheduleImportError("该导入批次已回滚，不能再次确认。", "BATCH_ROLLED_BACK")
        if batch.status != ScheduleImportBatch.Status.VALIDATED:
            raise ScheduleImportError("导入批次尚未完成校验。", "BATCH_NOT_VALIDATED")
        if batch.issues.filter(severity=ImportIssue.Severity.ERROR).exists():
            raise ScheduleImportError("导入批次仍有错误，不能确认。", "BATCH_HAS_ERRORS")

        season = Season.objects.select_for_update().get(id=batch.season_id)
        if season.version != expected_season_version:
            raise ScheduleImportError("赛季已被其他操作修改，请重新校验文件。", "VERSION_CONFLICT")
        if season.status != Season.Status.SETUP:
            raise ScheduleImportError("只有准备中的赛季可以确认赛程导入。", "SEASON_NOT_SETUP")
        with default_storage.open(batch.file_key, "rb") as source:
            content = source.read()
        if hashlib.sha256(content).hexdigest() != batch.file_sha256:
            raise ScheduleImportError("上传原文件校验值不一致。", "FILE_HASH_MISMATCH")

        # The season row serializes all schedule imports for this season. These additional
        # locks make the resource boundary explicit and keep future writers honest.
        list(
            CompetitionGroup.objects.select_for_update().filter(division__season=season)
        )
        list(
            ParticipantSlot.objects.select_for_update().filter(division__season=season)
        )
        list(Game.objects.select_for_update().filter(season=season))
        list(
            SlotReservation.objects.select_for_update().filter(
                season=season,
                status__in=[SlotReservation.Status.ACTIVE, SlotReservation.Status.CONVERTED],
            )
        )

        issues, analysis = _analyze_workbook(season, content)
        errors = [issue for issue in issues if issue.severity == ImportIssue.Severity.ERROR]
        if errors:
            raise ScheduleImportError(
                f"确认前复核失败：{errors[0].message}",
                "REVALIDATION_FAILED",
            )
        _lock_schedule_slots(season, analysis.placements)

        divisions = {item.code: item for item in Division.objects.filter(season=season)}
        group_objects = {
            (item.division.code, item.code): item
            for item in CompetitionGroup.objects.filter(division__season=season).select_related(
                "division"
            )
        }
        created_groups: list[CompetitionGroup] = []
        for key, item in sorted(analysis.groups.items()):
            if key in group_objects:
                continue
            group = CompetitionGroup(
                division=divisions[item.division_code],
                code=item.code,
                name=item.name,
                sort_order=item.sort_order,
                created_by_import_batch=batch,
            )
            group.full_clean()
            group.save()
            group_objects[key] = group
            created_groups.append(group)

        slot_objects = {
            (item.division.code, item.code): item
            for item in ParticipantSlot.objects.filter(division__season=season).select_related(
                "division", "group"
            )
        }
        created_slots: list[ParticipantSlot] = []
        for key, item in sorted(analysis.slots.items()):
            if key in slot_objects:
                continue
            group = (
                group_objects[(item.division_code, item.group_code)]
                if item.group_code
                else None
            )
            slot = ParticipantSlot(
                division=divisions[item.division_code],
                group=group,
                code=item.code,
                label=item.label,
                seed=item.seed,
                created_by_import_batch=batch,
            )
            slot.full_clean()
            slot.save()
            slot_objects[key] = slot
            created_slots.append(slot)

        periods = {item.code: item for item in _periods(season)}
        created_games: list[Game] = []
        for code, item in sorted(analysis.games.items()):
            placement = analysis.placements[code]
            game = Game(
                season=season,
                division=divisions[item.division_code],
                group=(
                    group_objects[(item.division_code, item.group_code)]
                    if item.group_code
                    else None
                ),
                code=item.code,
                stage=item.stage,
                round_number=item.round_number,
                date=placement.date,
                period=periods[placement.period_code],
                start_time=placement.start_time,
                venue_name=placement.venue_name,
                home_slot=slot_objects[(item.division_code, item.home_slot_code)],
                away_slot=slot_objects[(item.division_code, item.away_slot_code)],
                leader_adjustable=True,
                status=Game.Status.SCHEDULED,
                created_by_import_batch=batch,
            )
            game.full_clean()
            game.save()
            created_games.append(game)

        before = {
            "season_version": season.version,
            "existing_game_count": analysis.summary["existing_game_count"],
        }
        season.version += 1
        season.save(update_fields=["version", "updated_at"])
        batch.status = ScheduleImportBatch.Status.CONFIRMED
        batch.confirmed_at = now
        batch.summary = {
            **analysis.summary,
            "error_count": 0,
            "warning_count": sum(
                issue.severity == ImportIssue.Severity.WARNING for issue in issues
            ),
            "confirmed_season_version": season.version,
            "created_group_ids": [str(item.id) for item in created_groups],
            "created_slot_ids": [str(item.id) for item in created_slots],
            "created_game_ids": [str(item.id) for item in created_games],
        }
        batch.save(update_fields=["status", "confirmed_at", "summary", "updated_at"])
        AdminAuditLog.objects.create(
            actor=actor,
            action="SCHEDULE_IMPORT_CONFIRMED",
            object_type="ScheduleImportBatch",
            object_id=batch.id,
            before=before,
            after={
                "season_version": season.version,
                "created_group_ids": batch.summary["created_group_ids"],
                "created_slot_ids": batch.summary["created_slot_ids"],
                "created_game_ids": batch.summary["created_game_ids"],
            },
            metadata={"file_sha256": batch.file_sha256, "template_version": TEMPLATE_VERSION},
        )
        return batch


def confirm_schedule_import(
    *,
    actor: Account,
    batch_id: object,
    expected_season_version: int,
    now: datetime | None = None,
) -> ScheduleImportBatch:
    _require_superadmin(actor)
    try:
        return _confirm_schedule_import(
            actor=actor,
            batch_id=batch_id,
            expected_season_version=expected_season_version,
            now=now or timezone.now(),
        )
    except IntegrityError as error:
        raise ScheduleImportError(
            "确认时检测到并发写入冲突，未创建任何赛程，请重新校验。",
            "CONCURRENT_CONFLICT",
        ) from error


def _blocker(code: str, message: str, count: int = 1) -> dict[str, object]:
    return {"code": code, "message": message, "count": count}


def _reset_preview(season: Season, *, now: datetime) -> dict[str, object]:
    batches = ScheduleImportBatch.objects.filter(
        season=season,
        status=ScheduleImportBatch.Status.CONFIRMED,
    ).order_by("created_at")
    games = Game.objects.filter(created_by_import_batch__in=batches)
    slots = ParticipantSlot.objects.filter(created_by_import_batch__in=batches)
    groups = CompetitionGroup.objects.filter(created_by_import_batch__in=batches)
    game_ids = games.values_list("id", flat=True)
    slot_ids = slots.values_list("id", flat=True)
    group_ids = groups.values_list("id", flat=True)
    blockers: list[dict[str, object]] = []

    if season.status != Season.Status.SETUP:
        blockers.append(_blocker("SEASON_NOT_SETUP", "只有准备中的赛季可以重置导入。"))
    batch_count = batches.count()
    if batch_count == 0:
        blockers.append(_blocker("NO_CONFIRMED_IMPORTS", "本赛季没有可重置的已确认导入批次。"))

    draw_count = DrawAssignment.objects.filter(season=season).count()
    if draw_count:
        blockers.append(_blocker("DRAW_ASSIGNMENTS_EXIST", "赛季已有抽签映射。", draw_count))
    started_count = sum(
        _game_start(game, season.timezone) <= now
        for game in games
    )
    if started_count:
        blockers.append(
            _blocker("IMPORTED_GAMES_STARTED", "已有导入比赛到达开赛时间。", started_count)
        )
    result_count = games.filter(
        Q(home_score__isnull=False)
        | Q(away_score__isnull=False)
        | ~Q(status=Game.Status.SCHEDULED)
    ).count()
    if result_count:
        blockers.append(
            _blocker(
                "IMPORTED_GAMES_HAVE_RESULTS",
                "已有导入比赛包含赛果或非未赛状态。",
                result_count,
            )
        )
    request_count = RescheduleRequest.objects.filter(game_id__in=game_ids).count()
    if request_count:
        blockers.append(
            _blocker("RESCHEDULE_REQUESTS_EXIST", "导入比赛已有调赛申请。", request_count)
        )
    reservation_count = SlotReservation.objects.filter(
        season=season,
        status__in=[SlotReservation.Status.ACTIVE, SlotReservation.Status.CONVERTED],
    ).count()
    if reservation_count:
        blockers.append(
            _blocker(
                "RESERVATIONS_EXIST",
                "赛季已有有效或已转换的场次预留。",
                reservation_count,
            )
        )
    media_count = GameMediaAsset.objects.filter(game_id__in=game_ids).count()
    if media_count:
        blockers.append(_blocker("GAME_MEDIA_EXIST", "导入比赛已关联媒体文件。", media_count))

    external_game_slot_count = (
        Game.objects.filter(Q(home_slot_id__in=slot_ids) | Q(away_slot_id__in=slot_ids))
        .exclude(id__in=game_ids)
        .count()
    )
    if external_game_slot_count:
        blockers.append(
            _blocker(
                "IMPORTED_SLOTS_IN_USE",
                "导入签位已被非导入比赛引用。",
                external_game_slot_count,
            )
        )
    external_group_game_count = Game.objects.filter(group_id__in=group_ids).exclude(
        id__in=game_ids
    ).count()
    external_group_slot_count = ParticipantSlot.objects.filter(group_id__in=group_ids).exclude(
        id__in=slot_ids
    ).count()
    if external_group_game_count or external_group_slot_count:
        blockers.append(
            _blocker(
                "IMPORTED_GROUPS_IN_USE",
                "导入小组已被非导入签位或比赛引用。",
                external_group_game_count + external_group_slot_count,
            )
        )

    return {
        "season_id": str(season.id),
        "season_name": season.name,
        "season_version": season.version,
        "eligible": not blockers,
        "confirmed_batch_count": batch_count,
        "game_count": games.count(),
        "slot_count": slots.count(),
        "group_count": groups.count(),
        "batch_ids": [str(item) for item in batches.values_list("id", flat=True)],
        "blockers": blockers,
    }


def schedule_import_reset_preview(
    *, actor: Account, season: Season, now: datetime | None = None
) -> dict[str, object]:
    _require_superadmin(actor)
    return _reset_preview(season, now=now or timezone.now())


def _reset_schedule_imports(
    *,
    actor: Account,
    season_id: object,
    expected_season_version: int,
    season_name: str,
    now: datetime,
) -> dict[str, object]:
    with transaction.atomic():
        season = Season.objects.select_for_update().get(id=season_id)
        if season.version != expected_season_version:
            raise ScheduleImportError(
                "赛季已被其他操作修改，请重新执行重置预检。",
                "VERSION_CONFLICT",
            )
        if season_name.strip() != season.name:
            raise ScheduleImportError("输入的赛季名称不匹配。", "SEASON_NAME_MISMATCH")

        batches = list(
            ScheduleImportBatch.objects.select_for_update()
            .filter(season=season, status=ScheduleImportBatch.Status.CONFIRMED)
            .order_by("created_at")
        )
        list(Game.objects.select_for_update().filter(created_by_import_batch__in=batches))
        list(
            ParticipantSlot.objects.select_for_update().filter(
                created_by_import_batch__in=batches
            )
        )
        list(
            CompetitionGroup.objects.select_for_update().filter(
                created_by_import_batch__in=batches
            )
        )
        preview = _reset_preview(season, now=now)
        if preview["blockers"]:
            first = preview["blockers"][0]
            raise ScheduleImportError(str(first["message"]), "RESET_BLOCKED")

        games = Game.objects.filter(created_by_import_batch__in=batches)
        slots = ParticipantSlot.objects.filter(created_by_import_batch__in=batches)
        groups = CompetitionGroup.objects.filter(created_by_import_batch__in=batches)
        deleted = {
            "game_count": games.count(),
            "slot_count": slots.count(),
            "group_count": groups.count(),
            "batch_count": len(batches),
        }
        games.delete()
        slots.delete()
        groups.delete()

        before_version = season.version
        season.version += 1
        season.save(update_fields=["version", "updated_at"])
        for batch in batches:
            batch.status = ScheduleImportBatch.Status.ROLLED_BACK
            batch.summary = {
                **batch.summary,
                "rolled_back_at": now.isoformat(),
                "rolled_back_season_version": season.version,
            }
            batch.save(update_fields=["status", "summary", "updated_at"])
        AdminAuditLog.objects.create(
            actor=actor,
            action="SCHEDULE_IMPORTS_ROLLED_BACK",
            object_type="Season",
            object_id=season.id,
            before={"season_version": before_version, **deleted},
            after={"season_version": season.version, "remaining_imported_game_count": 0},
            metadata={"batch_ids": [str(batch.id) for batch in batches]},
        )
        return {
            "season_id": str(season.id),
            "season_version": season.version,
            "rolled_back_at": now,
            **deleted,
        }


def reset_schedule_imports(
    *,
    actor: Account,
    season_id: object,
    expected_season_version: int,
    season_name: str,
    now: datetime | None = None,
) -> dict[str, object]:
    _require_superadmin(actor)
    try:
        return _reset_schedule_imports(
            actor=actor,
            season_id=season_id,
            expected_season_version=expected_season_version,
            season_name=season_name,
            now=now or timezone.now(),
        )
    except ProtectedError as error:
        raise ScheduleImportError(
            "赛程对象仍被其他记录引用，未删除任何数据。",
            "RESET_PROTECTED",
        ) from error
