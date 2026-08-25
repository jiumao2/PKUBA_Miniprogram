from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import PurePath
from zipfile import BadZipFile, ZipFile

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.serializers.json import DjangoJSONEncoder
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from core.models import (
    Account,
    AdminAuditLog,
    CompetitionGroup,
    Division,
    DrawAssignment,
    Game,
    ParticipantSlot,
    RescheduleRequest,
    RosterImportBatch,
    RosterImportIssue,
    RosterPlayer,
    ScheduleSlotFamily,
    Season,
    SeasonLeaderBinding,
    Team,
    TeamConfirmation,
)

TEMPLATE_VERSION = "1.0.0"
HEADERS = ["球队标准名称*", "球员姓名*", "球衣号码（选填）"]
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ZIP_MEMBERS = 2_000
MAX_DATA_ROWS = 5_000
JERSEY_PATTERN = re.compile(r"^(?:00|0|[1-9][0-9]?)$")
INVALID_SHEET_CHARACTERS = re.compile(r"[\\/*?:\[\]]")
PKUBA_RED = "B52328"
PKUBA_RED_DARK = "8F1D22"
WARM_GRAY = "F5F2ED"
LINE = "D9D2C9"
INK = "2C2925"
MUTED = "6D675F"


class RosterManagementError(Exception):
    def __init__(self, message: str, code: str = "ROSTER_INVALID"):
        self.code = code
        super().__init__(message)


@dataclass(frozen=True)
class ParsedIssue:
    severity: str
    code: str
    message: str
    cell: str = ""
    context: dict[str, object] | None = None

    def model_kwargs(self) -> dict[str, object]:
        values = asdict(self)
        values["cell"] = str(values["cell"])[:64]
        values["context"] = values["context"] or {}
        return values


def _issue(
    code: str,
    message: str,
    *,
    severity: str = RosterImportIssue.Severity.ERROR,
    cell: str = "",
    context: dict[str, object] | None = None,
) -> ParsedIssue:
    return ParsedIssue(severity, code, message, cell, context)


def _require_superadmin(actor: Account) -> None:
    if not actor.is_pkuba_superadmin:
        raise RosterManagementError("只有超级管理员可以维护球队与名单。", "PERMISSION_DENIED")


def _normalize_text(value: object) -> str:
    return " ".join(unicodedata.normalize("NFKC", str(value or "")).strip().split())


def _validate_name(value: object, label: str, max_length: int) -> str:
    normalized = _normalize_text(value)
    if not normalized:
        raise RosterManagementError(f"{label}不能为空。", "REQUIRED_FIELD")
    if len(normalized) > max_length:
        raise RosterManagementError(f"{label}不能超过 {max_length} 个字符。", "FIELD_TOO_LONG")
    return normalized


def _normalize_jersey(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _normalize_text(value)


def _validate_jersey(value: object) -> str:
    jersey = _normalize_jersey(value)
    if jersey and not JERSEY_PATTERN.fullmatch(jersey):
        raise RosterManagementError(
            "球衣号码只允许 0–99 或 00，也可以留空。", "INVALID_JERSEY_NUMBER"
        )
    return jersey


def _sheet_titles(divisions: list[Division]) -> dict[str, str]:
    used = {"填写说明".casefold()}
    result: dict[str, str] = {}
    for division in divisions:
        base = INVALID_SHEET_CHARACTERS.sub("-", division.name).strip(" '") or division.code
        base = base[:31]
        title = base
        suffix = 2
        while title.casefold() in used:
            marker = f"-{suffix}"
            title = f"{base[: 31 - len(marker)]}{marker}"
            suffix += 1
        used.add(title.casefold())
        result[str(division.id)] = title
    return result


def generate_roster_template(season: Season) -> bytes:
    divisions = list(season.divisions.order_by("sort_order", "name"))
    if not divisions:
        raise RosterManagementError("请先在“赛季与组别”中创建至少一个组别。", "NO_DIVISIONS")
    titles = _sheet_titles(divisions)
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "填写说明"
    instructions.sheet_view.showGridLines = False
    instructions.freeze_panes = "A7"
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 76
    instructions.merge_cells("A1:B1")
    instructions["A1"] = f"{season.name} · 球队与名单填写模板"
    instructions["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor=PKUBA_RED)
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 38
    instructions["A3"] = "赛季"
    instructions["B3"] = season.name
    instructions["A4"] = "格式版本"
    instructions["B4"] = TEMPLATE_VERSION
    instructions["A5"] = "填写原则"
    instructions["B5"] = (
        "每名球员占一行；球队标准名称和球员姓名必填。"
        "同一球队的名称必须逐行完全一致，不要混用全称和缩写。"
    )
    instructions["A6"] = "号码规则"
    instructions["B6"] = "号码可留空；填写时只允许 0–99 或 00。号码列已设置为文本格式。"
    for row in range(3, 7):
        instructions.cell(row, 1).font = Font(bold=True, color=PKUBA_RED_DARK)
        instructions.cell(row, 1).fill = PatternFill("solid", fgColor=WARM_GRAY)
        instructions.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    start = 8
    instructions.cell(start, 1, "组别工作表")
    instructions.cell(start, 2, "应填写内容")
    for cell in instructions[start]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=PKUBA_RED_DARK)
    for index, division in enumerate(divisions, start=start + 1):
        instructions.cell(index, 1, titles[str(division.id)])
        instructions.cell(index, 2, f"{division.name}球队及球员；不要把其他组别球队填入本页。")
    example_row = start + len(divisions) + 3
    instructions.cell(example_row, 1, "示例（仅说明，不导入）")
    instructions.cell(example_row, 1).font = Font(bold=True, color=PKUBA_RED_DARK)
    instructions.cell(example_row + 1, 1, HEADERS[0])
    instructions.cell(example_row + 1, 2, HEADERS[1])
    instructions.cell(example_row + 1, 3, HEADERS[2])
    instructions.cell(example_row + 2, 1, "示例大学篮球队")
    instructions.cell(example_row + 2, 2, "张同学")
    instructions.cell(example_row + 2, 3, "00")
    instructions.column_dimensions["C"].width = 20
    instructions.auto_filter.ref = f"A{start}:B{start + len(divisions)}"

    thin = Side(style="thin", color=LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for division in divisions:
        sheet = workbook.create_sheet(titles[str(division.id)])
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:C1000"
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 22
        sheet.column_dimensions["C"].width = 20
        sheet.row_dimensions[1].height = 28
        for column, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=PKUBA_RED)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        for row in range(2, 1001):
            sheet.cell(row, 3).number_format = "@"
        required = FormulaRule(
            formula=['AND(COUNTA($A2:$C2)>0,OR($A2="",$B2=""))'],
            fill=PatternFill("solid", fgColor="FCE8E6"),
        )
        sheet.conditional_formatting.add("A2:B1000", required)
        jersey_validation = DataValidation(
            type="custom",
            formula1=(
                'OR(C2="",C2="00",C2="0",AND(LEN(C2)<=2,LEFT(C2,1)<>"0",'
                "ISNUMBER(--C2),--C2>=1,--C2<=99))"
            ),
            allow_blank=True,
            error="只允许 0–99、00 或留空。",
            errorTitle="球衣号码格式不正确",
            prompt="号码列为文本格式，可填写 00。",
            promptTitle="球衣号码（选填）",
            showErrorMessage=True,
            showInputMessage=True,
        )
        sheet.add_data_validation(jersey_validation)
        jersey_validation.add("C2:C1000")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _preflight_xlsx(content: bytes, source_name: str) -> None:
    if not PurePath(source_name).name.lower().endswith(".xlsx"):
        raise RosterManagementError("只允许上传 .xlsx 文件。", "INVALID_FILE_TYPE")
    if not content:
        raise RosterManagementError("上传文件为空。", "EMPTY_FILE")
    if len(content) > MAX_UPLOAD_BYTES:
        raise RosterManagementError("上传文件超过 10 MB。", "FILE_TOO_LARGE")
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise RosterManagementError("工作簿包含过多内部文件。", "XLSX_ZIP_BOMB")
            if sum(item.file_size for item in members) > MAX_UNCOMPRESSED_BYTES:
                raise RosterManagementError("工作簿解压后体积异常。", "XLSX_ZIP_BOMB")
            names = [item.filename.lower() for item in members]
            if any("vbaproject.bin" in name for name in names):
                raise RosterManagementError("不允许上传含宏的工作簿。", "MACRO_FORBIDDEN")
            if any(name.startswith("xl/externallinks/") for name in names):
                raise RosterManagementError(
                    "不允许上传含外部链接的工作簿。", "EXTERNAL_LINK_FORBIDDEN"
                )
            for member in members:
                if not member.filename.lower().endswith(".rels"):
                    continue
                relationships = archive.read(member).lower()
                if (
                    b'targetmode="external"' in relationships
                    or b"targetmode='external'" in relationships
                ):
                    raise RosterManagementError(
                        "不允许上传含外部链接的工作簿。", "EXTERNAL_LINK_FORBIDDEN"
                    )
    except BadZipFile as error:
        raise RosterManagementError("文件不是有效的 XLSX 工作簿。", "INVALID_XLSX") from error


def _name_key(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).casefold()
    value = re.sub(r"[\s\-_—·•,，.。()（）\[\]【】]", "", value)
    for suffix in ("男子篮球队", "女子篮球队", "男篮", "女篮", "篮球队", "代表队", "球队"):
        if value.endswith(suffix):
            value = value[: -len(suffix)]
            break
    return value


def _resolution_key(division_id: str, source_name: str) -> str:
    digest = hashlib.sha256(f"{division_id}\0{source_name}".encode()).hexdigest()[:16]
    return f"{division_id}:{digest}"


def _parse_workbook(
    season: Season,
    content: bytes,
    *,
    resolutions: dict[str, str] | None = None,
) -> tuple[list[ParsedIssue], dict[str, object]]:
    divisions = list(season.divisions.order_by("sort_order", "name"))
    titles = _sheet_titles(divisions)
    expected_sheets = ["填写说明", *[titles[str(item.id)] for item in divisions]]
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as error:
        raise RosterManagementError("无法读取 XLSX 工作簿。", "INVALID_XLSX") from error
    if workbook.sheetnames != expected_sheets:
        raise RosterManagementError(
            "工作表名称、数量或顺序与当前赛季模板不一致。", "SHEET_STRUCTURE_CHANGED"
        )
    if any(workbook[name].sheet_state != "visible" for name in expected_sheets):
        raise RosterManagementError("模板中的全部工作表必须可见。", "SHEET_VISIBILITY_INVALID")
    if _normalize_text(workbook["填写说明"]["B4"].value) != TEMPLATE_VERSION:
        raise RosterManagementError(
            f"不支持当前格式版本，请重新下载 {TEMPLATE_VERSION} 模板。",
            "UNSUPPORTED_TEMPLATE_VERSION",
        )
    issues: list[ParsedIssue] = []
    for blocker in roster_import_readiness(season)["blockers"]:
        issues.append(_issue(str(blocker["code"]), str(blocker["message"])))
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    issues.append(
                        _issue(
                            "FORMULA_FORBIDDEN",
                            f"工作表“{sheet.title}”不允许单元格公式。",
                            cell=f"{sheet.title}!{cell.coordinate}",
                        )
                    )
    resolutions = resolutions or {}
    rows: list[dict[str, object]] = []
    resolution_options: dict[str, dict[str, object]] = {}
    for division in divisions:
        title = titles[str(division.id)]
        sheet = workbook[title]
        actual_headers = [sheet.cell(1, column).value for column in range(1, 4)]
        if actual_headers != HEADERS:
            raise RosterManagementError(
                f"工作表“{title}”的字段名或顺序不正确。", "COLUMN_STRUCTURE_CHANGED"
            )
        if sheet.max_row - 1 > MAX_DATA_ROWS:
            raise RosterManagementError(
                f"工作表“{title}”超过 {MAX_DATA_ROWS} 行上限。", "TOO_MANY_ROWS"
            )
        for row_index in range(2, sheet.max_row + 1):
            team_cell = sheet.cell(row_index, 1)
            player_cell = sheet.cell(row_index, 2)
            jersey_cell = sheet.cell(row_index, 3)
            if all(cell.value in (None, "") for cell in (team_cell, player_cell, jersey_cell)):
                continue
            location = f"{title}!A{row_index}"
            source_team = _normalize_text(team_cell.value)
            player_name = _normalize_text(player_cell.value)
            if not source_team:
                issues.append(_issue("TEAM_REQUIRED", "球队标准名称不能为空。", cell=location))
            elif len(source_team) > 120:
                issues.append(
                    _issue("TEAM_NAME_TOO_LONG", "球队标准名称不能超过 120 个字符。", cell=location)
                )
            if not player_name:
                issues.append(
                    _issue(
                        "PLAYER_REQUIRED",
                        "球员姓名不能为空。",
                        cell=f"{title}!B{row_index}",
                    )
                )
            elif len(player_name) > 80:
                issues.append(
                    _issue(
                        "PLAYER_NAME_TOO_LONG",
                        "球员姓名不能超过 80 个字符。",
                        cell=f"{title}!B{row_index}",
                    )
                )
            jersey = _normalize_jersey(jersey_cell.value)
            if jersey and not JERSEY_PATTERN.fullmatch(jersey):
                issues.append(
                    _issue(
                        "INVALID_JERSEY_NUMBER",
                        "球衣号码只允许 0–99 或 00，也可以留空。",
                        cell=f"{title}!C{row_index}",
                        context={"value": jersey},
                    )
                )
            if (
                not source_team
                or not player_name
                or len(source_team) > 120
                or len(player_name) > 80
            ):
                continue
            key = _resolution_key(str(division.id), source_team)
            canonical_name = _normalize_text(resolutions.get(key, source_team))
            if not canonical_name or len(canonical_name) > 120:
                issues.append(
                    _issue(
                        "INVALID_CANONICAL_TEAM_NAME",
                        "统一后的球队标准名称不能为空且不能超过 120 个字符。",
                        cell=location,
                        context={"resolution_key": key},
                    )
                )
                canonical_name = source_team
            resolution_options[key] = {
                "key": key,
                "division_id": str(division.id),
                "division_name": division.name,
                "gender": division.gender,
                "source_name": source_team,
                "canonical_name": canonical_name,
            }
            rows.append(
                {
                    "division_id": str(division.id),
                    "division_code": division.code,
                    "division_name": division.name,
                    "gender": division.gender,
                    "sheet": title,
                    "row": row_index,
                    "source_team_name": source_team,
                    "canonical_team_name": canonical_name,
                    "resolution_key": key,
                    "player_name": player_name,
                    "jersey_number": jersey,
                }
            )
    if not rows:
        issues.append(_issue("NO_ROSTER_ROWS", "工作簿中没有可导入的球队与球员。"))

    teams: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        teams[(str(row["division_id"]), str(row["canonical_team_name"]).casefold())].append(row)
    previews: list[dict[str, object]] = []
    for (_division_id, _name), team_rows in teams.items():
        first = team_rows[0]
        player_seen: dict[str, dict[str, object]] = {}
        jersey_seen: dict[str, dict[str, object]] = {}
        for row in team_rows:
            player_key = str(row["player_name"]).casefold()
            if player_key in player_seen:
                issues.append(
                    _issue(
                        "DUPLICATE_PLAYER_NAME",
                        f"{row['canonical_team_name']}存在重复球员姓名“{row['player_name']}”。",
                        cell=f"{row['sheet']}!B{row['row']}",
                        context={"first_row": player_seen[player_key]["row"]},
                    )
                )
            else:
                player_seen[player_key] = row
            jersey = str(row["jersey_number"])
            if jersey and jersey in jersey_seen:
                issues.append(
                    _issue(
                        "DUPLICATE_JERSEY_NUMBER",
                        f"{row['canonical_team_name']}有多名球员使用 {jersey} 号。",
                        cell=f"{row['sheet']}!C{row['row']}",
                        context={"first_row": jersey_seen[jersey]["row"], "jersey_number": jersey},
                    )
                )
            elif jersey:
                jersey_seen[jersey] = row
        previews.append(
            {
                "division_id": first["division_id"],
                "division_code": first["division_code"],
                "division_name": first["division_name"],
                "gender": first["gender"],
                "name": first["canonical_team_name"],
                "source_names": sorted({str(item["source_team_name"]) for item in team_rows}),
                "player_count": len(team_rows),
                "players": [
                    {
                        "name": item["player_name"],
                        "jersey_number": item["jersey_number"],
                        "source_row": item["row"],
                    }
                    for item in team_rows
                ],
            }
        )

    by_gender_name: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    for item in previews:
        by_gender_name[(str(item["gender"]), str(item["name"]).casefold())].append(item)
    for (_gender, _name), matches in by_gender_name.items():
        division_ids = {str(item["division_id"]) for item in matches}
        if len(division_ids) > 1:
            issues.append(
                _issue(
                    "TEAM_DUPLICATE_ACROSS_DIVISIONS",
                    f"同一性别的多个组别重复出现标准球队“{matches[0]['name']}”。",
                    context={
                        "team_name": matches[0]["name"],
                        "divisions": [item["division_name"] for item in matches],
                    },
                )
            )

    all_options = list(resolution_options.values())
    near_pairs: set[tuple[str, str]] = set()
    for index, left in enumerate(all_options):
        for right in all_options[index + 1 :]:
            if left["canonical_name"].casefold() == right["canonical_name"].casefold():
                continue
            left_key = _name_key(str(left["canonical_name"]))
            right_key = _name_key(str(right["canonical_name"]))
            if not left_key or not right_key:
                continue
            ratio = SequenceMatcher(None, left_key, right_key).ratio()
            containment = (
                min(len(left_key), len(right_key)) >= 2
                and abs(len(left_key) - len(right_key)) <= 4
                and (left_key in right_key or right_key in left_key)
            )
            if left_key != right_key and ratio < 0.78 and not containment:
                continue
            pair = tuple(sorted((str(left["key"]), str(right["key"]))))
            if pair in near_pairs:
                continue
            near_pairs.add(pair)
            issues.append(
                _issue(
                    "SIMILAR_TEAM_NAMES",
                    f"“{left['canonical_name']}”与“{right['canonical_name']}”可能是同一球队。",
                    severity=RosterImportIssue.Severity.WARNING,
                    context={
                        "left_key": left["key"],
                        "right_key": right["key"],
                        "left_name": left["canonical_name"],
                        "right_name": right["canonical_name"],
                        "similarity": round(ratio, 3),
                    },
                )
            )

    division_stats = []
    families = ScheduleSlotFamily.objects.filter(
        season=season, stage__in=[Game.Stage.GROUP, Game.Stage.ROUND_ROBIN]
    )
    expected_by_division: dict[str, int] = defaultdict(int)
    for family in families:
        expected_by_division[str(family.division_id)] += family.slot_count
    for division in divisions:
        team_count = sum(1 for item in previews if item["division_id"] == str(division.id))
        player_count = sum(
            int(item["player_count"])
            for item in previews
            if item["division_id"] == str(division.id)
        )
        expected = expected_by_division.get(str(division.id), 0)
        mismatch = bool(expected and expected != team_count)
        if mismatch:
            issues.append(
                _issue(
                    "SLOT_FAMILY_TEAM_COUNT_MISMATCH",
                    f"{division.name}名单含 {team_count} 支球队，签位方案配置为 {expected} 支。"
                    "名单可以先建立，随后请前往“赛季与组别”修正签位方案。",
                    severity=RosterImportIssue.Severity.WARNING,
                    context={
                        "division_id": str(division.id),
                        "team_count": team_count,
                        "expected_team_count": expected,
                    },
                )
            )
        division_stats.append(
            {
                "division_id": str(division.id),
                "division_code": division.code,
                "division_name": division.name,
                "gender": division.gender,
                "team_count": team_count,
                "player_count": player_count,
                "expected_team_count": expected or None,
                "slot_count_mismatch": mismatch,
            }
        )
    error_count = sum(item.severity == RosterImportIssue.Severity.ERROR for item in issues)
    warning_count = sum(item.severity == RosterImportIssue.Severity.WARNING for item in issues)
    summary = {
        "team_count": len(previews),
        "player_count": len(rows),
        "division_stats": division_stats,
        "teams": sorted(
            previews,
            key=lambda item: (
                next(
                    division.sort_order
                    for division in divisions
                    if str(division.id) == item["division_id"]
                ),
                str(item["name"]),
            ),
        ),
        "name_resolutions": sorted(
            resolution_options.values(),
            key=lambda item: (str(item["division_name"]), str(item["source_name"])),
        ),
        "resolutions": {key: value["canonical_name"] for key, value in resolution_options.items()},
        "error_count": error_count,
        "warning_count": warning_count,
    }
    return issues, summary


def validate_roster_upload(
    *, actor: Account, season: Season, content: bytes, source_name: str
) -> RosterImportBatch:
    _require_superadmin(actor)
    digest = hashlib.sha256(content).hexdigest()
    safe_name = get_valid_filename(PurePath(source_name).name) or "roster.xlsx"
    batch = RosterImportBatch.objects.create(
        season=season,
        template_version=TEMPLATE_VERSION,
        file_key="",
        file_sha256=digest,
        base_season_version=season.version,
        uploaded_by=actor,
    )
    key = f"roster-imports/{season.id}/{batch.id}/{safe_name}"
    batch.file_key = default_storage.save(key, ContentFile(content))
    try:
        _preflight_xlsx(content, source_name)
        issues, summary = _parse_workbook(season, content)
    except RosterManagementError as error:
        issues = [_issue(error.code, str(error))]
        summary = {
            "team_count": 0,
            "player_count": 0,
            "division_stats": [],
            "teams": [],
            "name_resolutions": [],
            "resolutions": {},
            "error_count": 1,
            "warning_count": 0,
        }
    RosterImportIssue.objects.bulk_create(
        [RosterImportIssue(batch=batch, **item.model_kwargs()) for item in issues]
    )
    batch.status = RosterImportBatch.Status.VALIDATED
    batch.summary = summary
    batch.save(update_fields=["file_key", "status", "summary", "updated_at"])
    return batch


def resolve_roster_import(
    *, actor: Account, batch_id: object, resolutions: dict[str, str]
) -> RosterImportBatch:
    _require_superadmin(actor)
    with transaction.atomic():
        batch = (
            RosterImportBatch.objects.select_for_update().select_related("season").get(id=batch_id)
        )
        if batch.status != RosterImportBatch.Status.VALIDATED:
            raise RosterManagementError("只有已校验批次可以处理球队名称。", "BATCH_NOT_VALIDATED")
        allowed = {item["key"] for item in batch.summary.get("name_resolutions", [])}
        unknown = set(resolutions) - allowed
        if unknown:
            raise RosterManagementError(
                "名称处理包含未知记录，请刷新后重试。", "INVALID_RESOLUTION"
            )
        clean = {
            key: _validate_name(value, "球队标准名称", 120) for key, value in resolutions.items()
        }
        with default_storage.open(batch.file_key, "rb") as source:
            content = source.read()
        if hashlib.sha256(content).hexdigest() != batch.file_sha256:
            raise RosterManagementError("上传原文件校验值不一致。", "FILE_HASH_MISMATCH")
        merged = dict(batch.summary.get("resolutions", {}))
        merged.update(clean)
        issues, summary = _parse_workbook(batch.season, content, resolutions=merged)
        batch.issues.all().delete()
        RosterImportIssue.objects.bulk_create(
            [RosterImportIssue(batch=batch, **item.model_kwargs()) for item in issues]
        )
        batch.summary = summary
        batch.save(update_fields=["summary", "updated_at"])
        return batch


def _initial_import_blockers(
    season: Season, *, exclude_batch_id: object | None = None
) -> list[dict]:
    blockers: list[dict] = []
    confirmed = season.roster_imports.filter(status=RosterImportBatch.Status.CONFIRMED)
    if exclude_batch_id:
        confirmed = confirmed.exclude(id=exclude_batch_id)
    checks = [
        ("ROSTER_ALREADY_CONFIRMED", "本赛季已经确认过名单，不能重新导入。", confirmed.count()),
        (
            "GROUPS_EXIST",
            "本赛季已有实际小组数据。",
            CompetitionGroup.objects.filter(division__season=season).count(),
        ),
        (
            "SLOTS_EXIST",
            "本赛季已有实际签位数据。",
            ParticipantSlot.objects.filter(division__season=season).count(),
        ),
        (
            "DRAW_EXISTS",
            "本赛季已有抽签映射。",
            DrawAssignment.objects.filter(season=season).count(),
        ),
        ("GAMES_EXIST", "本赛季已有比赛。", Game.objects.filter(season=season).count()),
        (
            "LEADER_BINDINGS_EXIST",
            "本赛季已有领队绑定。",
            SeasonLeaderBinding.objects.filter(season=season).count(),
        ),
        (
            "RESCHEDULE_EXISTS",
            "本赛季已有调赛记录。",
            RescheduleRequest.objects.filter(game__season=season).count(),
        ),
        (
            "TEAM_CONFIRMATIONS_EXIST",
            "本赛季已有球队确认记录。",
            TeamConfirmation.objects.filter(team__season=season).count(),
        ),
    ]
    for code, message, count in checks:
        if count:
            blockers.append({"code": code, "message": message, "count": count})
    return blockers


def roster_import_readiness(season: Season) -> dict[str, object]:
    blockers = _initial_import_blockers(season)
    if season.status != Season.Status.SETUP:
        blockers.insert(
            0,
            {
                "code": "SEASON_NOT_SETUP",
                "message": "只有准备中的赛季可以执行首次名单导入。",
                "count": 1,
            },
        )
    if not season.divisions.exists():
        blockers.append({"code": "NO_DIVISIONS", "message": "请先创建赛季组别。", "count": 1})
    return {
        "season_id": str(season.id),
        "season_version": season.version,
        "ready": not blockers,
        "blockers": blockers,
    }


def _json_snapshot(value: object) -> object:
    return json.loads(json.dumps(value, cls=DjangoJSONEncoder))


def _team_snapshot(team: Team) -> dict[str, object]:
    return {
        "id": str(team.id),
        "season_id": str(team.season_id),
        "division_id": str(team.division_id),
        "name": team.name,
        "active": team.active,
        "version": team.version,
        "players": [
            {
                "id": str(player.id),
                "name": player.name,
                "jersey_number": player.jersey_number,
                "eligible": player.eligible,
                "active": player.active,
                "version": player.version,
            }
            for player in team.roster.order_by("created_at")
        ],
    }


def confirm_roster_import(
    *,
    actor: Account,
    batch_id: object,
    expected_season_version: int,
    warnings_acknowledged: bool,
) -> RosterImportBatch:
    _require_superadmin(actor)
    try:
        with transaction.atomic():
            batch = (
                RosterImportBatch.objects.select_for_update()
                .select_related("season")
                .get(id=batch_id)
            )
            if batch.status == RosterImportBatch.Status.CONFIRMED:
                return batch
            if batch.status != RosterImportBatch.Status.VALIDATED:
                raise RosterManagementError("导入批次尚未完成校验。", "BATCH_NOT_VALIDATED")
            first_error = (
                batch.issues.filter(severity=RosterImportIssue.Severity.ERROR)
                .order_by("created_at")
                .first()
            )
            if first_error:
                raise RosterManagementError(first_error.message, first_error.code)
            if (
                batch.issues.filter(severity=RosterImportIssue.Severity.WARNING).exists()
                and not warnings_acknowledged
            ):
                raise RosterManagementError("请先确认已阅读全部警告。", "WARNINGS_NOT_ACKNOWLEDGED")
            season = Season.objects.select_for_update().get(id=batch.season_id)
            if (
                season.version != expected_season_version
                or season.version != batch.base_season_version
            ):
                raise RosterManagementError(
                    "赛季已被其他操作修改，请重新上传名单。", "VERSION_CONFLICT"
                )
            if season.status != Season.Status.SETUP:
                raise RosterManagementError(
                    "只有准备中的赛季可以确认名单导入。", "SEASON_NOT_SETUP"
                )
            blockers = _initial_import_blockers(season, exclude_batch_id=batch.id)
            if blockers:
                raise RosterManagementError(blockers[0]["message"], blockers[0]["code"])
            with default_storage.open(batch.file_key, "rb") as source:
                content = source.read()
            if hashlib.sha256(content).hexdigest() != batch.file_sha256:
                raise RosterManagementError("上传原文件校验值不一致。", "FILE_HASH_MISMATCH")
            issues, summary = _parse_workbook(
                season,
                content,
                resolutions=dict(batch.summary.get("resolutions", {})),
            )
            errors = [item for item in issues if item.severity == RosterImportIssue.Severity.ERROR]
            if errors:
                raise RosterManagementError(
                    f"确认前复核失败：{errors[0].message}", "REVALIDATION_FAILED"
                )
            list(Team.objects.select_for_update().filter(season=season))
            list(RosterPlayer.objects.select_for_update().filter(team__season=season))
            before = {
                "season_version": season.version,
                "teams": [
                    _team_snapshot(team)
                    for team in Team.objects.filter(season=season).prefetch_related("roster")
                ],
            }
            RosterPlayer.objects.filter(team__season=season).delete()
            Team.objects.filter(season=season).delete()
            divisions = {str(item.id): item for item in season.divisions.all()}
            created_teams: list[Team] = []
            created_players: list[RosterPlayer] = []
            for item in summary["teams"]:
                team = Team(
                    season=season,
                    division=divisions[str(item["division_id"])],
                    created_by_roster_import_batch=batch,
                    name=str(item["name"]),
                )
                team.full_clean()
                team.save()
                created_teams.append(team)
                for player_item in item["players"]:
                    player = RosterPlayer(
                        team=team,
                        created_by_roster_import_batch=batch,
                        name=str(player_item["name"]),
                        jersey_number=str(player_item["jersey_number"]),
                    )
                    player.full_clean()
                    player.save()
                    created_players.append(player)
            season.version += 1
            season.save(update_fields=["version", "updated_at"])
            summary["confirmed_season_version"] = season.version
            summary["created_team_ids"] = [str(item.id) for item in created_teams]
            summary["created_player_ids"] = [str(item.id) for item in created_players]
            batch.status = RosterImportBatch.Status.CONFIRMED
            batch.summary = summary
            batch.confirmed_by = actor
            batch.confirmed_at = timezone.now()
            batch.save(
                update_fields=[
                    "status",
                    "summary",
                    "confirmed_by",
                    "confirmed_at",
                    "updated_at",
                ]
            )
            after = {
                "season_version": season.version,
                "teams": [_team_snapshot(team) for team in created_teams],
                "team_count": len(created_teams),
                "player_count": len(created_players),
            }
            AdminAuditLog.objects.create(
                actor=actor,
                action="roster.import.confirm",
                object_type="RosterImportBatch",
                object_id=batch.id,
                before=_json_snapshot(before),
                after=_json_snapshot(after),
                metadata={
                    "file_sha256": batch.file_sha256,
                    "template_version": TEMPLATE_VERSION,
                    "warnings_acknowledged": warnings_acknowledged,
                },
            )
            return batch
    except ProtectedError as error:
        raise RosterManagementError(
            "现有球队已被下游数据引用，整次名单确认已回滚。", "PROTECTED_TEAM_REFERENCE"
        ) from error
    except IntegrityError as error:
        raise RosterManagementError(
            "名单与并发数据发生冲突，整次确认已回滚。", "ROSTER_INTEGRITY_CONFLICT"
        ) from error
    except ValidationError as error:
        raise RosterManagementError(
            "名单未通过权威数据校验，整次确认已回滚。", "ROSTER_VALIDATION_FAILED"
        ) from error


def serialize_roster_dataset(season: Season) -> dict[str, object]:
    teams = list(
        Team.objects.filter(season=season)
        .select_related("division")
        .prefetch_related("roster")
        .order_by("division__sort_order", "name")
    )
    confirmed = (
        season.roster_imports.filter(status=RosterImportBatch.Status.CONFIRMED)
        .order_by("-confirmed_at")
        .first()
    )
    readiness = roster_import_readiness(season)
    return {
        "season_id": str(season.id),
        "season_name": season.name,
        "season_status": season.status,
        "season_version": season.version,
        "read_only": season.status == Season.Status.ARCHIVED,
        "team_count": len(teams),
        "active_team_count": sum(item.active for item in teams),
        "player_count": sum(item.roster.count() for item in teams),
        "active_player_count": sum(player.active for team in teams for player in team.roster.all()),
        "divisions": [
            {
                "id": str(item.id),
                "code": item.code,
                "name": item.name,
                "gender": item.gender,
                "sort_order": item.sort_order,
            }
            for item in season.divisions.order_by("sort_order", "name")
        ],
        "teams": [_team_snapshot(item) for item in teams],
        "import_state": {
            "allowed": bool(readiness["ready"]),
            "blockers": readiness["blockers"],
            "confirmed_batch_id": str(confirmed.id) if confirmed else None,
            "confirmed_at": confirmed.confirmed_at.isoformat() if confirmed else None,
        },
    }


def _validate_player_rows(rows: list[dict]) -> list[dict]:
    normalized = []
    names: set[str] = set()
    active_jerseys: set[str] = set()
    for row in rows:
        name = _validate_name(row.get("name"), "球员姓名", 80)
        key = name.casefold()
        if key in names:
            raise RosterManagementError("同一球队内球员姓名不能重复。", "DUPLICATE_PLAYER_NAME")
        names.add(key)
        jersey_number = _validate_jersey(row.get("jersey_number"))
        active = bool(row.get("active", True))
        if active and jersey_number:
            if jersey_number in active_jerseys:
                raise RosterManagementError(
                    f"同一球队内不能有多名启用球员使用 {jersey_number} 号。",
                    "DUPLICATE_JERSEY_NUMBER",
                )
            active_jerseys.add(jersey_number)
        normalized.append(
            {
                "id": row.get("id"),
                "expected_version": row.get("expected_version"),
                "name": name,
                "jersey_number": jersey_number,
                "eligible": bool(row.get("eligible", True)),
                "active": active,
            }
        )
    return normalized


def _validate_team_name_unique(
    *, season: Season, division: Division, name: str, exclude_team_id: object | None = None
) -> None:
    same_division = Team.objects.filter(division=division, name__iexact=name)
    if exclude_team_id:
        same_division = same_division.exclude(id=exclude_team_id)
    if same_division.exists():
        raise RosterManagementError("该组别已存在同名球队。", "DUPLICATE_TEAM_NAME")
    cross_division = Team.objects.filter(
        season=season, division__gender=division.gender, name__iexact=name
    ).exclude(division=division)
    if exclude_team_id:
        cross_division = cross_division.exclude(id=exclude_team_id)
    if cross_division.exists():
        raise RosterManagementError(
            "同一性别的其他组别已使用该标准球队名称。", "TEAM_DUPLICATE_ACROSS_DIVISIONS"
        )


def create_team_with_roster(
    *,
    actor: Account,
    season: Season,
    division_id: object,
    name: str,
    players: list[dict],
    expected_season_version: int,
) -> Team:
    _require_superadmin(actor)
    if season.status == Season.Status.ARCHIVED:
        raise RosterManagementError("已归档赛季只读。", "SEASON_ARCHIVED")
    normalized_name = _validate_name(name, "球队标准名称", 120)
    normalized_players = _validate_player_rows(players)
    with transaction.atomic():
        locked_season = Season.objects.select_for_update().get(id=season.id)
        if locked_season.status == Season.Status.ARCHIVED:
            raise RosterManagementError("已归档赛季只读。", "SEASON_ARCHIVED")
        if locked_season.version != expected_season_version:
            raise RosterManagementError("赛季已被其他操作修改，请刷新。", "VERSION_CONFLICT")
        division = Division.objects.get(id=division_id, season=locked_season)
        _validate_team_name_unique(season=locked_season, division=division, name=normalized_name)
        team = Team.objects.create(
            season=locked_season, division=division, name=normalized_name
        )
        for row in normalized_players:
            RosterPlayer.objects.create(
                team=team,
                name=row["name"],
                jersey_number=row["jersey_number"],
                eligible=row["eligible"],
                active=row["active"],
            )
        locked_season.version += 1
        locked_season.save(update_fields=["version", "updated_at"])
        after = _team_snapshot(team)
        AdminAuditLog.objects.create(
            actor=actor,
            action="roster.team.create",
            object_type="Team",
            object_id=team.id,
            before={},
            after=_json_snapshot(after),
            metadata={"season_version": locked_season.version},
        )
        return team


def _team_reference_counts(team: Team) -> dict[str, int]:
    return {
        "games": Game.objects.filter(Q(home_team=team) | Q(away_team=team)).count(),
        "draw_assignments": team.draw_assignments.count(),
        "leader_bindings": team.leader_bindings.count(),
        "reschedule_requests": team.reschedule_requests.count(),
        "confirmations": team.confirmations.count(),
    }


def _change_preview(team: Team, payload: dict) -> dict[str, object]:
    incoming = _validate_player_rows(payload.get("players", []))
    existing = {str(item.id): item for item in team.roster.all()}
    changed_players = []
    incoming_ids = set()
    added_players = []
    for row in incoming:
        if row["id"]:
            player = existing.get(str(row["id"]))
            if player is None:
                raise RosterManagementError("名单包含不属于该球队的球员。", "INVALID_PLAYER_ID")
            incoming_ids.add(str(player.id))
            changed = any(
                [
                    player.name != row["name"],
                    player.jersey_number != row["jersey_number"],
                    player.eligible != row["eligible"],
                    player.active != row["active"],
                ]
            )
            if changed:
                changed_players.append(str(player.id))
        else:
            added_players.append(row["name"])
    omitted = [str(item.id) for key, item in existing.items() if key not in incoming_ids]
    normalized_name = _validate_name(payload.get("name"), "球队标准名称", 120)
    active = bool(payload.get("active", True))
    dangerous = bool(
        normalized_name != team.name or active != team.active or changed_players or omitted
    )
    canonical = {
        "team_id": str(team.id),
        "team_version": team.version,
        "name": normalized_name,
        "active": active,
        "target_players": [
            {
                "id": row["id"],
                "expected_version": row.get("expected_version"),
                "name": row["name"],
                "jersey_number": row["jersey_number"],
                "eligible": row["eligible"],
                "active": row["active"],
            }
            for row in incoming
        ],
        "changed_player_ids": sorted(changed_players),
        "omitted_player_ids": sorted(omitted),
        "added_player_names": sorted(added_players),
    }
    token = hashlib.sha256(
        json.dumps(canonical, ensure_ascii=False, sort_keys=True).encode()
    ).hexdigest()
    return {
        "team_id": str(team.id),
        "requires_confirmation": team.season.status != Season.Status.SETUP and dangerous,
        "maintenance_token": token,
        "changes": canonical,
        "references": _team_reference_counts(team),
        "message": (
            "该修改会影响已开始赛季中的球队或现有球员，请二次确认。"
            if team.season.status != Season.Status.SETUP and dangerous
            else "可以直接保存。"
        ),
    }


def preview_team_change(*, actor: Account, team: Team, payload: dict) -> dict[str, object]:
    _require_superadmin(actor)
    team = (
        Team.objects.select_related("season", "division").prefetch_related("roster").get(id=team.id)
    )
    if team.season.status == Season.Status.ARCHIVED:
        raise RosterManagementError("已归档赛季只读。", "SEASON_ARCHIVED")
    return _change_preview(team, payload)


def save_team_roster(
    *, actor: Account, team_id: object, payload: dict, maintenance_token: str = ""
) -> Team:
    _require_superadmin(actor)
    try:
        with transaction.atomic():
            team = (
                Team.objects.select_for_update()
                .select_related("season", "division")
                .get(id=team_id)
            )
            if team.season.status == Season.Status.ARCHIVED:
                raise RosterManagementError("已归档赛季只读。", "SEASON_ARCHIVED")
            if team.version != int(payload.get("expected_team_version", 0)):
                raise RosterManagementError("球队已被其他操作修改，请刷新。", "VERSION_CONFLICT")
            list(RosterPlayer.objects.select_for_update().filter(team=team))
            preview = _change_preview(team, payload)
            if (
                preview["requires_confirmation"]
                and maintenance_token != preview["maintenance_token"]
            ):
                raise RosterManagementError(
                    "维护预览已变化，请重新预览并二次确认。", "MAINTENANCE_CONFIRMATION_REQUIRED"
                )
            normalized_name = str(preview["changes"]["name"])
            _validate_team_name_unique(
                season=team.season,
                division=team.division,
                name=normalized_name,
                exclude_team_id=team.id,
            )
            new_active = bool(payload.get("active", True))
            if team.active and not new_active:
                references = _team_reference_counts(team)
                if any(references.values()):
                    raise RosterManagementError(
                        "球队仍被比赛、抽签、领队或调赛数据引用，不能停用。",
                        "TEAM_DEACTIVATION_PROTECTED",
                    )
            before = _team_snapshot(team)
            normalized_players = _validate_player_rows(payload.get("players", []))
            existing = {str(item.id): item for item in team.roster.all()}
            incoming_ids: set[str] = set()
            for row in normalized_players:
                if row["id"]:
                    player = existing.get(str(row["id"]))
                    if player is None:
                        raise RosterManagementError(
                            "名单包含不属于该球队的球员。", "INVALID_PLAYER_ID"
                        )
                    if player.version != int(row.get("expected_version") or 0):
                        raise RosterManagementError(
                            f"球员“{player.name}”已被其他操作修改，请刷新。", "VERSION_CONFLICT"
                        )
                    incoming_ids.add(str(player.id))
                    changed = any(
                        [
                            player.name != row["name"],
                            player.jersey_number != row["jersey_number"],
                            player.eligible != row["eligible"],
                            player.active != row["active"],
                        ]
                    )
                    if changed:
                        player.name = row["name"]
                        player.jersey_number = row["jersey_number"]
                        player.eligible = row["eligible"]
                        player.active = row["active"]
                        player.version += 1
                        player.save(
                            update_fields=[
                                "name",
                                "jersey_number",
                                "eligible",
                                "active",
                                "version",
                                "updated_at",
                            ]
                        )
                else:
                    RosterPlayer.objects.create(
                        team=team,
                        name=row["name"],
                        jersey_number=row["jersey_number"],
                        eligible=row["eligible"],
                        active=row["active"],
                    )
            for player_id, player in existing.items():
                if player_id not in incoming_ids and player.active:
                    player.active = False
                    player.version += 1
                    player.save(update_fields=["active", "version", "updated_at"])
            team.name = normalized_name
            team.active = new_active
            team.version += 1
            team.save(update_fields=["name", "active", "version", "updated_at"])
            after = _team_snapshot(team)
            AdminAuditLog.objects.create(
                actor=actor,
                action="roster.team.save",
                object_type="Team",
                object_id=team.id,
                before=_json_snapshot(before),
                after=_json_snapshot(after),
                metadata={
                    "maintenance_confirmed": bool(preview["requires_confirmation"]),
                    "maintenance_token": preview["maintenance_token"],
                },
            )
            return team
    except IntegrityError as error:
        raise RosterManagementError(
            "球队名称或名单发生并发冲突。", "ROSTER_INTEGRITY_CONFLICT"
        ) from error
