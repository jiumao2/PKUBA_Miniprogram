from __future__ import annotations

import hashlib
from datetime import date, timedelta
from io import BytesIO
from itertools import combinations
from pathlib import PurePath
from uuid import UUID, uuid4

from django.db import transaction
from django.db.models import Max, Min
from django.utils.text import get_valid_filename
from openpyxl import load_workbook

from core.models import (
    Account,
    AdminAuditLog,
    Game,
    Period,
    ScheduleGridDraft,
    ScheduleGridDraftCell,
    ScheduleGridDraftColumn,
    ScheduleImportBatch,
    ScheduleSlotFamily,
    Season,
)
from core.services.schedule_imports_v3 import (
    EXPECTED_SHEETS,
    GRID_START_COLUMN,
    GRID_START_ROW,
    MATCHUP_PATTERN,
    MAX_GRID_COLUMNS,
    TEMPLATE_VERSION,
    WEEKDAY_LABELS,
    WOMEN_SUFFIX_PATTERN,
    ParsedGridColumn,
    ScheduleImportError,
    _default_grid_columns,
    _game_identity,
    _grid_dates_from_sheet,
    _normalize_venue_name,
    _parse_grid_columns,
    _preflight_xlsx,
    _scan_formulas,
    generate_schedule_template,
    schedule_import_readiness,
    validate_schedule_upload,
)


def _require_superadmin(actor: Account) -> None:
    if not actor.is_pkuba_superadmin:
        raise ScheduleImportError("只有超级管理员可以编排赛程。", "PERMISSION_DENIED")


def _assert_season_mutable(season: Season) -> None:
    if season.status == Season.Status.ARCHIVED:
        raise ScheduleImportError("已归档赛季只读。", "SEASON_ARCHIVED")


def _column_specs(draft: ScheduleGridDraft) -> list[ParsedGridColumn]:
    return [
        ParsedGridColumn(
            period_id=str(item.period_id),
            period_code=item.period.code,
            period_name=item.period.name,
            start_time=item.period.start_time.strftime("%H:%M"),
            venue_name=item.venue_name,
            final_only=item.final_only,
            sort_order=item.sort_order,
            source_column=GRID_START_COLUMN + item.sort_order - 1,
        )
        for item in draft.columns.select_related("period").order_by("sort_order")
    ]


def _draft_cells_for_workbook(
    draft: ScheduleGridDraft,
) -> tuple[dict[tuple[date, int], tuple[str, bool]], dict[str, bool]]:
    cells: dict[tuple[date, int], tuple[str, bool]] = {}
    leader_adjustable_by_cell: dict[str, bool] = {}
    date_rows = {
        target_date: index
        for index, target_date in enumerate(_draft_calendar_dates(draft), start=GRID_START_ROW)
    }
    for item in draft.cells.select_related("column").all():
        cells[(item.date, item.column.sort_order)] = (
            item.matchup,
            item.leader_adjustable,
        )
        row = date_rows[item.date]
        column = item.column.sort_order + GRID_START_COLUMN - 1
        from openpyxl.utils import get_column_letter

        leader_adjustable_by_cell[
            f"赛程网格!{get_column_letter(column)}{row}"
        ] = item.leader_adjustable
    return cells, leader_adjustable_by_cell


def _draft_calendar_dates(draft: ScheduleGridDraft) -> list[date]:
    bounds = draft.cells.aggregate(first=Min("date"), last=Max("date"))
    start = min(
        item for item in (draft.season.starts_on, bounds["first"]) if item is not None
    )
    end = max(
        item for item in (draft.season.ends_on, bounds["last"]) if item is not None
    )
    return [start + timedelta(days=offset) for offset in range((end - start).days + 1)]


def _create_default_columns(draft: ScheduleGridDraft) -> None:
    specs = _default_grid_columns(draft.season)
    if not specs:
        raise ScheduleImportError(
            "当前赛季没有可用于新草稿的默认时段。", "NO_DEFAULT_GRID_COLUMNS"
        )
    ScheduleGridDraftColumn.objects.bulk_create(
        [
            ScheduleGridDraftColumn(
                draft=draft,
                period_id=spec.period_id,
                venue_name=spec.venue_name,
                final_only=spec.final_only,
                sort_order=spec.sort_order,
            )
            for spec in specs
        ]
    )


@transaction.atomic
def get_or_create_schedule_draft(
    *, actor: Account, season: Season
) -> ScheduleGridDraft:
    _require_superadmin(actor)
    if season.status == Season.Status.ARCHIVED:
        draft = ScheduleGridDraft.objects.filter(season=season).first()
        if draft is None:
            raise ScheduleImportError(
                "已归档赛季没有保存过赛程草稿，且不能新建草稿。",
                "SEASON_ARCHIVED",
            )
        return draft
    draft, created = ScheduleGridDraft.objects.get_or_create(
        season=season,
        defaults={"updated_by": actor},
    )
    draft = ScheduleGridDraft.objects.select_for_update().get(id=draft.id)
    if not draft.columns.exists():
        _create_default_columns(draft)
        created = True
    if created:
        AdminAuditLog.objects.create(
            actor=actor,
            action="SCHEDULE_GRID_DRAFT_CREATED",
            object_type="ScheduleGridDraft",
            object_id=draft.id,
            before={},
            after={
                "season_id": str(season.id),
                "version": draft.version,
                "column_count": draft.columns.count(),
            },
            metadata={"template_version": TEMPLATE_VERSION},
        )
    return draft


def _normalized_matchup(value: str) -> tuple[str, str] | None:
    match = MATCHUP_PATTERN.fullmatch(value.strip())
    if match is None:
        return None
    left, right = match.groups()
    suffix = "WOMEN" if WOMEN_SUFFIX_PATTERN.search(value) else "MEN"
    return suffix, "|".join(sorted((left, right)))


def _matchup_pool(season: Season, draft: ScheduleGridDraft) -> list[dict[str, object]]:
    scheduled = {
        normalized
        for value in draft.cells.values_list("matchup", flat=True)
        if (normalized := _normalized_matchup(value)) is not None
    }
    existing_semantics: set[str] = set()
    for game in (
        Game.objects.filter(season=season)
        .exclude(status=Game.Status.VOID)
        .select_related("division", "home_slot", "away_slot")
    ):
        if game.home_slot_id and game.away_slot_id:
            _code, semantic = _game_identity(
                game.division.code,
                game.stage,
                game.home_slot.code,
                game.away_slot.code,
            )
            existing_semantics.add(semantic)

    rows: list[dict[str, object]] = []
    for family in ScheduleSlotFamily.objects.filter(season=season).select_related(
        "division"
    ).order_by("sort_order"):
        codes = [f"{family.prefix}{number}" for number in range(1, family.slot_count + 1)]
        pairs = (
            list(combinations(codes, 2))
            if family.stage in {Game.Stage.GROUP, Game.Stage.ROUND_ROBIN}
            else list(zip(codes[::2], codes[1::2], strict=True))
        )
        for home, away in pairs:
            suffix = "（女）" if family.division.gender == "WOMEN" else ""
            matchup = f"{home}vs{away}{suffix}"
            gender_key = "WOMEN" if suffix else "MEN"
            _code, semantic = _game_identity(
                family.division.code, family.stage, home, away
            )
            rows.append(
                {
                    "key": semantic,
                    "matchup": matchup,
                    "division_code": family.division.code,
                    "division_name": family.division.name,
                    "gender": family.division.gender,
                    "stage": family.stage,
                    "stage_name": family.get_stage_display(),
                    "scheduled": (gender_key, "|".join(sorted((home, away))))
                    in scheduled,
                    "already_formal": semantic in existing_semantics,
                }
            )
    return rows


def serialize_schedule_draft(draft: ScheduleGridDraft) -> dict[str, object]:
    season = draft.season
    columns = list(draft.columns.select_related("period").order_by("sort_order"))
    cells = list(draft.cells.select_related("column").order_by("date", "column__sort_order"))
    readiness = schedule_import_readiness(season)
    return {
        "id": str(draft.id),
        "season_id": str(season.id),
        "season_version": season.version,
        "version": draft.version,
        "template_version": TEMPLATE_VERSION,
        "source_name": draft.source_name,
        "updated_at": draft.updated_at,
        "columns": [
            {
                "id": str(item.id),
                "period_id": str(item.period_id),
                "period_code": item.period.code,
                "period_name": item.period.name,
                "start_time": item.period.start_time.strftime("%H:%M"),
                "venue_name": item.venue_name,
                "final_only": item.final_only,
                "sort_order": item.sort_order,
            }
            for item in columns
        ],
        "cells": [
            {
                "id": str(item.id),
                "column_id": str(item.column_id),
                "date": item.date.isoformat(),
                "matchup": item.matchup,
                "leader_adjustable": item.leader_adjustable,
            }
            for item in cells
        ],
        "dates": [
            {"date": item.isoformat(), "weekday": WEEKDAY_LABELS[item.weekday()]}
            for item in _draft_calendar_dates(draft)
        ],
        "periods": [
            {
                "id": str(item.id),
                "code": item.code,
                "name": item.name,
                "start_time": item.start_time.strftime("%H:%M"),
            }
            for item in Period.objects.filter(season=season).order_by("sort_order")
        ],
        "matchup_pool": _matchup_pool(season, draft),
        "summary": {
            "expected_game_count": readiness["expected_game_count"],
            "draft_game_count": len(cells),
            "locked_game_count": sum(not item.leader_adjustable for item in cells),
            "column_count": len(columns),
            "calendar_day_count": readiness["calendar_day_count"],
        },
    }


@transaction.atomic
def replace_schedule_draft(
    *,
    actor: Account,
    season: Season,
    expected_version: int,
    columns: list[dict[str, object]],
    cells: list[dict[str, object]],
    source_name: str | None = None,
    source_sha256: str | None = None,
) -> ScheduleGridDraft:
    _require_superadmin(actor)
    _assert_season_mutable(season)
    draft = get_or_create_schedule_draft(actor=actor, season=season)
    draft = ScheduleGridDraft.objects.select_for_update().get(id=draft.id)
    if draft.version != expected_version:
        raise ScheduleImportError(
            "赛程草稿已被其他窗口修改，请刷新后重试。", "DRAFT_VERSION_CONFLICT"
        )
    if not 1 <= len(columns) <= MAX_GRID_COLUMNS:
        raise ScheduleImportError(
            f"赛程网格必须包含 1 至 {MAX_GRID_COLUMNS} 个排期列。",
            "INVALID_GRID_COLUMN_COUNT",
        )

    periods = {
        str(item.id): item for item in Period.objects.filter(season=season)
    }
    normalized_columns: list[dict[str, object]] = []
    seen_ids: set[UUID] = set()
    seen_pairs: set[tuple[str, str]] = set()
    for sort_order, raw in enumerate(columns, start=1):
        try:
            column_id = UUID(str(raw.get("id") or uuid4()))
        except ValueError as error:
            raise ScheduleImportError("赛程列标识无效。", "INVALID_DRAFT_COLUMN_ID") from error
        if column_id in seen_ids:
            raise ScheduleImportError("赛程列标识重复。", "DUPLICATE_DRAFT_COLUMN_ID")
        seen_ids.add(column_id)
        period_id = str(raw.get("period_id") or "")
        period = periods.get(period_id)
        if period is None:
            raise ScheduleImportError("赛程列使用了其他赛季或未知时段。", "UNKNOWN_GRID_PERIOD")
        venue_name, marker_final_only = _normalize_venue_name(raw.get("venue_name"))
        if not venue_name or len(venue_name) > 120:
            raise ScheduleImportError("场地名称必须为 1 至 120 个字符。", "INVALID_GRID_VENUE")
        duplicate_key = (period_id, venue_name.casefold())
        if duplicate_key in seen_pairs:
            raise ScheduleImportError(
                f"{period.start_time.strftime('%H:%M')} + {venue_name} 重复。",
                "DUPLICATE_GRID_COLUMN",
            )
        seen_pairs.add(duplicate_key)
        normalized_columns.append(
            {
                "id": column_id,
                "period": period,
                "venue_name": venue_name,
                "final_only": bool(raw.get("final_only")) or marker_final_only,
                "sort_order": sort_order,
            }
        )

    normalized_cells: list[dict[str, object]] = []
    seen_cells: set[tuple[date, UUID]] = set()
    for raw in cells:
        matchup = str(raw.get("matchup") or "").strip()
        if not matchup:
            continue
        if len(matchup) > 64:
            raise ScheduleImportError("单元格对阵不能超过 64 个字符。", "MATCHUP_TOO_LONG")
        try:
            target_date = date.fromisoformat(str(raw.get("date") or ""))
            column_id = UUID(str(raw.get("column_id") or ""))
        except ValueError as error:
            raise ScheduleImportError(
                "草稿单元格日期或列标识无效。", "INVALID_DRAFT_CELL"
            ) from error
        if column_id not in seen_ids:
            raise ScheduleImportError("草稿单元格引用了未知列。", "UNKNOWN_DRAFT_COLUMN")
        key = (target_date, column_id)
        if key in seen_cells:
            raise ScheduleImportError("同一草稿单元格重复提交。", "DUPLICATE_DRAFT_CELL")
        seen_cells.add(key)
        normalized_cells.append(
            {
                "date": target_date,
                "column_id": column_id,
                "matchup": matchup,
                "leader_adjustable": bool(raw.get("leader_adjustable", True)),
            }
        )

    before = {
        "version": draft.version,
        "column_count": draft.columns.count(),
        "cell_count": draft.cells.count(),
    }
    draft.cells.all().delete()
    draft.columns.all().delete()
    ScheduleGridDraftColumn.objects.bulk_create(
        [ScheduleGridDraftColumn(draft=draft, **item) for item in normalized_columns]
    )
    ScheduleGridDraftCell.objects.bulk_create(
        [ScheduleGridDraftCell(draft=draft, **item) for item in normalized_cells]
    )
    draft.version += 1
    draft.updated_by = actor
    if source_name is not None:
        draft.source_name = get_valid_filename(PurePath(source_name).name)[:255]
    if source_sha256 is not None:
        draft.source_sha256 = source_sha256
    draft.save(
        update_fields=[
            "version",
            "updated_by",
            "source_name",
            "source_sha256",
            "updated_at",
        ]
    )
    draft.validation_batches.filter(status=ScheduleImportBatch.Status.VALIDATED).update(
        status=ScheduleImportBatch.Status.REJECTED
    )
    AdminAuditLog.objects.create(
        actor=actor,
        action="SCHEDULE_GRID_DRAFT_UPDATED",
        object_type="ScheduleGridDraft",
        object_id=draft.id,
        before=before,
        after={
            "version": draft.version,
            "column_count": len(normalized_columns),
            "cell_count": len(normalized_cells),
        },
        metadata={"source_name": draft.source_name},
    )
    return draft


def import_schedule_draft_xlsx(
    *,
    actor: Account,
    season: Season,
    expected_version: int,
    content: bytes,
    source_name: str,
) -> ScheduleGridDraft:
    _require_superadmin(actor)
    _assert_season_mutable(season)
    _preflight_xlsx(content)
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as error:
        raise ScheduleImportError("无法读取 XLSX 工作簿。", "INVALID_XLSX") from error
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise ScheduleImportError("工作表名称、数量或顺序不正确。", "SHEET_STRUCTURE_CHANGED")
    if any(workbook[name].sheet_state != "visible" for name in EXPECTED_SHEETS):
        raise ScheduleImportError("三个工作表必须全部可见。", "SHEET_VISIBILITY_INVALID")
    if str(workbook["填写说明"]["B4"].value or "").strip() != TEMPLATE_VERSION:
        raise ScheduleImportError(
            f"只支持当前 {TEMPLATE_VERSION} 格式。", "UNSUPPORTED_TEMPLATE_VERSION"
        )
    formula_issues: list = []
    _scan_formulas(workbook, formula_issues)
    if formula_issues:
        raise ScheduleImportError(formula_issues[0].message, formula_issues[0].code)

    sheet = workbook["赛程网格"]
    specs = _parse_grid_columns(season, sheet)
    expected_dates = _grid_dates_from_sheet(
        sheet,
        end_column=GRID_START_COLUMN + len(specs) - 1,
    )

    column_payload: list[dict[str, object]] = []
    column_ids: dict[int, UUID] = {}
    for spec in specs:
        column_id = uuid4()
        column_ids[spec.source_column] = column_id
        column_payload.append(
            {
                "id": column_id,
                "period_id": spec.period_id,
                "venue_name": spec.venue_name,
                "final_only": spec.final_only,
            }
        )
    cell_payload: list[dict[str, object]] = []
    for row_index, target_date in enumerate(expected_dates, start=GRID_START_ROW):
        for spec in specs:
            value = sheet.cell(row_index, spec.source_column).value
            if value in (None, ""):
                continue
            cell_payload.append(
                {
                    "column_id": column_ids[spec.source_column],
                    "date": target_date,
                    "matchup": str(value).strip(),
                    "leader_adjustable": True,
                }
            )
    return replace_schedule_draft(
        actor=actor,
        season=season,
        expected_version=expected_version,
        columns=column_payload,
        cells=cell_payload,
        source_name=source_name,
        source_sha256=hashlib.sha256(content).hexdigest(),
    )


@transaction.atomic
def export_schedule_draft_xlsx(*, actor: Account, season: Season) -> bytes:
    _require_superadmin(actor)
    draft = get_or_create_schedule_draft(actor=actor, season=season)
    draft = ScheduleGridDraft.objects.select_for_update().get(id=draft.id)
    cells, _leader_map = _draft_cells_for_workbook(draft)
    return generate_schedule_template(
        season,
        columns=_column_specs(draft),
        cells=cells,
        calendar_dates=_draft_calendar_dates(draft),
    )


@transaction.atomic
def validate_schedule_draft(
    *, actor: Account, season: Season, expected_version: int
) -> ScheduleImportBatch:
    _require_superadmin(actor)
    _assert_season_mutable(season)
    draft = get_or_create_schedule_draft(actor=actor, season=season)
    draft = ScheduleGridDraft.objects.select_for_update().get(id=draft.id)
    if draft.version != expected_version:
        raise ScheduleImportError(
            "草稿已发生变化，请等待自动保存完成后重新核对。", "DRAFT_VERSION_CONFLICT"
        )
    cells, leader_map = _draft_cells_for_workbook(draft)
    content = generate_schedule_template(
        season,
        columns=_column_specs(draft),
        cells=cells,
        calendar_dates=_draft_calendar_dates(draft),
    )
    return validate_schedule_upload(
        actor=actor,
        season=season,
        content=content,
        source_name=f"{season.name}-online-draft-v{draft.version}.xlsx",
        source_kind=ScheduleImportBatch.SourceKind.ONLINE_DRAFT,
        source_draft=draft,
        source_draft_version=draft.version,
        leader_adjustable_by_cell=leader_map,
    )
